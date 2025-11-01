from clases.UsuarioSistema import UsuarioSistema
from utils import pedir_input, validar_rut, validar_email, validar_telefono, validar_contrasena
from datetime import datetime
import getpass
import hashlib
from openpyxl import Workbook
import os

if not os.path.exists("./informes"):
    os.makedirs("./informes")


class Admin(UsuarioSistema):

    def __init__(self, Nombre, Apellido, RUT, Email, Telefono, NombreUsuario, Contraseña, id_Direccion, id_Rol, id_UsuarioSistema=None):
        super().__init__(Nombre, Apellido, RUT, Email, Telefono, NombreUsuario, Contraseña, id_Direccion, id_Rol, id_UsuarioSistema)
    
    
    # Creacion del Usuario: Admin/Empleado en sistema 
    # Falta rellenar datos hacia DetalleEmpleado y EmpleadoProyecto
    
    # Contraseñas deben ir encriptadas

    # Falta agregar la opción de cancelar operación para cada una de las funciones

    # --- CREAR USUARIOS: LISTO ---

    def crear_usuario(self, cnx, cursor):
        try:
            nombre = pedir_input("Nombre [0 para cancelar]: ", cnx)
            if nombre is None: return

            apellido = pedir_input("Apellido [0 para cancelar]: ", cnx)
            if apellido is None: return

            while True:
                # validar rut con algoritmo
                rut = pedir_input("RUT [0 para cancelar]: ", cnx)
                if rut is None: 
                    return
                if not rut:
                    print("El Rut no puede estar vacío.")
                    continue
                if validar_rut(rut):
                    break
                else:
                    print("RUT Inválido. Reintente.")

            while True:
                # match con @ y gmail
                email = pedir_input("Email [0 para cancelar]: ", cnx)

                if email is None: 
                    return

                if not email:
                    print("El Email no puede estar vacío.")
                    continue
                if validar_email(email):
                    break
                else:
                    print("Email Inválido. Reintente.")

            # telefonos solo con prefijo +569 o +52, es opcional
            while True:
                telefono = pedir_input("Teléfono (opcional, 0 para cancelar): ", cnx, opcional=True)
                
                if telefono is None: 
                    return
                if not telefono:
                    telefono = None 
                    break
                if validar_telefono(telefono):
                    break
                else:
                    print(f"Teléfono {telefono} inválido. Reintente.")

            while True:
                # max 20 caracteres y unico
                nombre_usuario = pedir_input("Nombre de usuario [0 para cancelar]: ", cnx)
                if nombre_usuario is None: 
                    return
                if len(nombre_usuario) > 20:
                    print("El nombre de usuario no puede superar 20 carácteres.")

                cursor.execute("SELECT COUNT(*) FROM UsuarioSistema WHERE NombreUsuario = %s", (nombre_usuario,))
                if cursor.fetchone()[0] > 0:
                    print("⚠️ El nombre de usuario ya existe. Elige otro.")
                    continue

                break

            # tiene que llegar hasheada a la base de datos
            # y ademas no mostrarse en consola
            while True:
                contrasena = getpass.getpass("Contraseña [0 para cancelar]: ").strip()
                if contrasena.lower() in ("0", "cancelar", "exit"):
                    print("Operación cancelada.")
                    return

                if not validar_contrasena(contrasena):
                    print("La contraseña debe tener mínimo 8 caracteres, una mayúscula, un número y un símbolo.")
                    continue

                # confirmación opcional
                confirmar = getpass.getpass("Confirme su contraseña: ").strip()
                if contrasena != confirmar:
                    print("Las contraseñas no coinciden. Intente nuevamente.")
                    continue

                break

            # Convertir la contraseña a hash
            contrasena_hash = hashlib.sha256(contrasena.encode()).hexdigest()


            """--- RELLENAR DIRECCIONES """

            print("Ingrese la dirección del usuario:")
            calle = pedir_input("Calle [0 para cancelar]: ", cnx)
            if calle is None: return

            while True:
                numero = pedir_input("Número [0 para cancelar]: ", cnx)
                if numero is None: return
                try:
                    numero = int(numero)
                    break
                except ValueError:
                    print("Número inválido. Debe ser un número entero.")

            # Debe solo pedir strings
            ciudad = pedir_input("Ciudad [0 para cancelar]: ", cnx)
            if ciudad is None: return

            # Debe solo pedir strings
            region = pedir_input("Región [0 para cancelar]: ", cnx)
            if region is None: return

            # Debe solo pedir strings
            pais = pedir_input("País [0 para cancelar]: ", cnx)
            if pais is None: return

            while True:
                codigo_postal = pedir_input("Código Postal [0 para cancelar]: ", cnx)
                if codigo_postal is None: 
                    return
                try:
                    codigo_postal = int(codigo_postal)
                    if not (100000 <= codigo_postal <= 9999999):
                        print("Código postal fuera de rango válido.")
                        continue
                    break
                except ValueError:
                    print("Código postal inválido. Debe ser un número entero.")


            cursor.execute("""
                INSERT INTO Direccion (Calle, Numero, Ciudad, Region, Pais, CodigoPostal)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (calle, numero, ciudad, region, pais, codigo_postal))
            cnx.commit()
            id_direccion = cursor.lastrowid

            while True:
                tipo_rol = pedir_input("Tipo de rol (admin/empleado, 0 para cancelar): ", cnx)
                if tipo_rol is None: return
                tipo_rol = tipo_rol.lower()
                if tipo_rol in ("admin", "empleado"):
                    break
                print("Opción inválida. Debe ser 'admin' o 'empleado'.")

            cursor.execute("INSERT INTO Rol (tipo_Rol) VALUES (%s)", (tipo_rol,))
            cnx.commit()
            id_rol = cursor.lastrowid

            cursor.execute("""
                    INSERT INTO UsuarioSistema 
                    (Nombre, Apellido, RUT, Email, Telefono, NombreUsuario, Contraseña, Direccion_idDireccion, Rol_idRol)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (nombre, apellido, rut, email, telefono, nombre_usuario, contrasena_hash, id_direccion, id_rol))
            
            cnx.commit()
            id_usuario = cursor.lastrowid

            while True:
                salario = pedir_input("Salario (en CLP, 0 para cancelar): ", cnx)
                if salario is None:
                    return

                # Evitar valores vacíos (por seguridad extra)
                if not salario:
                    print("El salario no puede estar vacío.")
                    continue

                # Validar formato y rango
                try:
                    salario = int(salario)
                    if salario <= 0:
                        print("El salario debe ser mayor que 0.")
                        continue
                    if salario > 10_000_000:  # ejemplo de tope
                        print("El salario no puede superar los 10 millones de CLP.")
                        continue
                    break
                except ValueError:
                    print("Salario inválido. Debe ser un número entero (sin puntos ni símbolos).")


            tipo_empleado = pedir_input("Tipo de empleado (por defecto 'No especificado', 0 para cancelar): ", cnx, opcional=True)
            if tipo_empleado is None: 
                return
            if not tipo_empleado:
                tipo_empleado = "No especificado"


            while True:
                fecha_contrato = pedir_input("Fecha de contrato (YYYY-MM-DD, dejar vacío para actual, 0 para cancelar): ", cnx, opcional=True)
                if fecha_contrato is None:
                    return

                if not fecha_contrato:
                    fecha_contrato = datetime.today().strftime("%Y-%m-%d")
                    print(f"Se usará la fecha actual: {fecha_contrato}")
                    break

                try:
                    datetime.strptime(fecha_contrato, "%Y-%m-%d")
                    break
                except ValueError:
                    print("⚠️ Formato inválido. Use YYYY-MM-DD, por ejemplo: 2025-03-21.")
                    continue

            # una sola inserción
            cursor.execute("""
                INSERT INTO DetalleEmpleado (FechaContrato, Salario, TipoEmpleado, UsuarioSistema_idUsuarioSistema)
                VALUES (%s, %s, %s, %s)
            """, (fecha_contrato, salario, tipo_empleado, id_usuario))


            cnx.commit()
            print(f"Usuario creado con ID: {id_usuario}, Rol: {tipo_rol}")

        except Exception as e:
            cnx.rollback()
            print(f"Ocurrió un error: {e}")

    def editar_usuario(self, cnx, cursor):
        try:
            print("\n--- EDITAR USUARIO ---")

            # --- LISTADO DE USUARIOS DISPONIBLES ---
            cursor.execute("""
                SELECT U.id_UsuarioSistema, U.Nombre, U.Apellido, U.NombreUsuario, R.tipo_Rol
                FROM UsuarioSistema U
                INNER JOIN Rol R ON U.Rol_idRol = R.id_Rol
                ORDER BY U.id_UsuarioSistema
            """)
            usuarios = cursor.fetchall()

            if not usuarios:
                print("No hay usuarios registrados.")
                return

            print("\n--- LISTADO DE USUARIOS ---")
            for u in usuarios:
                print(f"ID: {u[0]} - {u[1]} {u[2]} (Usuario: {u[3]}, Rol: {u[4]})")

            id_usuario = pedir_input("Ingrese el ID del usuario a editar [0 para cancelar]: ", cnx)
            if id_usuario is None or not id_usuario.isdigit():
                print("Operación cancelada o ID inválido.")
                return
            id_usuario = int(id_usuario)

            # --- DATOS ACTUALES DEL USUARIO ---
            cursor.execute("""
                SELECT U.Nombre, U.Apellido, U.RUT, U.Email, U.Telefono, U.NombreUsuario, U.Contraseña,
                    U.Direccion_idDireccion, U.Rol_idRol,
                    D.Calle, D.Numero, D.Ciudad, D.Region, D.Pais, D.CodigoPostal,
                    R.tipo_Rol,
                    E.id_DetalleEmpleado, E.FechaContrato, E.Salario, E.TipoEmpleado
                FROM UsuarioSistema U
                INNER JOIN Direccion D ON U.Direccion_idDireccion = D.id_Direccion
                INNER JOIN Rol R ON U.Rol_idRol = R.id_Rol
                LEFT JOIN DetalleEmpleado E ON U.id_UsuarioSistema = E.UsuarioSistema_idUsuarioSistema
                WHERE U.id_UsuarioSistema = %s
            """, (id_usuario,))
            usuario = cursor.fetchone()

            if not usuario:
                print("Usuario no encontrado.")
                return

            print("\n--- EDITAR DATOS PERSONALES ---")

            # --- DATOS PERSONALES ---
            nombre = pedir_input(f"Nombre actual [{usuario[0]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
            if nombre is None: return
            nombre = nombre or usuario[0]

            apellido = pedir_input(f"Apellido actual [{usuario[1]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
            if apellido is None: return
            apellido = apellido or usuario[1]

            # --- RUT ---
            while True:
                rut = pedir_input(f"RUT actual [{usuario[2]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
                if rut is None: return
                if not rut:
                    rut = usuario[2]
                    break
                if validar_rut(rut):
                    break
                print("RUT inválido. Reintente.")

            # --- EMAIL ---
            while True:
                email = pedir_input(f"Email actual [{usuario[3]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
                if email is None: return
                if not email:
                    email = usuario[3]
                    break
                if validar_email(email):
                    break
                print("Email inválido. Reintente.")

            # --- TELÉFONO ---
            while True:
                telefono = pedir_input(f"Teléfono actual [{usuario[4]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
                if telefono is None: return
                if not telefono:
                    telefono = usuario[4]
                    break
                if validar_telefono(telefono):
                    break
                print("Teléfono inválido. Reintente.")

            # --- NOMBRE DE USUARIO ---
            while True:
                nombre_usuario = pedir_input(f"Nombre de usuario actual [{usuario[5]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
                if nombre_usuario is None: return
                if not nombre_usuario:
                    nombre_usuario = usuario[5]
                    break
                if len(nombre_usuario) > 20:
                    print("El nombre de usuario no puede superar 20 caracteres.")
                    continue
                cursor.execute("SELECT COUNT(*) FROM UsuarioSistema WHERE NombreUsuario = %s AND id_UsuarioSistema != %s", (nombre_usuario, id_usuario))
                if cursor.fetchone()[0] > 0:
                    print("El nombre de usuario ya existe. Intente con otro.")
                    continue
                break

            # --- CONTRASEÑA ---
            contrasena = getpass.getpass("Nueva contraseña [dejar vacío para mantener, 0 para cancelar]: ").strip()
            if contrasena.lower() in ("0", "cancelar", "exit"):
                print("Operación cancelada.")
                return
            if contrasena:
                if validar_contrasena(contrasena):
                    confirmar = getpass.getpass("Confirme la nueva contraseña: ").strip()
                    if contrasena == confirmar:
                        contrasena = hashlib.sha256(contrasena.encode()).hexdigest()
                    else:
                        print("Las contraseñas no coinciden. Se mantendrá la anterior.")
                        contrasena = usuario[6]
                else:
                    print("Contraseña inválida. Se mantendrá la actual.")
                    contrasena = usuario[6]
            else:
                contrasena = usuario[6]

            # --- DIRECCIÓN ---
            print("\n--- EDITAR DIRECCIÓN ---")
            calle = pedir_input(f"Calle [{usuario[9]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
            if calle is None: return
            calle = calle or usuario[9]

            while True:
                numero = pedir_input(f"Número [{usuario[10]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
                if numero is None: return
                if not numero:
                    numero = usuario[10]
                    break
                try:
                    numero = int(numero)
                    break
                except ValueError:
                    print("Número inválido. Debe ser un número entero.")

            ciudad = pedir_input(f"Ciudad [{usuario[11]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True) or usuario[11]
            region = pedir_input(f"Región [{usuario[12]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True) or usuario[12]
            pais = pedir_input(f"País [{usuario[13]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True) or usuario[13]

            while True:
                codigo_postal = pedir_input(f"Código Postal [{usuario[14]}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
                if codigo_postal is None: return
                if not codigo_postal:
                    codigo_postal = usuario[14]
                    break
                try:
                    codigo_postal = int(codigo_postal)
                    if not (100000 <= codigo_postal <= 9999999):
                        print("Código postal fuera de rango.")
                        continue
                    break
                except ValueError:
                    print("Código postal inválido. Debe ser numérico.")

            # --- ROL ---
            tipo_rol = pedir_input(f"Tipo de rol actual [{usuario[15]}] (admin/empleado) [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
            if tipo_rol is None: return
            tipo_rol = tipo_rol.lower() if tipo_rol else usuario[15]
            if tipo_rol not in ("admin", "empleado"):
                tipo_rol = usuario[15]

            cursor.execute("UPDATE Rol SET tipo_Rol = %s WHERE id_Rol = %s", (tipo_rol, usuario[8]))

            # --- ACTUALIZAR DIRECCIÓN ---
            cursor.execute("""
                UPDATE Direccion
                SET Calle=%s, Numero=%s, Ciudad=%s, Region=%s, Pais=%s, CodigoPostal=%s
                WHERE id_Direccion=%s
            """, (calle, numero, ciudad, region, pais, codigo_postal, usuario[7]))

            # --- ACTUALIZAR USUARIO ---
            cursor.execute("""
                UPDATE UsuarioSistema
                SET Nombre=%s, Apellido=%s, RUT=%s, Email=%s, Telefono=%s, NombreUsuario=%s, Contraseña=%s
                WHERE id_UsuarioSistema=%s
            """, (nombre, apellido, rut, email, telefono, nombre_usuario, contrasena, id_usuario))

            # --- DETALLE EMPLEADO ---
            print("\n--- EDITAR DETALLE DE EMPLEADO ---")
            id_detalle, fecha_actual, salario_actual, tipo_actual = usuario[16:20]

            while True:
                salario = pedir_input(f"Salario actual [{salario_actual}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
                if salario is None: return
                if not salario:
                    salario = salario_actual
                    break
                try:
                    salario = int(salario)
                    if salario <= 0:
                        print("El salario debe ser mayor que 0.")
                        continue
                    break
                except ValueError:
                    print("Salario inválido. Debe ser un número entero.")

            tipo_empleado = pedir_input(f"Tipo de empleado actual [{tipo_actual}] [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
            if tipo_empleado is None: return
            tipo_empleado = tipo_empleado or tipo_actual

            from datetime import datetime
            while True:
                fecha_contrato = pedir_input(f"Fecha de contrato actual [{fecha_actual}] (YYYY-MM-DD) [dejar vacío para mantener, 0 para cancelar]: ", cnx, opcional=True)
                if fecha_contrato is None: return
                if not fecha_contrato:
                    fecha_contrato = fecha_actual
                    break
                try:
                    datetime.strptime(fecha_contrato, "%Y-%m-%d")
                    break
                except ValueError:
                    print("Formato inválido. Use YYYY-MM-DD.")
                    continue

            if id_detalle:
                cursor.execute("""
                    UPDATE DetalleEmpleado
                    SET FechaContrato=%s, Salario=%s, TipoEmpleado=%s
                    WHERE id_DetalleEmpleado=%s
                """, (fecha_contrato, salario, tipo_empleado, id_detalle))
            else:
                cursor.execute("""
                    INSERT INTO DetalleEmpleado (UsuarioSistema_idUsuarioSistema, FechaContrato, Salario, TipoEmpleado)
                    VALUES (%s, %s, %s, %s)
                """, (id_usuario, fecha_contrato, salario, tipo_empleado))

            cnx.commit()
            print("\nUsuario actualizado correctamente.\n")

        except Exception as e:
            try:
                cnx.rollback()
            except:
                pass
            print(f"Ocurrió un error: {e}")

    def eliminar_usuario(self, cnx, cursor, current_user_id):
        try:
            while True:
                usuario_a_eliminar = pedir_input("Ingrese el ID del usuario a eliminar [0 para cancelar]: ", cnx)
                if usuario_a_eliminar is None:
                    return
                if usuario_a_eliminar.isdigit():
                    usuario_a_eliminar = int(usuario_a_eliminar)
                    break
                print("Error: Debe ingresar un número entero válido.")

            cursor.execute("""
                SELECT U.id_UsuarioSistema, U.NombreUsuario, U.Direccion_idDireccion, U.Rol_idRol, R.tipo_Rol
                FROM UsuarioSistema U
                JOIN Rol R ON U.Rol_idRol = R.id_Rol
                WHERE U.id_UsuarioSistema = %s
            """, (usuario_a_eliminar,))
            usuario = cursor.fetchone()

            if not usuario:
                print(f"No se encontró ningún usuario con ID: {usuario_a_eliminar}")
                return

            id_usuario, nombre_usuario, id_direccion, id_rol, tipo_rol = usuario

            if id_usuario == current_user_id:
                print("Acción prohibida: no puedes eliminar el usuario con el que estás autenticado.")
                return

            if tipo_rol.lower() == "admin":
                cursor.execute("""
                    SELECT COUNT(*) FROM UsuarioSistema U
                    JOIN Rol R ON U.Rol_idRol = R.id_Rol
                    WHERE R.tipo_Rol = 'admin'
                """)
                cantidad_admins = cursor.fetchone()[0]
                if cantidad_admins <= 1:
                    print("Acción prohibida: no se puede eliminar al último usuario con rol 'admin'.")
                    return

            print(f"Usuario a eliminar: ID={id_usuario}, NombreUsuario='{nombre_usuario}', Rol='{tipo_rol}'")
            print("ADVERTENCIA: Esta acción es irreversible y eliminará todos los datos asociados.")
            confirmacion = pedir_input("Escriba DELETE (en mayúsculas) para confirmar o 0 para cancelar: ", cnx)
            if confirmacion is None or confirmacion.strip().upper() != "DELETE":
                print("Operación cancelada.")
                cnx.rollback()
                return

            cursor.execute("DELETE FROM DetalleEmpleado WHERE UsuarioSistema_idUsuarioSistema = %s", (id_usuario,))
            cursor.execute("DELETE FROM UsuarioSistema WHERE id_UsuarioSistema = %s", (id_usuario,))

            if id_direccion is not None:
                cursor.execute("SELECT COUNT(*) FROM UsuarioSistema WHERE Direccion_idDireccion = %s", (id_direccion,))
                if cursor.fetchone()[0] == 0:
                    cursor.execute("DELETE FROM Direccion WHERE id_Direccion = %s", (id_direccion,))

            if id_rol is not None:
                cursor.execute("SELECT COUNT(*) FROM UsuarioSistema WHERE Rol_idRol = %s", (id_rol,))
                if cursor.fetchone()[0] == 0:
                    cursor.execute("DELETE FROM Rol WHERE id_Rol = %s", (id_rol,))

            cnx.commit()
            print(f"Usuario '{nombre_usuario}' (ID {id_usuario}) eliminado correctamente.")

        except Exception as err:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Se produjo un error al eliminar el usuario: {err}")

    def listar_usuarios(self, cursor):
        try:
            cursor.execute("""
                SELECT U.id_UsuarioSistema, U.Nombre, U.Apellido, U.RUT, U.Email, U.Telefono,
                    U.NombreUsuario, R.tipo_Rol,
                    D.Calle, D.Numero, D.Ciudad, D.Region, D.Pais, D.CodigoPostal,
                    E.FechaContrato, E.Salario, E.TipoEmpleado
                FROM UsuarioSistema U
                INNER JOIN Rol R ON U.Rol_idRol = R.id_Rol
                INNER JOIN Direccion D ON U.Direccion_idDireccion = D.id_Direccion
                LEFT JOIN DetalleEmpleado E ON U.id_UsuarioSistema = E.UsuarioSistema_idUsuarioSistema
                ORDER BY U.id_UsuarioSistema
            """)
            usuarios = cursor.fetchall()

            if not usuarios:
                print("No hay usuarios registrados.")
                return

            pagina = 0
            por_pagina = 5
            total = len(usuarios)

            while True:
                inicio = pagina * por_pagina
                fin = inicio + por_pagina
                pag_usuarios = usuarios[inicio:fin]

                if not pag_usuarios:
                    print("No hay más usuarios para mostrar.")
                    break

                print(f"\n--- PÁGINA {pagina + 1} DE {(total + por_pagina - 1) // por_pagina} ---")
                for u in pag_usuarios:
                    print(f"""
    ID: {u[0]} - {u[1]} {u[2]} (Usuario: {u[6]}, Rol: {u[7]})
    RUT: {u[3]}, Email: {u[4]}, Teléfono: {u[5]}
    Dirección: {u[8]} {u[9]}, {u[10]}, {u[11]}, {u[12]}, CP: {u[13]}
    Detalle Empleado → Fecha Contrato: {u[14]}, Salario: {u[15]}, Tipo: {u[16]}
    ----------------------------------------
    """)

                if fin >= total:
                    print("Fin del listado de usuarios.")
                    break

                continuar = pedir_input("¿Ver siguiente página? (s/n, 0 para cancelar): ")
                if continuar is None or continuar.strip().lower() != 's':
                    print("Operación cancelada o fin del listado.")
                    break

                pagina += 1

        except Exception as e:
            print(f"Ocurrió un error al listar usuarios: {e}")

    def buscar_usuarios_especificos(self, cursor):
        print("\n--- BUSCAR USUARIOS ESPECÍFICOS ---")
        print("""
    1. ID de usuario
    2. Nombre de usuario
    3. Correo electrónico
    4. Teléfono
    0. Cancelar
    """)

        opcion = pedir_input("Seleccione una opción: ")
        if opcion is None or opcion == "0":
            return

        filtro = ""
        valor = None

        if opcion == "1":
            id_usuario = pedir_input("Ingrese el ID del usuario [0 para cancelar]: ")
            if id_usuario is None: return
            if not id_usuario.isdigit():
                print("Error: El ID debe ser un número entero.")
                return
            filtro = "U.id_UsuarioSistema = %s"
            valor = (int(id_usuario),)

        elif opcion == "2":
            nombre = pedir_input("Ingrese el nombre de usuario o parte de él [0 para cancelar]: ")
            if nombre is None or not nombre:
                print("Campo vacío o cancelado.")
                return
            filtro = "LOWER(U.NombreUsuario) LIKE LOWER(%s)"
            valor = (f"%{nombre}%",)

        elif opcion == "3":
            correo = pedir_input("Ingrese el correo o parte del correo [0 para cancelar]: ")
            if correo is None or not correo:
                print("Campo vacío o cancelado.")
                return
            filtro = "LOWER(U.Email) LIKE LOWER(%s)"
            valor = (f"%{correo}%",)

        elif opcion == "4":
            telefono = pedir_input("Ingrese el teléfono o parte del teléfono [0 para cancelar]: ")
            if telefono is None or not telefono:
                print("Campo vacío o cancelado.")
                return
            filtro = "U.Telefono LIKE %s"
            valor = (f"%{telefono}%",)

        else:
            print("Opción no válida.")
            return

        try:
            cursor.execute(f"""
                SELECT 
                    U.id_UsuarioSistema, U.Nombre, U.Apellido, U.RUT, U.Email, U.Telefono, 
                    U.NombreUsuario, R.tipo_Rol,
                    D.Calle, D.Numero, D.Ciudad, D.Region, D.Pais, D.CodigoPostal,
                    E.FechaContrato, E.Salario, E.TipoEmpleado
                FROM UsuarioSistema U
                INNER JOIN Rol R ON U.Rol_idRol = R.id_Rol
                INNER JOIN Direccion D ON U.Direccion_idDireccion = D.id_Direccion
                LEFT JOIN DetalleEmpleado E ON U.id_UsuarioSistema = E.UsuarioSistema_idUsuarioSistema
                WHERE {filtro}
            """, valor)

            resultados = cursor.fetchall()

            if not resultados:
                print("No se encontraron usuarios que coincidan con la búsqueda.")
                return

            print("\n--- RESULTADOS ENCONTRADOS ---")
            for u in resultados:
                print(f"""
    ID: {u[0]} - {u[1]} {u[2]} (Usuario: {u[6]}, Rol: {u[7]})
    RUT: {u[3]}, Email: {u[4]}, Teléfono: {u[5]}
    Dirección: {u[8]} {u[9]}, {u[10]}, {u[11]}, {u[12]}, CP: {u[13]}
    Detalle Empleado → Fecha Contrato: {u[14]}, Salario: {u[15]}, Tipo: {u[16]}
    ----------------------------------------
    """)

        except Exception as err:
            print(f"Error al buscar usuarios: {err}")







    # --- DEPARTAMENTOS: LISTO ---

    def crear_departamento(self, cnx, cursor):
        try:
            print("\n--- CREAR DEPARTAMENTO ---")

            # Nombre obligatorio
            while True:
                nombre = pedir_input("Nombre del departamento [0 para cancelar]: ", cnx)
                if nombre is None:
                    return
                if not nombre.strip():
                    print("El nombre no puede estar vacío.")
                    continue

                if len(nombre) > 50:
                    print("El nombre no puede superar los 50 caracteres.")
                    continue

                # Evitar duplicados
                cursor.execute("SELECT COUNT(*) FROM Departamento WHERE LOWER(NombreDepartamento) = LOWER(%s)", (nombre,))
                if cursor.fetchone()[0] > 0:
                    print("Ya existe un departamento con ese nombre.")
                    continue
                break

            # Tipo obligatorio
            while True:
                tipo = pedir_input("Tipo de departamento [0 para cancelar]: ", cnx)
                if tipo is None:
                    return
                if not tipo.strip():
                    print("El tipo no puede estar vacío.")
                    continue

                if len(tipo) > 50:
                    print("El tipo no puede superar los 50 caracteres.")
                    continue
                break

            # Descripción opcional
            descripcion = pedir_input("Descripción del departamento (opcional, 0 para cancelar): ", cnx, opcional=True)
            if descripcion is None:
                return
            if descripcion and len(descripcion) > 200:
                print("La descripción no puede superar los 200 caracteres.")
                return

            # Inserción segura
            cursor.execute("""
                INSERT INTO Departamento (NombreDepartamento, TipoDepartamento, DescripcionDepartamento, Direccion_idDireccion)
                VALUES (%s, %s, %s, %s)
            """, (nombre, tipo, descripcion, getattr(self, "Direccion_idDireccion", None)))

            cnx.commit()
            id_departamento = cursor.lastrowid
            print(f"Departamento creado correctamente (ID: {id_departamento}) → {nombre} / {tipo}")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al crear el departamento: {e}")

    def editar_departamento(self, cnx, cursor):
        try:
            print("\n--- EDITAR DEPARTAMENTO ---")

            # Listar departamentos existentes
            cursor.execute("SELECT id_Departamento, NombreDepartamento, TipoDepartamento FROM Departamento ORDER BY id_Departamento")
            departamentos = cursor.fetchall()

            if not departamentos:
                print("No hay departamentos registrados.")
                return

            print("\nDepartamentos existentes:")
            for dep in departamentos:
                print(f"  {dep[0]} - {dep[1]} ({dep[2]})")

            # Seleccionar ID válido
            while True:
                id_dep = pedir_input("Ingrese el ID del departamento a editar [0 para cancelar]: ", cnx)
                if id_dep is None:
                    return
                if not id_dep.isdigit():
                    print("Debe ingresar un número válido.")
                    continue

                id_dep = int(id_dep)
                cursor.execute("""
                    SELECT id_Departamento, NombreDepartamento, TipoDepartamento, DescripcionDepartamento
                    FROM Departamento
                    WHERE id_Departamento = %s
                """, (id_dep,))
                depto = cursor.fetchone()

                if not depto:
                    print("ID de departamento no existe. Intente nuevamente.")
                    continue
                break

            actual_nombre, actual_tipo, actual_desc = depto[1], depto[2], depto[3]

            # Nuevos valores
            nuevo_nombre = pedir_input(
                f"Nuevo nombre (actual: {actual_nombre}) [ENTER para mantener, 0 para cancelar]: ", cnx, opcional=True)
            if nuevo_nombre is None:
                return
            if not nuevo_nombre:
                nuevo_nombre = actual_nombre
            elif len(nuevo_nombre) > 50:
                print("El nombre no puede superar los 50 caracteres.")
                return

            # Validar duplicados si cambia el nombre
            if nuevo_nombre != actual_nombre:
                cursor.execute(
                    "SELECT COUNT(*) FROM Departamento WHERE LOWER(NombreDepartamento) = LOWER(%s) AND id_Departamento != %s",
                    (nuevo_nombre, id_dep))
                if cursor.fetchone()[0] > 0:
                    print("Ya existe otro departamento con ese nombre.")
                    return

            nuevo_tipo = pedir_input(
                f"Nuevo tipo (actual: {actual_tipo}) [ENTER para mantener, 0 para cancelar]: ", cnx, opcional=True)
            if nuevo_tipo is None:
                return
            if not nuevo_tipo:
                nuevo_tipo = actual_tipo
            elif len(nuevo_tipo) > 50:
                print("El tipo no puede superar los 50 caracteres.")
                return

            nueva_descripcion = pedir_input(
                f"Nueva descripción (actual: {actual_desc if actual_desc else '(sin descripción)'}) "
                f"[ENTER para mantener, 0 para cancelar]: ", cnx, opcional=True)
            if nueva_descripcion is None:
                return
            if not nueva_descripcion:
                nueva_descripcion = actual_desc
            elif len(nueva_descripcion) > 200:
                print("La descripción no puede superar los 200 caracteres.")
                return

            # Evitar actualizaciones vacías
            if (nuevo_nombre == actual_nombre and
                    nuevo_tipo == actual_tipo and
                    nueva_descripcion == actual_desc):
                print("No se realizaron cambios.")
                return

            # Actualizar
            cursor.execute("""
                UPDATE Departamento
                SET NombreDepartamento = %s,
                    TipoDepartamento = %s,
                    DescripcionDepartamento = %s
                WHERE id_Departamento = %s
            """, (nuevo_nombre, nuevo_tipo, nueva_descripcion, id_dep))
            cnx.commit()

            print(f"Departamento ID {id_dep} actualizado correctamente.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al editar el departamento: {e}")

    def listar_departamentos(self, cursor):
        try:
            print("\n--- LISTADO DE DEPARTAMENTOS ---")

            cursor.execute("""
                SELECT id_Departamento, NombreDepartamento, TipoDepartamento, DescripcionDepartamento
                FROM Departamento
                ORDER BY id_Departamento
            """)
            departamentos = cursor.fetchall()

            if not departamentos:
                print("No hay departamentos registrados.")
                return

            cantidad_por_pagina = 5
            total = len(departamentos)
            pagina = 0

            while True:
                inicio = pagina * cantidad_por_pagina
                fin = inicio + cantidad_por_pagina
                bloque = departamentos[inicio:fin]

                for dep in bloque:
                    print(f"""
                            ID: {dep[0]}
                            Nombre: {dep[1]}
                            Tipo: {dep[2]}
                            Descripción: {dep[3] if dep[3] else "(sin descripción)"}
                            ------------------------------
                            """.rstrip())

                if fin >= total:
                    print("No hay más departamentos.")
                    break

                # Control de paginación con posibilidad de cancelar
                avanzar = pedir_input("Presione ENTER para ver los siguientes [dejar vacío para mantener, 0 para cancelar]: ", opcional=True)
                if avanzar is None:
                    # pedir_input ya imprimió "Operación cancelada."
                    return
                # Cualquier entrada distinta de None continúa
                pagina += 1

        except Exception as e:
            print(f"Ocurrió un error al listar departamentos: {e}")

    def eliminar_departamento(self, cnx, cursor):
        try:
            print("\n--- ELIMINAR DEPARTAMENTO ---")

            # Mostrar lista de departamentos
            cursor.execute("""
                SELECT id_Departamento, NombreDepartamento, TipoDepartamento
                FROM Departamento
                ORDER BY id_Departamento
            """)
            departamentos = cursor.fetchall()

            if not departamentos:
                print("No hay departamentos para eliminar.")
                return

            print("\nDepartamentos existentes:")
            for dep in departamentos:
                print(f"  {dep[0]} - {dep[1]} ({dep[2]})")

            # Seleccionar ID válido
            while True:
                id_dep = pedir_input("Ingrese el ID del departamento a eliminar [0 para cancelar]: ", cnx)
                if id_dep is None:
                    return
                if not id_dep.isdigit():
                    print("Debe ingresar un número válido.")
                    continue

                id_dep = int(id_dep)
                if not any(dep[0] == id_dep for dep in departamentos):
                    print("ID no válido, intente nuevamente.")
                    continue
                break

            # Verificar dependencias (empleados asignados)
            cursor.execute("""
                SELECT COUNT(*) FROM EmpleadoDepartamento WHERE Departamento_idDepartamento = %s
            """, (id_dep,))
            empleados_asignados = cursor.fetchone()[0]

            if empleados_asignados > 0:
                print(f"No se puede eliminar el departamento. Hay {empleados_asignados} empleados asignados.")
                return

            # Confirmación fuerte antes de eliminar
            print("ADVERTENCIA: Esta acción eliminará permanentemente el registro.")
            confirmar = pedir_input("Escriba DELETE (en mayúsculas) para confirmar o 0 para cancelar: ", cnx, opcional=True)
            if confirmar is None or confirmar.strip().upper() != "DELETE":
                print("Operación cancelada.")
                cnx.rollback()
                return

            # Eliminar registro
            cursor.execute("DELETE FROM Departamento WHERE id_Departamento = %s", (id_dep,))
            cnx.commit()

            print(f"Departamento con ID {id_dep} eliminado correctamente.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al eliminar el departamento: {e}")





    # --- PROYECTOS: LISTO ---
        
    def crear_proyecto(self, cnx, cursor):
        try:
            print("\n--- CREAR PROYECTO ---")

            # --- NOMBRE OBLIGATORIO ---
            while True:
                nombre = pedir_input("Nombre del proyecto [0 para cancelar]: ", cnx)
                if nombre is None:
                    return
                if not nombre.strip():
                    print("El nombre no puede estar vacío.")
                    continue
                if len(nombre) > 100:
                    print("El nombre no puede superar los 100 caracteres.")
                    continue

                # Validar duplicado
                cursor.execute("SELECT COUNT(*) FROM Proyecto WHERE LOWER(NombreProyecto) = LOWER(%s)", (nombre,))
                if cursor.fetchone()[0] > 0:
                    print("Ya existe un proyecto con ese nombre. Elija otro.")
                    continue
                break

            # --- DESCRIPCIÓN OPCIONAL ---
            descripcion = pedir_input("Descripción del proyecto (opcional, 0 para cancelar): ", cnx, opcional=True)
            if descripcion is None:
                return
            if descripcion and len(descripcion) > 300:
                print("La descripción no puede superar los 300 caracteres.")
                return

            # --- FECHA OPCIONAL ---
            fecha_inicio = pedir_input(
                "Fecha de inicio (YYYY-MM-DD, dejar vacío si no aplica, 0 para cancelar): ",
                cnx, opcional=True
            )
            if fecha_inicio is None:
                return

            if fecha_inicio:
                try:
                    datetime.strptime(fecha_inicio, "%Y-%m-%d")
                except ValueError:
                    print("⚠️ Fecha inválida. Se dejará sin fecha de inicio.")
                    fecha_inicio = None
            else:
                fecha_inicio = None

            # --- INSERCIÓN ---
            cursor.execute("""
                INSERT INTO Proyecto (NombreProyecto, DescripcionProyecto, FechaInicioProyecto)
                VALUES (%s, %s, %s)
            """, (nombre, descripcion, fecha_inicio))

            cnx.commit()
            id_proyecto = cursor.lastrowid
            print(f"Proyecto '{nombre}' creado correctamente con ID {id_proyecto}.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Error al crear el proyecto: {e}")

    def listar_proyectos(self, cursor):
        try:
            print("\n--- LISTADO DE PROYECTOS ---")

            cursor.execute("""
                SELECT id_Proyecto, NombreProyecto, FechaInicioProyecto, DescripcionProyecto
                FROM Proyecto
                ORDER BY id_Proyecto
            """)
            proyectos = cursor.fetchall()

            if not proyectos:
                print("No hay proyectos registrados.")
                return

            cantidad_por_pagina = 5
            total = len(proyectos)
            pagina = 0

            while True:
                inicio = pagina * cantidad_por_pagina
                fin = inicio + cantidad_por_pagina
                bloque = proyectos[inicio:fin]

                for p in bloque:
                    print(f"""
                    ID: {p[0]}
                    Nombre: {p[1]}
                    Fecha inicio: {p[2] if p[2] else '(sin definir)'}
                    Descripción: {p[3] if p[3] else '(sin descripción)'}
                    ----------------------------------------
                    """.rstrip())

                if fin >= total:
                    print("Fin del listado de proyectos.")
                    break

                opcion = pedir_input("¿Ver siguiente página? (s/n, 0 para cancelar): ", opcional=True)
                if opcion is None or opcion.lower() != "s":
                    print("Operación cancelada o fin del listado.")
                    break

                pagina += 1

        except Exception as e:
            print(f"Ocurrió un error al listar proyectos: {e}")

    def editar_proyecto(self, cnx, cursor):
        try:
            print("\n--- EDITAR PROYECTO ---")

            # Mostrar proyectos disponibles
            cursor.execute("SELECT id_Proyecto, NombreProyecto, DescripcionProyecto, FechaInicioProyecto FROM Proyecto")
            proyectos = cursor.fetchall()

            if not proyectos:
                print("No hay proyectos registrados.")
                return

            print("Proyectos existentes:")
            for p in proyectos:
                print(f"{p[0]} - {p[1]} | Fecha inicio: {p[3]}")

            # Seleccionar ID válido
            while True:
                id_pro = pedir_input("Ingrese el ID del proyecto a editar [0 para cancelar]: ", cnx)
                if id_pro is None:
                    return
                if id_pro.isdigit():
                    id_pro = int(id_pro)
                    cursor.execute("SELECT * FROM Proyecto WHERE id_Proyecto = %s", (id_pro,))
                    proyecto = cursor.fetchone()
                    if proyecto:
                        break
                    print("ID de proyecto no existe. Intente nuevamente.")
                else:
                    print("Debe ingresar un número válido.")

            # Campos actuales
            actual_id, actual_nombre, actual_descripcion, actual_fecha = proyecto

            # Pedir nuevos valores
            nuevo_nombre = pedir_input(
                f"Nuevo nombre (actual: {actual_nombre}) [ENTER para mantener, 0 para cancelar]: ",
                cnx, opcional=True
            )
            if nuevo_nombre is None:
                return
            nuevo_nombre = nuevo_nombre.strip() or actual_nombre

            # Validar duplicado si cambió el nombre
            if nuevo_nombre.lower() != actual_nombre.lower():
                cursor.execute(
                    "SELECT COUNT(*) FROM Proyecto WHERE LOWER(NombreProyecto) = LOWER(%s) AND id_Proyecto != %s",
                    (nuevo_nombre, id_pro)
                )
                if cursor.fetchone()[0] > 0:
                    print("Ya existe otro proyecto con ese nombre.")
                    return

            nueva_descripcion = pedir_input(
                f"Nueva descripción (actual: {actual_descripcion if actual_descripcion else '(sin descripción)'}) "
                "[ENTER para mantener, 0 para cancelar]: ",
                cnx, opcional=True
            )
            if nueva_descripcion is None:
                return
            nueva_descripcion = nueva_descripcion or actual_descripcion

            nueva_fecha = pedir_input(
                f"Nueva fecha de inicio (actual: {actual_fecha if actual_fecha else '(sin definir)'}) "
                "(YYYY-MM-DD, dejar vacío para mantener, 0 para cancelar): ",
                cnx, opcional=True
            )
            if nueva_fecha is None:
                return

            # Validar fecha si se cambió
            if nueva_fecha:
                try:
                    datetime.strptime(nueva_fecha, "%Y-%m-%d")
                except ValueError:
                    print("Fecha inválida. Se mantendrá la anterior.")
                    nueva_fecha = actual_fecha
            else:
                nueva_fecha = actual_fecha

            # Verificar si hubo cambios
            if (
                nuevo_nombre == actual_nombre
                and nueva_descripcion == actual_descripcion
                and nueva_fecha == actual_fecha
            ):
                print("No se realizaron cambios.")
                return

            # Actualizar proyecto
            cursor.execute("""
                UPDATE Proyecto
                SET NombreProyecto = %s,
                    DescripcionProyecto = %s,
                    FechaInicioProyecto = %s
                WHERE id_Proyecto = %s
            """, (nuevo_nombre, nueva_descripcion, nueva_fecha, id_pro))

            cnx.commit()
            print(f"Proyecto '{nuevo_nombre}' (ID {id_pro}) actualizado correctamente.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al editar el proyecto: {e}")

    def eliminar_proyecto(self, cnx, cursor):   
        try:
            print("\n--- ELIMINAR PROYECTO ---")

            cursor.execute("SELECT id_Proyecto, NombreProyecto FROM Proyecto ORDER BY id_Proyecto")
            proyectos = cursor.fetchall()

            if not proyectos:
                print("No hay proyectos para eliminar.")
                return

            for p in proyectos:
                print(f"{p[0]} - {p[1]}")

            # Seleccionar ID válido
            while True:
                id_pro = pedir_input("Ingrese el ID del proyecto a eliminar [0 para cancelar]: ", cnx)
                if id_pro is None:
                    return
                if id_pro.isdigit():
                    id_pro = int(id_pro)
                    if any(p[0] == id_pro for p in proyectos):
                        break
                    print("ID no válido, intente de nuevo.")
                else:
                    print("Debe ingresar un número válido.")

            # Verificar si tiene empleados asignados
            cursor.execute("""
                SELECT COUNT(*) 
                FROM EmpleadoProyecto 
                WHERE Proyecto_idProyecto = %s AND Activo = TRUE
            """, (id_pro,))
            count = cursor.fetchone()[0]

            if count > 0:
                print(f"No se puede eliminar el proyecto. Tiene {count} empleados asignados actualmente.")
                return

            # Confirmar eliminación
            confirm = pedir_input("Escriba DELETE para confirmar o 0 para cancelar: ", cnx)
            if confirm is None or confirm.strip().upper() != "DELETE":
                print("Operación cancelada.")
                return

            # Eliminar proyecto
            cursor.execute("SELECT NombreProyecto FROM Proyecto WHERE id_Proyecto = %s", (id_pro,))
            nombre = cursor.fetchone()[0]

            cursor.execute("DELETE FROM Proyecto WHERE id_Proyecto = %s", (id_pro,))
            cnx.commit()
            print(f"Proyecto '{nombre}' (ID {id_pro}) eliminado correctamente.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al eliminar el proyecto: {e}")










    # Asignaciones: LISTAS

    # --- EmpleadoDepartamento ---

    def asignar_empleado_a_departamento(self, cnx, cursor):
        try:
            print("\n--- ASIGNAR EMPLEADO A DEPARTAMENTO ---")

            # --- LISTAR USUARIOS DISPONIBLES ---
            cursor.execute("""
                SELECT id_UsuarioSistema, Nombre, Apellido 
                FROM UsuarioSistema
                ORDER BY id_UsuarioSistema
            """)
            usuarios = cursor.fetchall()
            if not usuarios:
                print("No hay usuarios registrados.")
                return

            print("\nUsuarios disponibles:")
            for u in usuarios:
                print(f"  {u[0]} - {u[1]} {u[2]}")

            # --- SELECCIONAR USUARIO ---
            while True:
                id_usuario = pedir_input("Ingrese el ID del usuario a asignar [0 para cancelar]: ", cnx)
                if id_usuario is None:
                    return
                if not id_usuario.isdigit():
                    print("Debe ingresar un número válido.")
                    continue

                id_usuario = int(id_usuario)
                if not any(u[0] == id_usuario for u in usuarios):
                    print("No existe un usuario con ese ID.")
                    continue
                break

            # --- LISTAR DEPARTAMENTOS DISPONIBLES ---
            cursor.execute("""
                SELECT id_Departamento, NombreDepartamento 
                FROM Departamento
                ORDER BY id_Departamento
            """)
            departamentos = cursor.fetchall()
            if not departamentos:
                print("No hay departamentos registrados.")
                return

            print("\nDepartamentos disponibles:")
            for d in departamentos:
                print(f"  {d[0]} - {d[1]}")

            # --- SELECCIONAR DEPARTAMENTO ---
            while True:
                id_departamento = pedir_input("Ingrese el ID del departamento [0 para cancelar]: ", cnx)
                if id_departamento is None:
                    return
                if not id_departamento.isdigit():
                    print("Debe ingresar un número válido.")
                    continue

                id_departamento = int(id_departamento)
                if not any(d[0] == id_departamento for d in departamentos):
                    print("No existe un departamento con ese ID.")
                    continue
                break

            # --- VERIFICAR ASIGNACIÓN EXISTENTE ---
            cursor.execute("""
                SELECT id_EmpleadoDepartamento, Activo 
                FROM EmpleadoDepartamento
                WHERE UsuarioSistema_idUsuarioSistema = %s AND Departamento_idDepartamento = %s
            """, (id_usuario, id_departamento))
            asignacion = cursor.fetchone()

            if asignacion:
                if asignacion[1]:  # ya activo
                    print("El usuario ya está asignado a este departamento.")
                    return
                else:
                    # Reactivar asignación
                    cursor.execute("""
                        UPDATE EmpleadoDepartamento 
                        SET Activo = TRUE, FechaAsignacion = CURRENT_TIMESTAMP
                        WHERE id_EmpleadoDepartamento = %s
                    """, (asignacion[0],))
                    cnx.commit()
                    print("Asignación reactivada correctamente.")
                    return

            # --- CREAR NUEVA ASIGNACIÓN ---
            cursor.execute("""
                INSERT INTO EmpleadoDepartamento (Departamento_idDepartamento, UsuarioSistema_idUsuarioSistema)
                VALUES (%s, %s)
            """, (id_departamento, id_usuario))
            cnx.commit()

            print(f"Empleado ID {id_usuario} asignado correctamente al departamento ID {id_departamento}.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al asignar el empleado: {e}")

    def listar_empleados_en_departamento(self, cursor):
        try:
            print("\n--- LISTADO DE EMPLEADOS EN DEPARTAMENTOS ---")

            cursor.execute("""
                SELECT 
                    ED.id_EmpleadoDepartamento,
                    U.id_UsuarioSistema, U.Nombre, U.Apellido, U.NombreUsuario,
                    D.id_Departamento, D.NombreDepartamento, 
                    R.tipo_Rol,
                    ED.Activo, ED.FechaAsignacion
                FROM EmpleadoDepartamento ED
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
                JOIN Rol R ON U.Rol_idRol = R.id_Rol
                ORDER BY D.NombreDepartamento, U.Nombre
            """)
            asignaciones = cursor.fetchall()

            if not asignaciones:
                print("No hay empleados asignados a ningún departamento.")
                return

            por_pagina = 5
            total = len(asignaciones)
            pagina = 0

            while True:
                inicio = pagina * por_pagina
                fin = inicio + por_pagina
                grupo = asignaciones[inicio:fin]

                if not grupo:
                    print("No hay más resultados para mostrar.")
                    break

                print(f"\n--- PÁGINA {pagina + 1} DE {((total - 1) // por_pagina) + 1} ---")
                for a in grupo:
                    estado = "Activo" if a[8] else "Inactivo"
                    print(f"""
                    ID Asignación: {a[0]} | Fecha: {a[9]}
                    Empleado: {a[2]} {a[3]} ({a[4]})
                    Departamento: {a[6]} (ID {a[5]})
                    Rol: {a[7]}
                    Estado: {estado}
                    ----------------------------------------
                    """.rstrip())

                if fin >= total:
                    print("Fin del listado.")
                    break

                continuar = pedir_input("¿Ver siguiente página? (s/n, 0 para cancelar): ", opcional=True)
                if continuar is None or continuar.lower() != 's':
                    print("Operación cancelada o fin del listado.")
                    break

                pagina += 1

        except Exception as e:
            print(f"Ocurrió un error al listar los empleados: {e}")

    def editar_asignacion_empleado_departamento(self, cnx, cursor):
        try:
            print("\n--- EDITAR ASIGNACIÓN DE EMPLEADO A DEPARTAMENTO ---")

            # --- LISTAR ASIGNACIONES EXISTENTES ---
            cursor.execute("""
                SELECT 
                    ED.id_EmpleadoDepartamento, ED.Activo, ED.FechaAsignacion,
                    U.id_UsuarioSistema, U.Nombre, U.Apellido, U.NombreUsuario,
                    D.id_Departamento, D.NombreDepartamento
                FROM EmpleadoDepartamento ED
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
                ORDER BY ED.id_EmpleadoDepartamento
            """)
            asignaciones = cursor.fetchall()

            if not asignaciones:
                print("No hay asignaciones registradas.")
                return

            print("\nAsignaciones disponibles:")
            for a in asignaciones:
                estado = "Activo" if a[1] else "Inactivo"
                print(f"{a[0]} - {a[4]} {a[5]} ({a[6]}) | Departamento: {a[8]} | Estado: {estado}")

            # --- SELECCIONAR ASIGNACIÓN ---
            id_asignacion = pedir_input("Ingrese el ID de la asignación a editar [0 para cancelar]: ", cnx)
            if id_asignacion is None:
                return
            if not id_asignacion.isdigit():
                print("Debe ingresar un número válido.")
                return

            id_asignacion = int(id_asignacion)
            asignacion = next((a for a in asignaciones if a[0] == id_asignacion), None)
            if not asignacion:
                print("No existe una asignación con ese ID.")
                return

            id_usuario = asignacion[3]
            id_departamento_actual = asignacion[7]
            estado_actual = asignacion[1]

            print(f"\nEmpleado: {asignacion[4]} {asignacion[5]} ({asignacion[6]})")
            print(f"Departamento actual: {asignacion[8]}")
            print(f"Estado actual: {'Activo' if estado_actual else 'Inactivo'}")

            # --- OPCIONES ---
            print("""
    1. Cambiar departamento
    2. Activar / Desactivar asignación
    0. Cancelar
    """)
            opcion = pedir_input("Seleccione una opción: ", cnx, opcional=True)
            if opcion is None or opcion == "0":
                return

            # --- OPCIÓN 1: CAMBIAR DEPARTAMENTO ---
            if opcion == "1":
                cursor.execute("SELECT id_Departamento, NombreDepartamento FROM Departamento ORDER BY id_Departamento")
                departamentos = cursor.fetchall()
                if not departamentos:
                    print("No hay departamentos registrados.")
                    return

                print("\nDepartamentos disponibles:")
                for d in departamentos:
                    print(f"{d[0]} - {d[1]}")

                nuevo_id = pedir_input("Ingrese el ID del nuevo departamento [0 para cancelar]: ", cnx)
                if nuevo_id is None:
                    return
                if not nuevo_id.isdigit():
                    print("Debe ingresar un número válido.")
                    return

                nuevo_id = int(nuevo_id)
                if nuevo_id == id_departamento_actual:
                    print("El empleado ya pertenece a ese departamento.")
                    return

                # Comprobar asignación previa
                cursor.execute("""
                    SELECT id_EmpleadoDepartamento, Activo 
                    FROM EmpleadoDepartamento
                    WHERE UsuarioSistema_idUsuarioSistema = %s AND Departamento_idDepartamento = %s
                """, (id_usuario, nuevo_id))
                existente = cursor.fetchone()

                if existente:
                    if existente[1]:
                        print("El empleado ya está activo en ese departamento.")
                        return
                    else:
                        cursor.execute("""
                            UPDATE EmpleadoDepartamento 
                            SET Activo = TRUE, FechaAsignacion = CURRENT_TIMESTAMP
                            WHERE id_EmpleadoDepartamento = %s
                        """, (existente[0],))
                        cnx.commit()
                        print("Empleado reactivado en el nuevo departamento.")
                        return

                # Actualizar registro existente
                cursor.execute("""
                    UPDATE EmpleadoDepartamento
                    SET Departamento_idDepartamento = %s, FechaAsignacion = CURRENT_TIMESTAMP
                    WHERE id_EmpleadoDepartamento = %s
                """, (nuevo_id, id_asignacion))
                cnx.commit()
                print("Departamento actualizado correctamente.")

            # --- OPCIÓN 2: ACTIVAR / DESACTIVAR ---
            elif opcion == "2":
                nuevo_estado = not estado_actual
                cursor.execute("""
                    UPDATE EmpleadoDepartamento
                    SET Activo = %s, FechaAsignacion = CURRENT_TIMESTAMP
                    WHERE id_EmpleadoDepartamento = %s
                """, (nuevo_estado, id_asignacion))
                cnx.commit()
                print(f"Asignación {'activada' if nuevo_estado else 'desactivada'} correctamente.")

            else:
                print("Opción no válida.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al editar la asignación: {e}")

    def eliminar_asignacion_empleado_departamento(self, cnx, cursor):
        try:
            print("\n--- DESASIGNAR EMPLEADO DE DEPARTAMENTO ---")

            # --- LISTAR ASIGNACIONES ---
            cursor.execute("""
                SELECT 
                    ED.id_EmpleadoDepartamento, ED.Activo, ED.FechaAsignacion,
                    U.Nombre, U.Apellido, U.NombreUsuario,
                    D.NombreDepartamento
                FROM EmpleadoDepartamento ED
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
                ORDER BY ED.id_EmpleadoDepartamento
            """)
            asignaciones = cursor.fetchall()

            if not asignaciones:
                print("No hay asignaciones registradas.")
                return

            print("\nAsignaciones disponibles:")
            for a in asignaciones:
                estado = "Activo" if a[1] else "Inactivo"
                print(f"{a[0]} - {a[3]} {a[4]} ({a[5]}) | Departamento: {a[6]} | Estado: {estado}")

            # --- PEDIR ID ---
            id_asignacion = pedir_input("Ingrese el ID de la asignación a eliminar [0 para cancelar]: ", cnx)
            if id_asignacion is None:
                return
            if not id_asignacion.isdigit():
                print("Debe ingresar un número válido.")
                return

            id_asignacion = int(id_asignacion)
            asignacion = next((a for a in asignaciones if a[0] == id_asignacion), None)
            if not asignacion:
                print("No existe una asignación con ese ID.")
                return

            # --- VERIFICAR DEPENDENCIAS (proyectos activos) ---
            cursor.execute("""
                SELECT COUNT(*) 
                FROM EmpleadoProyecto 
                WHERE EmpleadoDepartamento_idEmpleadoDepartamento = %s
            """, (id_asignacion,))
            proyectos_asociados = cursor.fetchone()[0]

            if proyectos_asociados > 0:
                print(f"No se puede eliminar la asignación. El empleado participa en {proyectos_asociados} proyecto(s).")
                print("Debe desasignarlo de esos proyectos antes de eliminar esta relación.")
                return

            # --- CONFIRMACIÓN ---
            confirm = pedir_input(
                f"Confirme la eliminación de la asignación del empleado {asignacion[3]} {asignacion[4]} "
                f"en el departamento '{asignacion[6]}'.\nEscriba DELETE para confirmar o 0 para cancelar: ",
                cnx
            )
            if confirm is None or confirm.strip().upper() != "DELETE":
                print("Operación cancelada.")
                return

            # --- ELIMINAR ---
            cursor.execute("DELETE FROM EmpleadoDepartamento WHERE id_EmpleadoDepartamento = %s", (id_asignacion,))
            cnx.commit()
            print("Asignación eliminada correctamente.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al eliminar la asignación: {e}")

    # --- EmpleadoProyecto ---

    def asignar_empleado_a_proyecto(self, cnx, cursor):
        try:
            print("\n--- ASIGNAR EMPLEADO A PROYECTO ---")

            # --- EMPLEADOS ACTIVOS EN DEPARTAMENTOS ---
            cursor.execute("""
                SELECT 
                    ED.id_EmpleadoDepartamento,
                    U.id_UsuarioSistema, U.Nombre, U.Apellido, U.NombreUsuario,
                    D.NombreDepartamento
                FROM EmpleadoDepartamento ED
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
                WHERE ED.Activo = TRUE
                ORDER BY U.Nombre
            """)
            empleados = cursor.fetchall()

            if not empleados:
                print("No hay empleados activos en ningún departamento.")
                return

            print("\nEmpleados disponibles:")
            for e in empleados:
                print(f"  {e[0]} - {e[2]} {e[3]} ({e[4]}) | Departamento: {e[5]}")

            # --- SELECCIONAR EMPLEADO ---
            id_empleado_dep = pedir_input("Ingrese el ID del empleado a asignar [0 para cancelar]: ", cnx)
            if id_empleado_dep is None:
                return
            if not id_empleado_dep.isdigit():
                print("Debe ingresar un número válido.")
                return
            id_empleado_dep = int(id_empleado_dep)

            empleado = next((e for e in empleados if e[0] == id_empleado_dep), None)
            if not empleado:
                print("No existe un empleado con ese ID.")
                return

            # --- PROYECTOS DISPONIBLES ---
            cursor.execute("SELECT id_Proyecto, NombreProyecto FROM Proyecto ORDER BY id_Proyecto")
            proyectos = cursor.fetchall()
            if not proyectos:
                print("No hay proyectos registrados.")
                return

            print("\nProyectos disponibles:")
            for p in proyectos:
                print(f"  {p[0]} - {p[1]}")

            id_proyecto = pedir_input("Ingrese el ID del proyecto [0 para cancelar]: ", cnx)
            if id_proyecto is None:
                return
            if not id_proyecto.isdigit():
                print("Debe ingresar un número válido.")
                return
            id_proyecto = int(id_proyecto)

            proyecto = next((p for p in proyectos if p[0] == id_proyecto), None)
            if not proyecto:
                print("No existe un proyecto con ese ID.")
                return

            # --- VERIFICAR EXISTENCIA ---
            cursor.execute("""
                SELECT id_DetalleProyecto, Activo
                FROM EmpleadoProyecto
                WHERE EmpleadoDepartamento_idEmpleadoDepartamento = %s AND Proyecto_idProyecto = %s
            """, (id_empleado_dep, id_proyecto))
            existente = cursor.fetchone()

            if existente:
                if existente[1]:
                    print("El empleado ya está asignado activamente a este proyecto.")
                    return
                else:
                    cursor.execute("""
                        UPDATE EmpleadoProyecto
                        SET Activo = TRUE, FechaProyectoInscrito = CURRENT_TIMESTAMP
                        WHERE id_DetalleProyecto = %s
                    """, (existente[0],))
                    cnx.commit()
                    print("Asignación reactivada correctamente.")
                    return

            # --- HORAS ---
            while True:
                horas = pedir_input("Ingrese cantidad de horas asignadas (mayor a 0) [0 para cancelar]: ", cnx)
                if horas is None:
                    return
                if not horas.isdigit() or int(horas) <= 0:
                    print("Debe ingresar un número entero mayor a 0.")
                    continue
                horas = int(horas)
                break

            # --- DESCRIPCIÓN ---
            descripcion = pedir_input("Descripción de la tarea (opcional, 0 para cancelar): ", cnx, opcional=True)
            if descripcion is None:
                return

            # --- INSERTAR NUEVA ASIGNACIÓN ---
            cursor.execute("""
                INSERT INTO EmpleadoProyecto 
                (EmpleadoDepartamento_idEmpleadoDepartamento, Proyecto_idProyecto, CantidadHorasEmpleadoProyecto, DescripcionTareaProyecto)
                VALUES (%s, %s, %s, %s)
            """, (id_empleado_dep, id_proyecto, horas, descripcion))
            cnx.commit()

            print(f"Empleado '{empleado[2]} {empleado[3]}' asignado correctamente al proyecto '{proyecto[1]}'.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al asignar el empleado al proyecto: {e}")

    def editar_empleado_proyecto(self, cnx, cursor):
        try:
            print("\n--- EDITAR ASIGNACIÓN DE EMPLEADO A PROYECTO ---")

            # --- LISTAR ASIGNACIONES ACTIVAS ---
            cursor.execute("""
                SELECT 
                    EP.id_DetalleProyecto,
                    U.Nombre, U.Apellido,
                    P.NombreProyecto,
                    EP.CantidadHorasEmpleadoProyecto,
                    EP.DescripcionTareaProyecto
                FROM EmpleadoProyecto EP
                JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                WHERE EP.Activo = TRUE
                ORDER BY EP.id_DetalleProyecto
            """)
            asignaciones = cursor.fetchall()

            if not asignaciones:
                print("No hay asignaciones activas para editar.")
                return

            print("\nAsignaciones activas:")
            for a in asignaciones:
                print(f"{a[0]} - {a[1]} {a[2]} | Proyecto: {a[3]} | Horas: {a[4]} | Tarea: {a[5]}")

            # --- SELECCIONAR ASIGNACIÓN ---
            id_detalle = pedir_input("Ingrese el ID del detalle a editar [0 para cancelar]: ", cnx)
            if id_detalle is None:
                return
            if not id_detalle.isdigit():
                print("Debe ingresar un número válido.")
                return
            id_detalle = int(id_detalle)

            detalle = next((a for a in asignaciones if a[0] == id_detalle), None)
            if not detalle:
                print("No existe una asignación activa con ese ID.")
                return

            print(f"\nEmpleado: {detalle[1]} {detalle[2]}")
            print(f"Proyecto: {detalle[3]}")
            print(f"Horas actuales: {detalle[4]}")
            print(f"Tarea actual: {detalle[5]}")

            # --- NUEVOS DATOS ---
            horas = pedir_input("Nueva cantidad de horas (ENTER para mantener, 0 para cancelar): ", cnx, opcional=True)
            if horas is None:
                return
            if horas == "":
                horas = detalle[4]
            elif not horas.isdigit() or int(horas) <= 0:
                print("Debe ingresar un número válido mayor a 0.")
                return
            else:
                horas = int(horas)

            descripcion = pedir_input("Nueva descripción de tarea (ENTER para mantener, 0 para cancelar): ", cnx, opcional=True)
            if descripcion is None:
                return
            if descripcion == "":
                descripcion = detalle[5]

            # --- CONFIRMACIÓN ---
            confirmar = pedir_input("¿Desea actualizar esta asignación? (s/n, 0 para cancelar): ", cnx, opcional=True)
            if confirmar is None or confirmar.lower() != "s":
                print("Operación cancelada.")
                return

            # --- ACTUALIZAR ---
            cursor.execute("""
                UPDATE EmpleadoProyecto
                SET CantidadHorasEmpleadoProyecto = %s, DescripcionTareaProyecto = %s
                WHERE id_DetalleProyecto = %s
            """, (horas, descripcion, id_detalle))
            cnx.commit()

            print("Asignación actualizada correctamente.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al editar la asignación: {e}")

    def desasignar_empleado_a_proyecto(self, cnx, cursor):
        try:
            print("\n--- DESASIGNAR EMPLEADO DE PROYECTO ---")

            cursor.execute("""
                SELECT 
                    EP.id_DetalleProyecto,
                    U.Nombre, U.Apellido,
                    P.NombreProyecto,
                    EP.CantidadHorasEmpleadoProyecto,
                    EP.DescripcionTareaProyecto
                FROM EmpleadoProyecto EP
                JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                WHERE EP.Activo = TRUE
                ORDER BY EP.id_DetalleProyecto
            """)
            asignaciones = cursor.fetchall()

            if not asignaciones:
                print("No hay asignaciones activas de empleados a proyectos.")
                return

            print("\nAsignaciones activas:")
            for a in asignaciones:
                print(f"{a[0]} - {a[1]} {a[2]} | Proyecto: {a[3]} | Horas: {a[4]} | Tarea: {a[5]}")

            id_asignacion = pedir_input("Ingrese el ID de la asignación a desasignar [0 para cancelar]: ", cnx)
            if id_asignacion is None:
                return
            if not id_asignacion.isdigit():
                print("Debe ingresar un número válido.")
                return
            id_asignacion = int(id_asignacion)

            asignacion = next((a for a in asignaciones if a[0] == id_asignacion), None)
            if not asignacion:
                print("No existe una asignación activa con ese ID.")
                return

            confirmar = pedir_input(
                f"¿Está seguro de desasignar al empleado {asignacion[1]} {asignacion[2]} del proyecto '{asignacion[3]}'? (s/n, 0 para cancelar): ",
                cnx, opcional=True
            )
            if confirmar is None or confirmar.lower() != "s":
                print("Operación cancelada.")
                return

            cursor.execute("""
                UPDATE EmpleadoProyecto
                SET Activo = FALSE
                WHERE id_DetalleProyecto = %s
            """, (id_asignacion,))
            cnx.commit()

            print(f"Empleado {asignacion[1]} {asignacion[2]} desasignado correctamente del proyecto '{asignacion[3]}'.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Ocurrió un error al desasignar al empleado del proyecto: {e}")

    def ver_detalle_empleado_proyecto(self, cursor):
        try:
            print("\n--- VER DETALLES DE ASIGNACIÓN DE EMPLEADOS A PROYECTOS ---")

            print("1. Solo asignaciones activas")
            print("2. Ver historial completo (activas + inactivas)")
            opcion = pedir_input("Seleccione una opción [0 para cancelar]: ", opcional=True)
            if opcion is None or opcion == "0":
                return
            if opcion not in ["1", "2"]:
                print("Opción inválida.")
                return

            filtro_activo = "EP.Activo = TRUE" if opcion == "1" else "1=1"

            cursor.execute(f"""
                SELECT 
                    EP.id_DetalleProyecto,
                    U.Nombre, U.Apellido, U.NombreUsuario,
                    P.NombreProyecto,
                    EP.FechaProyectoInscrito,
                    EP.CantidadHorasEmpleadoProyecto,
                    EP.DescripcionTareaProyecto,
                    CASE WHEN EP.Activo = TRUE THEN 'Activo' ELSE 'Inactivo' END AS Estado,
                    Dpto.NombreDepartamento
                FROM EmpleadoProyecto EP
                JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                JOIN Departamento Dpto ON ED.Departamento_idDepartamento = Dpto.id_Departamento
                WHERE {filtro_activo}
                ORDER BY EP.id_DetalleProyecto
            """)
            asignaciones = cursor.fetchall()

            if not asignaciones:
                print("No se encontraron asignaciones con el criterio seleccionado.")
                return

            print("\n--- LISTADO DE ASIGNACIONES ---")
            for a in asignaciones:
                print(f"""
                        ID Detalle Proyecto:   {a[0]}
                        Empleado:              {a[1]} {a[2]} ({a[3]})
                        Departamento:          {a[9]}
                        Proyecto:              {a[4]}
                        Fecha Inscripción:     {a[5]}
                        Horas Asignadas:       {a[6]}
                        Descripción Tarea:     {a[7] if a[7] else '(Sin descripción)'}
                        Estado:                {a[8]}
                        ----------------------------------------
                        """.rstrip())

            ver_detalle = pedir_input("¿Desea ver un detalle específico? (s/n, 0 para cancelar): ", opcional=True)
            if ver_detalle is None or ver_detalle.lower() != "s":
                print("Operación finalizada.")
                return

            id_detalle = pedir_input("Ingrese el ID del detalle a consultar [0 para cancelar]: ", opcional=True)
            if id_detalle is None or not id_detalle.isdigit():
                print("Debe ingresar un número válido.")
                return
            id_detalle = int(id_detalle)

            cursor.execute("""
                SELECT 
                    EP.id_DetalleProyecto,
                    U.Nombre, U.Apellido, U.NombreUsuario,
                    P.NombreProyecto,
                    EP.FechaProyectoInscrito,
                    EP.CantidadHorasEmpleadoProyecto,
                    EP.DescripcionTareaProyecto,
                    CASE WHEN EP.Activo = TRUE THEN 'Activo' ELSE 'Inactivo' END AS Estado,
                    Dpto.NombreDepartamento
                FROM EmpleadoProyecto EP
                JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                JOIN Departamento Dpto ON ED.Departamento_idDepartamento = Dpto.id_Departamento
                WHERE EP.id_DetalleProyecto = %s
            """, (id_detalle,))
            detalle = cursor.fetchone()

            if not detalle:
                print("No existe una asignación con ese ID.")
                return

            print("\n" + "="*55)
            print("     DETALLE DE EMPLEADO EN PROYECTO")
            print("="*55)
            print(f"ID Detalle Proyecto: {detalle[0]}")
            print(f"Empleado:            {detalle[1]} {detalle[2]} ({detalle[3]})")
            print(f"Departamento:        {detalle[9]}")
            print(f"Proyecto:            {detalle[4]}")
            print(f"Fecha Inscripción:   {detalle[5]}")
            print(f"Horas Asignadas:     {detalle[6]}")
            print(f"Descripción Tarea:   {detalle[7] if detalle[7] else '(Sin descripción)'}")
            print(f"Estado:              {detalle[8]}")
            print("="*55 + "\n")

        except Exception as e:
            print(f"Ocurrió un error al visualizar los detalles: {e}")
















    # --- Informes ---

    # Exportar a Excel
    # LISTO
    
    def generar_informe_empleado_proyecto(self, cnx, cursor):
        """
        Genera un informe con las asignaciones de empleados a proyectos.
        Exporta los datos a Excel y guarda un registro en la tabla 'Informe'.
        """
        try:
            print("\n--- GENERAR INFORME DE EMPLEADOS Y PROYECTOS ---")

            # Selección de filtro
            print("""
                    1. Informe general (todos los empleados y proyectos)
                    2. Informe por proyecto específico
                    3. Informe por departamento
                    0. Cancelar
                    """)
            opcion = pedir_input("Seleccione una opción: ", cnx)
            if opcion is None or opcion == "0":
                return

            filtro_sql = ""
            params = ()
            tipo_informe = "General"

            if opcion == "2":  # Por proyecto
                cursor.execute("SELECT id_Proyecto, NombreProyecto FROM Proyecto ORDER BY id_Proyecto")
                proyectos = cursor.fetchall()
                if not proyectos:
                    print("No hay proyectos registrados.")
                    return

                print("\nProyectos disponibles:")
                for p in proyectos:
                    print(f"  {p[0]} - {p[1]}")

                id_proyecto = pedir_input("Ingrese el ID del proyecto [0 para cancelar]: ", cnx)
                if id_proyecto is None or not id_proyecto.isdigit():
                    print("Operación cancelada.")
                    return
                filtro_sql = "WHERE P.id_Proyecto = %s"
                params = (int(id_proyecto),)
                tipo_informe = "Por Proyecto"

            elif opcion == "3":  # Por departamento
                cursor.execute("SELECT id_Departamento, NombreDepartamento FROM Departamento ORDER BY id_Departamento")
                departamentos = cursor.fetchall()
                if not departamentos:
                    print("No hay departamentos registrados.")
                    return

                print("\nDepartamentos disponibles:")
                for d in departamentos:
                    print(f"  {d[0]} - {d[1]}")

                id_dep = pedir_input("Ingrese el ID del departamento [0 para cancelar]: ", cnx)
                if id_dep is None or not id_dep.isdigit():
                    print("Operación cancelada.")
                    return
                filtro_sql = "WHERE D.id_Departamento = %s"
                params = (int(id_dep),)
                tipo_informe = "Por Departamento"

            # Consulta principal
            cursor.execute(f"""
                SELECT 
                    EP.id_DetalleProyecto,
                    CONCAT(U.Nombre, ' ', U.Apellido) AS Empleado,
                    D.NombreDepartamento,
                    P.NombreProyecto,
                    EP.FechaProyectoInscrito,
                    EP.CantidadHorasEmpleadoProyecto,
                    EP.DescripcionTareaProyecto,
                    CASE WHEN EP.Activo THEN 'Activo' ELSE 'Inactivo' END AS Estado
                FROM EmpleadoProyecto EP
                JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
                {filtro_sql}
                ORDER BY D.NombreDepartamento, P.NombreProyecto, U.Nombre
            """, params)
            datos = cursor.fetchall()

            if not datos:
                print("No se encontraron registros para generar el informe.")
                return

            # Crear Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe Empleados-Proyectos"

            encabezados = [
                "ID Detalle", "Empleado", "Departamento", "Proyecto",
                "Fecha Inscripción", "Horas Asignadas", "Descripción", "Estado"
            ]
            ws.append(encabezados)

            for fila in datos:
                ws.append(list(fila))

            nombre_archivo = f"informe_{tipo_informe.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            ruta = f"./informes/{nombre_archivo}"
            wb.save(ruta)

            descripcion = f"Informe {tipo_informe.lower()} generado con {len(datos)} registros."

            cursor.execute("""
                INSERT INTO Informe (NombreInforme, FechaConsulta, DescripcionInforme, TipoInforme, EmpleadoProyecto_idDetalleProyecto)
                VALUES (%s, NOW(), %s, %s, %s)
            """, (nombre_archivo, descripcion, tipo_informe, datos[0][0]))
            cnx.commit()

            print(f"\nInforme '{tipo_informe}' generado correctamente.")
            print(f"Archivo: {ruta}\nRegistros: {len(datos)}")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Error al generar el informe: {e}")

    def generar_informe_por_empleado(self, cnx, cursor):
        """
        Genera un informe completo de un empleado específico,
        incluyendo sus datos personales y proyectos asociados.
        Exporta a Excel y registra el informe en la base de datos.
        Compatible con la estructura de la tabla 'Informe' (sin valores NULL).
        """
        try:
            print("\n--- GENERAR INFORME POR EMPLEADO ---")

            # Mostrar empleados disponibles
            cursor.execute("""
                SELECT U.id_UsuarioSistema, CONCAT(U.Nombre, ' ', U.Apellido) AS Empleado, R.tipo_Rol
                FROM UsuarioSistema U
                JOIN Rol R ON U.Rol_idRol = R.id_Rol
                WHERE R.tipo_Rol = 'empleado'
                ORDER BY U.Apellido
            """)
            empleados = cursor.fetchall()

            if not empleados:
                print("No hay empleados registrados.")
                return

            print("\nEmpleados disponibles:")
            for e in empleados:
                print(f"  {e[0]} - {e[1]} ({e[2]})")

            id_emp = pedir_input("Ingrese el ID del empleado [0 para cancelar]: ", cnx)
            if id_emp is None or not id_emp.isdigit() or id_emp == "0":
                return

            id_emp = int(id_emp)

            # Datos personales del empleado
            cursor.execute("""
                SELECT 
                    U.Nombre, U.Apellido, U.RUT, U.Email, U.Telefono,
                    D.Calle, D.Numero, D.Ciudad, D.Region, D.Pais, D.CodigoPostal
                FROM UsuarioSistema U
                LEFT JOIN Direccion D ON U.Direccion_idDireccion = D.id_Direccion
                WHERE U.id_UsuarioSistema = %s
            """, (id_emp,))
            datos_personales = cursor.fetchone()

            if not datos_personales:
                print("No se encontró información personal del empleado.")
                return

            # Proyectos del empleado
            cursor.execute("""
                SELECT 
                    EP.id_DetalleProyecto,
                    P.NombreProyecto,
                    Dep.NombreDepartamento,
                    EP.FechaProyectoInscrito,
                    EP.CantidadHorasEmpleadoProyecto,
                    EP.DescripcionTareaProyecto,
                    CASE WHEN EP.Activo THEN 'Activo' ELSE 'Inactivo' END AS Estado
                FROM EmpleadoProyecto EP
                JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                JOIN Departamento Dep ON ED.Departamento_idDepartamento = Dep.id_Departamento
                WHERE ED.UsuarioSistema_idUsuarioSistema = %s
                ORDER BY EP.FechaProyectoInscrito DESC
            """, (id_emp,))
            proyectos = cursor.fetchall()

            if not proyectos:
                print("El empleado no tiene proyectos asignados. No se puede generar el informe.")
                return

            # Tomar un id_DetalleProyecto válido para cumplir la restricción NOT NULL
            id_detalle = proyectos[0][0]

            # Crear carpeta si no existe
            carpeta = "./informes"
            os.makedirs(carpeta, exist_ok=True)

            # Crear Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe Empleado"

            ws.append(["INFORME DE EMPLEADO"])
            ws.append([])

            encabezados_personales = [
                "Nombre", "Apellido", "RUT", "Email", "Teléfono",
                "Calle", "Número", "Ciudad", "Región", "País", "Código Postal"
            ]
            ws.append(encabezados_personales)
            ws.append(list(datos_personales))
            ws.append([])

            ws.append(["PROYECTOS ASOCIADOS"])
            ws.append([])
            encabezados_proyectos = [
                "ID Detalle", "Proyecto", "Departamento", "Fecha Inscripción",
                "Horas Asignadas", "Descripción Tarea", "Estado"
            ]
            ws.append(encabezados_proyectos)

            for fila in proyectos:
                ws.append(list(fila))

            nombre_archivo = f"informe_empleado_{id_emp}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            ruta = os.path.join(carpeta, nombre_archivo)
            wb.save(ruta)

            descripcion = f"Informe individual del empleado ID {id_emp}, con {len(proyectos)} proyectos listados."

            # Registrar informe en la base de datos
            cursor.execute("""
                INSERT INTO Informe (NombreInforme, FechaConsulta, DescripcionInforme, TipoInforme, EmpleadoProyecto_idDetalleProyecto)
                VALUES (%s, NOW(), %s, %s, %s)
            """, (nombre_archivo, descripcion, "Por Empleado", id_detalle))
            cnx.commit()

            print("\nInforme generado correctamente.")
            print(f"Archivo: {ruta}")
            print(f"Proyectos incluidos: {len(proyectos)}")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Error al generar informe por empleado: {e}")

    def listar_informes(self, cursor):
        """
        Lista todos los informes registrados en la tabla 'Informe'.
        Muestra nombre, fecha, tipo y descripción.
        """
        try:
            print("\n--- LISTADO DE INFORMES GENERADOS ---")

            cursor.execute("""
                SELECT id_Informe, NombreInforme, FechaConsulta, TipoInforme, DescripcionInforme
                FROM Informe
                ORDER BY FechaConsulta DESC
            """)
            informes = cursor.fetchall()

            if not informes:
                print("No hay informes registrados.")
                return

            por_pagina = 5
            total = len(informes)
            pagina = 0

            while True:
                inicio = pagina * por_pagina
                fin = inicio + por_pagina
                bloque = informes[inicio:fin]

                print(f"\n--- PÁGINA {pagina + 1} DE {(total + por_pagina - 1) // por_pagina} ---")
                for inf in bloque:
                    print(f"""
                    ID: {inf[0]}
                    Nombre: {inf[1]}
                    Fecha: {inf[2]}
                    Tipo: {inf[3]}
                    Descripción: {inf[4]}
                    ----------------------------------------
                    """.rstrip())

                if fin >= total:
                    print("Fin del listado de informes.")
                    break

                continuar = pedir_input("¿Ver siguiente página? (s/n, 0 para cancelar): ", opcional=True)
                if continuar is None or continuar.lower() != "s":
                    print("Operación cancelada o fin del listado.")
                    break

                pagina += 1

        except Exception as e:
            print(f"Error al listar informes: {e}")

    def ver_detalle_informe(self, cursor):
        """
        Muestra la información completa de un informe específico.
        """
        try:
            print("\n--- VER DETALLE DE INFORME ---")

            cursor.execute("SELECT id_Informe, NombreInforme FROM Informe ORDER BY FechaConsulta DESC")
            informes = cursor.fetchall()

            if not informes:
                print("No hay informes disponibles.")
                return

            for inf in informes:
                print(f"{inf[0]} - {inf[1]}")

            id_inf = pedir_input("Ingrese el ID del informe a consultar [0 para cancelar]: ", opcional=True)
            if id_inf is None or not id_inf.isdigit():
                print("Operación cancelada.")
                return

            id_inf = int(id_inf)
            cursor.execute("""
                SELECT id_Informe, NombreInforme, FechaConsulta, DescripcionInforme, TipoInforme
                FROM Informe
                WHERE id_Informe = %s
            """, (id_inf,))
            detalle = cursor.fetchone()

            if not detalle:
                print("No se encontró el informe solicitado.")
                return

            print("\n" + "="*50)
            print("        DETALLE DE INFORME")
            print("="*50)
            print(f"ID: {detalle[0]}")
            print(f"Nombre: {detalle[1]}")
            print(f"Fecha: {detalle[2]}")
            print(f"Tipo: {detalle[4]}")
            print(f"Descripción: {detalle[3]}")
            print("="*50 + "\n")

        except Exception as e:
            print(f"Error al ver detalle de informe: {e}")

    def eliminar_informe(self, cnx, cursor):
        """
        Elimina un informe del sistema y su registro de la base de datos.
        """
        try:
            print("\n--- ELIMINAR INFORME ---")

            cursor.execute("SELECT id_Informe, NombreInforme, FechaConsulta FROM Informe ORDER BY FechaConsulta DESC")
            informes = cursor.fetchall()

            if not informes:
                print("No hay informes registrados.")
                return

            for inf in informes:
                print(f"{inf[0]} - {inf[1]} ({inf[2]})")

            id_inf = pedir_input("Ingrese el ID del informe a eliminar [0 para cancelar]: ", cnx)
            if id_inf is None or not id_inf.isdigit():
                print("Operación cancelada.")
                return
            id_inf = int(id_inf)

            cursor.execute("SELECT NombreInforme FROM Informe WHERE id_Informe = %s", (id_inf,))
            registro = cursor.fetchone()
            if not registro:
                print("Informe no encontrado.")
                return

            nombre = registro[0]
            ruta = f"./informes/{nombre}"

            print(f"\nADVERTENCIA: Esto eliminará el informe '{nombre}' y su registro.")
            confirmar = pedir_input("Escriba DELETE para confirmar o 0 para cancelar: ", cnx)
            if confirmar is None or confirmar.strip().upper() != "DELETE":
                print("Operación cancelada.")
                return

            # Eliminar archivo físico (si existe)
            import os
            if os.path.exists(ruta):
                os.remove(ruta)
                print("Archivo Excel eliminado correctamente.")
            else:
                print("Advertencia: No se encontró el archivo físico, solo se eliminará el registro.")

            # Eliminar registro de la BD
            cursor.execute("DELETE FROM Informe WHERE id_Informe = %s", (id_inf,))
            cnx.commit()

            print(f"Informe '{nombre}' eliminado correctamente.")

        except Exception as e:
            try:
                cnx.rollback()
            except Exception:
                pass
            print(f"Error al eliminar el informe: {e}")

    def exportar_excel_empleado_proyecto(self, cursor):
        """
        Exporta directamente a Excel los datos de empleados y proyectos,
        sin registrar el informe en la base de datos.
        """
        try:
            print("\n--- EXPORTAR EXCEL DE EMPLEADOS Y PROYECTOS ---")

            print("""
                1. Exportar todos
                2. Exportar por proyecto
                3. Exportar por departamento
                0. Cancelar
                """)
            opcion = pedir_input("Seleccione una opción: ", opcional=True)
            if opcion is None or opcion == "0":
                return

            filtro_sql = ""
            params = ()
            tipo = "General"

            if opcion == "2":  # Por proyecto
                cursor.execute("SELECT id_Proyecto, NombreProyecto FROM Proyecto ORDER BY id_Proyecto")
                proyectos = cursor.fetchall()
                if not proyectos:
                    print("No hay proyectos registrados.")
                    return

                for p in proyectos:
                    print(f"{p[0]} - {p[1]}")
                id_p = pedir_input("Ingrese el ID del proyecto [0 para cancelar]: ", opcional=True)
                if id_p is None or not id_p.isdigit():
                    print("Operación cancelada.")
                    return
                filtro_sql = "WHERE P.id_Proyecto = %s"
                params = (int(id_p),)
                tipo = "Por Proyecto"

            elif opcion == "3":  # Por departamento
                cursor.execute("SELECT id_Departamento, NombreDepartamento FROM Departamento ORDER BY id_Departamento")
                departamentos = cursor.fetchall()
                if not departamentos:
                    print("No hay departamentos registrados.")
                    return

                for d in departamentos:
                    print(f"{d[0]} - {d[1]}")
                id_d = pedir_input("Ingrese el ID del departamento [0 para cancelar]: ", opcional=True)
                if id_d is None or not id_d.isdigit():
                    print("Operación cancelada.")
                    return
                filtro_sql = "WHERE D.id_Departamento = %s"
                params = (int(id_d),)
                tipo = "Por Departamento"

            # Consulta principal
            cursor.execute(f"""
                SELECT 
                    EP.id_DetalleProyecto,
                    CONCAT(U.Nombre, ' ', U.Apellido) AS Empleado,
                    D.NombreDepartamento,
                    P.NombreProyecto,
                    EP.FechaProyectoInscrito,
                    EP.CantidadHorasEmpleadoProyecto,
                    EP.DescripcionTareaProyecto,
                    CASE WHEN EP.Activo THEN 'Activo' ELSE 'Inactivo' END AS Estado
                FROM EmpleadoProyecto EP
                JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
                {filtro_sql}
                ORDER BY D.NombreDepartamento, P.NombreProyecto, U.Nombre
            """, params)
            resultados = cursor.fetchall()

            if not resultados:
                print("No se encontraron registros para exportar.")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Empleados-Proyectos"

            encabezados = [
                "ID Detalle", "Empleado", "Departamento",
                "Proyecto", "Fecha Inscripción",
                "Horas Asignadas", "Descripción", "Estado"
            ]
            ws.append(encabezados)

            for fila in resultados:
                ws.append(list(fila))

            nombre_archivo = f"export_{tipo.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            ruta = f"./informes/{nombre_archivo}"
            wb.save(ruta)

            print(f"\nExportación completada correctamente.")
            print(f"Archivo guardado en: {ruta}")
            print(f"Registros exportados: {len(resultados)}")

        except Exception as e:
            print(f"Error al exportar Excel: {e}")



