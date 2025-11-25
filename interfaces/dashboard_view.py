import tkinter as tk
from tkinter import messagebox
from tkinter import ttk   # Importar librerías útiles de TKinter
import hashlib
from datetime import datetime, date # manipulación de fechas

from tkcalendar import DateEntry  # calendario dinámico

from servicios.economia_service import EconomiaService # Servicios de API

import os
from openpyxl import Workbook

# validadores de clase
from utils import validar_rut, validar_email, validar_telefono, validar_contrasena

# Para gráficos en Tkinter
try:
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
except ImportError:
    FigureCanvasTkAgg = None
    Figure = None




class DashboardView(tk.Frame):
    def __init__(self, master, cnx, cursor, usuario, rol, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.cnx = cnx
        self.cursor = cursor
        self.usuario = usuario      # objeto UsuarioSistema / Admin
        self.rol = rol

        # Service de economía
        self.economia_service = EconomiaService(self.cnx, self.cursor, self.usuario)

        # Labels para mostrar valores económicos (los llenamos en _build_admin_layout)
        self.lbl_usd = None
        self.lbl_uf = None
        self.lbl_eur = None
        self.lbl_fecha_eco = None


        # panel central donde iremos cambiando el contenido
        self.panel_central = None
        self.lbl_contenido = None

        self._build_ui()

    # =========================================================
    # UI GENERAL
    # =========================================================
    def _build_ui(self):
        # ---------- HEADER ----------
        header = tk.Frame(self, bg="#e0e0e0")
        header.pack(fill="x", pady=5)

        titulo = f"Bienvenido {self.usuario.Nombre} {self.usuario.Apellido} | Dashboard {self.rol.capitalize()}"
        lbl_header = tk.Label(header, text=titulo, font=("Arial", 14, "bold"))
        lbl_header.pack(pady=8)

        # ---------- CUERPO PRINCIPAL ----------
        cuerpo = tk.Frame(self)
        cuerpo.pack(fill="both", expand=True, padx=10, pady=10)

        if self.rol == "admin":
            self._build_admin_layout(cuerpo)
        else:
            self._build_empleado_layout(cuerpo)

        # ---------- PIE (cerrar sesión) ----------
        pie = tk.Frame(self)
        pie.pack(fill="x", pady=5)

        btn_logout = tk.Button(pie, text="Cerrar sesión", command=self._logout)
        btn_logout.pack(pady=5)

    # =========================================================
    # LAYOUT ADMIN
    # =========================================================
    def _build_admin_layout(self, parent):
        # columnas: menú izquierda, contenido centro, economía derecha
        parent.columnconfigure(0, weight=0)  # menú
        parent.columnconfigure(1, weight=1)  # contenido
        parent.columnconfigure(2, weight=0)  # economía

        # ---- MENÚ IZQUIERDA ----
        menu = tk.Frame(parent, bd=1, relief="groove")
        menu.grid(row=0, column=0, sticky="nsw", padx=(0, 10))

        tk.Label(menu, text="Opciones Admin", font=("Arial", 11, "bold")).pack(pady=5, padx=5)

        # USUARIOS
        btn_usuarios = tk.Button(menu, text="Usuarios", command=self._show_usuarios)
        btn_usuarios.pack(fill="x", padx=5, pady=2)

         # DEPARTAMENTOS
        btn_dep = tk.Button(
            menu,
            text="Departamentos",
            command=self._show_departamentos
        )
        btn_dep.pack(fill="x", padx=5, pady=2)

        # PROYECTOS
        btn_proj = tk.Button(
            menu,
            text="Proyectos",
            command=self._show_proyectos
        )
        btn_proj.pack(fill="x", padx=5, pady=2)

        # ASIGNACIONES
        btn_asig = tk.Button(
            menu,
            text="Asignaciones",
            command=self._show_asignaciones
        )
        btn_asig.pack(fill="x", padx=5, pady=2)

        # INFORMES
        btn_inf = tk.Button(
            menu,
            text="Informes",
            command=self._show_informes
        )
        btn_inf.pack(fill="x", padx=5, pady=2)


        # HERRAMIENTAS BD
        btn_db = tk.Button(menu, text="Herramientas BD", command=self._show_bd_tools)
        btn_db.pack(fill="x", padx=5, pady=10)

        # ---- PANEL CENTRAL (contenido) ----
        self.panel_central = tk.Frame(parent, bd=1, relief="groove")
        self.panel_central.grid(row=0, column=1, sticky="nsew")
        parent.rowconfigure(0, weight=1)

        self.lbl_contenido = tk.Label(
            self.panel_central,
            text="Selecciona una opción del menú de la izquierda.",
            font=("Arial", 12),
        )
        self.lbl_contenido.pack(padx=10, pady=10, anchor="nw")

        # ---- PANEL DERECHA (Economía) ----
        panel_eco = tk.Frame(parent, bd=1, relief="groove", width=240)
        panel_eco.grid(row=0, column=2, sticky="ns", padx=(10, 0))

        tk.Label(panel_eco, text="Economía", font=("Arial", 11, "bold")).pack(pady=5)

        # Indicadores principales
        self.lbl_usd = tk.Label(panel_eco, text="USD/CLP (dólar): —", anchor="w")
        self.lbl_usd.pack(fill="x", padx=8, pady=2)

        self.lbl_usd_inter = tk.Label(panel_eco, text="USD Acuerdo: —", anchor="w")
        self.lbl_usd_inter.pack(fill="x", padx=8, pady=2)

        self.lbl_eur = tk.Label(panel_eco, text="EUR/CLP: —", anchor="w")
        self.lbl_eur.pack(fill="x", padx=8, pady=2)

        self.lbl_uf = tk.Label(panel_eco, text="UF (CLP): —", anchor="w")
        self.lbl_uf.pack(fill="x", padx=8, pady=2)

        self.lbl_utm = tk.Label(panel_eco, text="UTM (CLP): —", anchor="w")
        self.lbl_utm.pack(fill="x", padx=8, pady=2)

        self.lbl_ipc = tk.Label(panel_eco, text="IPC (%): —", anchor="w")
        self.lbl_ipc.pack(fill="x", padx=8, pady=2)

        self.lbl_ipsa = tk.Label(panel_eco, text="IPSA (pts): —", anchor="w")
        self.lbl_ipsa.pack(fill="x", padx=8, pady=2)

        self.lbl_fecha_eco = tk.Label(
            panel_eco,
            text="Última actualización: —",
            font=("Arial", 8),
            anchor="w",
            justify="left"
        )
        self.lbl_fecha_eco.pack(fill="x", padx=8, pady=(8, 4))

        # Botones de acción
        tk.Button(
            panel_eco,
            text="Actualizar ahora",
            command=self._actualizar_economia_desde_api
        ).pack(fill="x", padx=8, pady=(5, 2))

        tk.Button(
            panel_eco,
            text="Ver historial…",
            command=self._ver_historial_economia_gui
        ).pack(fill="x", padx=8, pady=(0, 4))

        tk.Button(
            panel_eco,
            text="Ver gráfico…",
            command=self._ver_grafico_economia_gui
        ).pack(fill="x", padx=8, pady=(0, 4))

        tk.Button(
            panel_eco,
            text="Exportar historial…",
            command=self._exportar_historial_economia_gui
        ).pack(fill="x", padx=8, pady=(0, 8))

        # Cargar valores iniciales desde la BD (si existen)
        self._refrescar_panel_economia()


    # =========================================================
    # ECONOMÍA
    # =========================================================

    def _refrescar_panel_economia(self):
        """Lee los últimos valores guardados y los muestra en el panel derecho."""
        try:
            datos = self.economia_service.obtener_ultimos_valores()
        except Exception as e:
            print(f"[Dashboard] Error obteniendo valores económicos: {e}")
            datos = {}

        def set_label(lbl_name, codigo, fmt_texto):
            """
            Helper interno:
            - lbl_name: nombre del atributo (self.lbl_usd, etc.)
            - codigo: CódigoIndicador en la BD (USD_CLP, IPC, etc.)
            - fmt_texto: cadena con un {: .2f} para el valor
            """
            lbl = getattr(self, lbl_name, None)
            if not lbl:
                return

            info = datos.get(codigo)
            if info:
                try:
                    texto = fmt_texto.format(info["valor"])
                except Exception:
                    texto = fmt_texto.format(0.0)
                lbl.config(text=texto)
            else:
                base = fmt_texto.split(":")[0]  # parte antes de los dos puntos
                lbl.config(text=f"{base}: —")

        # Mapeo a nuestros indicadores
        set_label("lbl_usd",        "USD_CLP",             "USD/CLP (dólar): {:.2f}")
        set_label("lbl_usd_inter",  "USD_INTERCAMBIO_CLP", "USD Acuerdo: {:.2f}")
        set_label("lbl_eur",        "EUR_CLP",             "EUR/CLP: {:.2f}")
        set_label("lbl_uf",         "UF_CLP",              "UF (CLP): {:.2f}")
        set_label("lbl_utm",        "UTM_CLP",             "UTM (CLP): {:.2f}")
        set_label("lbl_ipc",        "IPC",                 "IPC (%): {:.2f}")
        set_label("lbl_ipsa",       "IPSA",                "IPSA (pts): {:.2f}")

        # Fecha de la última actualización (la más reciente entre todos)
        if hasattr(self, "lbl_fecha_eco") and self.lbl_fecha_eco:
            fechas = [
                info["fecha"]
                for info in datos.values()
                if info.get("fecha") is not None
            ]
            if fechas:
                ultima = max(fechas)
                self.lbl_fecha_eco.config(
                    text=f"Última actualización: {ultima.strftime('%Y-%m-%d %H:%M')}"
                )
            else:
                self.lbl_fecha_eco.config(text="Última actualización: —")

    def _actualizar_economia_desde_api(self):
        """
        Llama al service para consultar la API pública y guardar
        los valores, luego refresca el panel.
        """
        try:
            resultados = self.economia_service.actualizar_desde_api()
            if resultados:
                messagebox.showinfo(
                    "Economía",
                    "Indicadores actualizados correctamente desde la API."
                )
            else:
                messagebox.showwarning(
                    "Economía",
                    "No se pudieron actualizar los indicadores."
                )
        except Exception as e:
            messagebox.showerror(
                "Economía",
                f"Ocurrió un error al consultar la API:\n{e}"
            )

        # En cualquier caso, intentamos refrescar la vista
        self._refrescar_panel_economia()
 
    def _ver_historial_economia_gui(self):
        """
        Muestra un historial de consultas económicas con filtros
        por fecha, indicador y (si es admin) por usuario.
        """
        win = tk.Toplevel(self)
        win.title("Historial de consultas económicas")
        win.geometry("800x450")
        win.grab_set()

        # ---------- FRAME DE FILTROS ----------
        filtros = tk.Frame(win)
        filtros.pack(fill="x", padx=10, pady=5)

        tk.Label(filtros, text="Desde:").grid(row=0, column=0, padx=5, pady=2, sticky="e")
        if DateEntry:
            entry_desde = DateEntry(filtros, width=12, date_pattern="yyyy-mm-dd")
        else:
            entry_desde = tk.Entry(filtros, width=12)
        entry_desde.grid(row=0, column=1, padx=5, pady=2, sticky="w")

        tk.Label(filtros, text="Hasta:").grid(row=0, column=2, padx=5, pady=2, sticky="e")
        if DateEntry:
            entry_hasta = DateEntry(filtros, width=12, date_pattern="yyyy-mm-dd")
        else:
            entry_hasta = tk.Entry(filtros, width=12)
        entry_hasta.grid(row=0, column=3, padx=5, pady=2, sticky="w")


        tk.Label(filtros, text="Indicador:").grid(row=0, column=4, padx=5, pady=2, sticky="e")
        valores_ind = ["Todos"] + list(self.economia_service.INDICADORES_API.keys())
        cb_ind = ttk.Combobox(
            filtros,
            state="readonly",
            width=20,
            values=valores_ind
        )
        cb_ind.grid(row=0, column=5, padx=5, pady=2, sticky="w")
        cb_ind.current(0)

        # Solo admin puede filtrar por usuario
        entry_usuario = None
        if self.rol == "admin":
            tk.Label(filtros, text="Usuario:").grid(row=1, column=0, padx=5, pady=2, sticky="e")
            entry_usuario = tk.Entry(filtros, width=15)
            entry_usuario.grid(row=1, column=1, padx=5, pady=2, sticky="w")

        # Botones filtros
        def aplicar_filtro():
            desde = entry_desde.get().strip() or None
            hasta = entry_hasta.get().strip() or None
            indicador = cb_ind.get()
            if indicador == "Todos":
                indicador = None

            usuario_filtro = None
            if entry_usuario is not None:
                usuario_filtro = entry_usuario.get().strip() or None

            # Validar fechas
            from datetime import datetime
            try:
                if desde:
                    datetime.strptime(desde, "%Y-%m-%d")
                if hasta:
                    datetime.strptime(hasta, "%Y-%m-%d")
            except ValueError:
                messagebox.showwarning(
                    "Formato de fecha",
                    "Use el formato YYYY-MM-DD (ej: 2025-11-25)."
                )
                return

            cargar_datos(desde, hasta, indicador, usuario_filtro)

        def limpiar_filtro():
            entry_desde.delete(0, tk.END)
            entry_hasta.delete(0, tk.END)
            cb_ind.current(0)
            if entry_usuario is not None:
                entry_usuario.delete(0, tk.END)
            cargar_datos(None, None, None, None)

        tk.Button(filtros, text="Filtrar", command=aplicar_filtro)\
          .grid(row=1, column=4, padx=5, pady=4, sticky="e")
        tk.Button(filtros, text="Limpiar", command=limpiar_filtro)\
          .grid(row=1, column=5, padx=5, pady=4, sticky="w")

        # ---------- FRAME TABLA + SCROLL ----------
        frame_tabla = tk.Frame(win)
        frame_tabla.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("fecha", "codigo", "nombre", "valor", "usuario")
        tree = ttk.Treeview(frame_tabla, columns=cols, show="headings")
        tree.pack(side="left", fill="both", expand=True)

        headers = ["Fecha", "Código", "Indicador", "Valor", "Usuario"]
        for col, h in zip(cols, headers):
            tree.heading(col, text=h)
            tree.column(col, anchor="w")

        scrollbar = ttk.Scrollbar(frame_tabla, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)

        # ---------- CARGA DE DATOS ----------
        def cargar_datos(fecha_desde, fecha_hasta, codigo_indicador, usuario_filtro):
            # Limpiar tabla
            for item in tree.get_children():
                tree.delete(item)

            where = []
            params = []

            # Filtro por rol: si es empleado, solo sus consultas
            if self.rol != "admin":
                where.append("c.UsuarioSistema_idUsuarioSistema = %s")
                params.append(self.usuario.id_UsuarioSistema)

            # Filtros dinámicos
            if fecha_desde:
                where.append("c.FechaConsulta >= %s")
                params.append(f"{fecha_desde} 00:00:00")
            if fecha_hasta:
                where.append("c.FechaConsulta <= %s")
                params.append(f"{fecha_hasta} 23:59:59")
            if codigo_indicador:
                where.append("i.CodigoIndicador = %s")
                params.append(codigo_indicador)
            if usuario_filtro:
                where.append("u.NombreUsuario = %s")
                params.append(usuario_filtro)

            where_sql = ""
            if where:
                where_sql = "WHERE " + " AND ".join(where)

            try:
                self.cursor.execute(
                    f"""
                    SELECT 
                        c.FechaConsulta,
                        i.CodigoIndicador,
                        i.NombreIndicador,
                        c.ValorPrincipal,
                        u.NombreUsuario
                    FROM ConsultaEconomica c
                    JOIN IndicadorEconomico i
                        ON c.IndicadorEconomico_idIndicadorEconomico = i.id_IndicadorEconomico
                    JOIN UsuarioSistema u
                        ON c.UsuarioSistema_idUsuarioSistema = u.id_UsuarioSistema
                    {where_sql}
                    ORDER BY c.FechaConsulta DESC
                    """,
                    tuple(params)
                )
                filas = self.cursor.fetchall()

                for f in filas:
                    fecha, codigo, nombre, valor, usuario = f
                    fecha_str = fecha.strftime("%Y-%m-%d %H:%M") if fecha else ""
                    tree.insert(
                        "",
                        "end",
                        values=(fecha_str, codigo, nombre, f"{float(valor):.2f}", usuario),
                    )

            except Exception as e:
                messagebox.showerror(
                    "Economía",
                    f"No se pudo cargar el historial:\n{e}"
                )

        # Carga inicial (sin filtros)
        cargar_datos(None, None, None, None)


    def _exportar_historial_economia_gui(self):
        """
        Ventana para exportar el historial económico a Excel,
        con filtros simples (fechas + indicador).
        """
        win = tk.Toplevel(self)
        win.title("Exportar historial económico")
        win.geometry("350x220")
        win.grab_set()

        tk.Label(win, text="Fecha desde:").grid(row=0, column=0, padx=8, pady=5, sticky="e")
        if DateEntry:
            entry_desde = DateEntry(win, width=12, date_pattern="yyyy-mm-dd")
        else:
            entry_desde = tk.Entry(win, width=15)
        entry_desde.grid(row=0, column=1, padx=8, pady=5, sticky="w")

        tk.Label(win, text="Fecha hasta:").grid(row=1, column=0, padx=8, pady=5, sticky="e")
        if DateEntry:
            entry_hasta = DateEntry(win, width=12, date_pattern="yyyy-mm-dd")
        else:
            entry_hasta = tk.Entry(win, width=15)
        entry_hasta.grid(row=1, column=1, padx=8, pady=5, sticky="w")

        tk.Label(win, text="Indicador:").grid(row=2, column=0, padx=8, pady=5, sticky="e")
        cb_ind = ttk.Combobox(
            win,
            state="readonly",
            width=22,
            values=["Todos"] + list(self.economia_service.INDICADORES_API.keys())
        )
        
        cb_ind.grid(row=2, column=1, padx=8, pady=5, sticky="w")
        cb_ind.current(0)  # "Todos"

                # ==== FECHAS DINÁMICAS DESDE LA BD ====
        try:
            self.cursor.execute("""
                SELECT 
                    MIN(FechaConsulta),
                    MAX(FechaConsulta)
                FROM ConsultaEconomica
            """)
            row = self.cursor.fetchone()
            if row and row[0] and row[1]:
                fecha_min, fecha_max = row

                if DateEntry and hasattr(entry_desde, "set_date"):
                    entry_desde.set_date(fecha_min)
                    entry_hasta.set_date(fecha_max)
                else:
                    entry_desde.insert(0, fecha_min.strftime("%Y-%m-%d"))
                    entry_hasta.insert(0, fecha_max.strftime("%Y-%m-%d"))
        except Exception as e:
            print("[Exportar] No se pudieron precargar fechas:", e)

        def actualizar_fechas_por_indicador(event=None):
            cod = cb_ind.get()
            if cod == "Todos":
                query = """
                    SELECT MIN(FechaConsulta), MAX(FechaConsulta)
                    FROM ConsultaEconomica
                """
                params = ()
            else:
                query = """
                    SELECT MIN(c.FechaConsulta), MAX(c.FechaConsulta)
                    FROM ConsultaEconomica c
                    JOIN IndicadorEconomico i
                        ON c.IndicadorEconomico_idIndicadorEconomico = i.id_IndicadorEconomico
                    WHERE i.CodigoIndicador = %s
                """
                params = (cod,)

            try:
                self.cursor.execute(query, params)
                row = self.cursor.fetchone()
                if row and row[0] and row[1]:
                    fecha_min, fecha_max = row

                    if DateEntry and hasattr(entry_desde, "set_date"):
                        # 👉 usar el calendario
                        entry_desde.set_date(fecha_min)
                        entry_hasta.set_date(fecha_max)
                    else:
                        # fallback a Entry normal
                        entry_desde.delete(0, tk.END)
                        entry_hasta.delete(0, tk.END)
                        entry_desde.insert(0, fecha_min.strftime("%Y-%m-%d"))
                        entry_hasta.insert(0, fecha_max.strftime("%Y-%m-%d"))
            except:
                pass


        cb_ind.bind("<<ComboboxSelected>>", actualizar_fechas_por_indicador)

        def exportar():

            fecha_desde = entry_desde.get().strip() or None
            fecha_hasta = entry_hasta.get().strip() or None
            indicador = cb_ind.get()
            if indicador == "Todos":
                indicador = None  # sin filtro

            # Validar formato de fechas si se ingresan
            try:
                if fecha_desde:
                    datetime.strptime(fecha_desde, "%Y-%m-%d")
                if fecha_hasta:
                    datetime.strptime(fecha_hasta, "%Y-%m-%d")
            except ValueError:
                messagebox.showwarning(
                    "Formato de fecha",
                    "Use el formato YYYY-MM-DD (ej: 2025-11-22)."
                )
                return

            try:
                ruta = self.economia_service.exportar_historial_excel(
                    fecha_desde=fecha_desde,
                    fecha_hasta=fecha_hasta,
                    codigo_indicador=indicador
                )
                if ruta:
                    messagebox.showinfo(
                        "Exportación completada",
                        f"Historial exportado correctamente en:\n{ruta}"
                    )
                    win.destroy()
                else:
                    messagebox.showwarning(
                        "Sin datos",
                        "No se encontraron registros con esos filtros."
                    )
            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Ocurrió un error al exportar:\n{e}"
                )

        tk.Button(win, text="Exportar", command=exportar)\
            .grid(row=3, column=1, padx=8, pady=15, sticky="e")

    def _ver_grafico_economia_gui(self):
        """
        Ventana para visualizar un gráfico de evolución de un indicador
        en un rango de fechas dinámico y exportar esa serie a Excel
        (con gráfico incluido).
        """
        import matplotlib
        matplotlib.use("TkAgg")
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from datetime import datetime

        win = tk.Toplevel(self)
        win.title("Gráfico de indicador económico")
        win.geometry("900x500")
        win.grab_set()

        frm_filtros = tk.Frame(win)
        frm_filtros.pack(fill="x", padx=8, pady=4)

        # --- Controles de filtro ---
        tk.Label(frm_filtros, text="Indicador:").grid(row=0, column=0, padx=4, pady=2, sticky="e")
        cb_ind = ttk.Combobox(
            frm_filtros,
            state="readonly",
            width=22,
            values=["UF_CLP", "IVP", "IPC", "UTM_CLP", "USD_CLP", "EUR_CLP"]
        )
        cb_ind.grid(row=0, column=1, padx=4, pady=2, sticky="w")
        cb_ind.current(0)

        tk.Label(frm_filtros, text="Desde:").grid(row=1, column=0, padx=4, pady=2, sticky="e")
        if DateEntry:
            entry_desde = DateEntry(frm_filtros, width=12, date_pattern="yyyy-mm-dd")
        else:
            entry_desde = tk.Entry(frm_filtros, width=12)
        entry_desde.grid(row=1, column=1, padx=4, pady=2, sticky="w")

        tk.Label(frm_filtros, text="Hasta:").grid(row=1, column=2, padx=4, pady=2, sticky="e")
        if DateEntry:
            entry_hasta = DateEntry(frm_filtros, width=12, date_pattern="yyyy-mm-dd")
        else:
            entry_hasta = tk.Entry(frm_filtros, width=12)
        entry_hasta.grid(row=1, column=3, padx=4, pady=2, sticky="w")


        # Para admin: opción "solo mis consultas"
        solo_mias_var = tk.BooleanVar(value=False)
        if self.rol == "admin":
            chk = tk.Checkbutton(
                frm_filtros,
                text="Solo mis consultas",
                variable=solo_mias_var
            )
            chk.grid(row=0, column=3, padx=4, pady=2, sticky="w")

        # Precargar rango global al inicio (todas las consultas)
        try:
            self.cursor.execute("""
                SELECT MIN(FechaConsulta), MAX(FechaConsulta)
                FROM ConsultaEconomica
            """)
            row = self.cursor.fetchone()
            if row and row[0] and row[1]:
                f_min, f_max = row
                if DateEntry and hasattr(entry_desde, "set_date"):
                    entry_desde.set_date(f_min)
                    entry_hasta.set_date(f_max)
                else:
                    entry_desde.insert(0, f_min.strftime("%Y-%m-%d"))
                    entry_hasta.insert(0, f_max.strftime("%Y-%m-%d"))
        except Exception as e:
            print("[Grafico] No se pudieron precargar fechas:", e)

        # --- Frame para el gráfico ---
        frm_grafico = tk.Frame(win, bd=1, relief="sunken")
        frm_grafico.pack(fill="both", expand=True, padx=8, pady=8)

        canvas_mpl = None  # lo rellenaremos luego

        def dibujar_grafico():
            nonlocal canvas_mpl

            cod = cb_ind.get()
            fecha_desde = entry_desde.get().strip() or None
            fecha_hasta = entry_hasta.get().strip() or None

            # Validar formato de fechas
            try:
                if fecha_desde:
                    datetime.strptime(fecha_desde, "%Y-%m-%d")
                if fecha_hasta:
                    datetime.strptime(fecha_hasta, "%Y-%m-%d")
            except ValueError:
                messagebox.showwarning(
                    "Fechas",
                    "Use el formato YYYY-MM-DD en ambas fechas."
                )
                return

            solo_usuario_id = None
            if solo_mias_var.get():
                solo_usuario_id = self.usuario.id_UsuarioSistema

            try:
                # serie = self.economia_service.obtener_serie(
                #     codigo_indicador=cod,
                #     fecha_desde=fecha_desde,
                #     fecha_hasta=fecha_hasta,
                #     solo_usuario_id=solo_usuario_id
                # )


                serie = self.economia_service.consultar_indicador_periodo_api(
                    codigo_indicador=cod,
                    fecha_desde=fecha_desde or None,
                    fecha_hasta=fecha_hasta or None,
                    registrar_en_bd=True   # si quieres que lo deje guardado
                )

            except Exception as e:
                messagebox.showerror(
                    "Gráfico",
                    f"Error al obtener datos:\n{e}"
                )
                return

            if not serie:
                messagebox.showinfo(
                    "Gráfico",
                    "No hay datos en ese rango de fechas para el indicador seleccionado."
                )
                return

            # Destruir gráfico anterior si lo hubiera
            if canvas_mpl is not None:
                canvas_mpl.get_tk_widget().destroy()
                canvas_mpl = None

            # Construir nuevo gráfico
            fig = Figure(figsize=(7, 3.8), dpi=100)
            ax = fig.add_subplot(111)

            fechas = [p[0] for p in serie]
            valores = [p[1] for p in serie]

            ax.plot(fechas, valores, marker="o", linestyle="-")
            ax.set_title(f"Evolución {cod}")
            ax.set_xlabel("Fecha")
            ax.set_ylabel("Valor")
            fig.autofmt_xdate(rotation=45)

            canvas_mpl = FigureCanvasTkAgg(fig, master=frm_grafico)
            canvas_mpl.draw()
            canvas_mpl.get_tk_widget().pack(fill="both", expand=True)

        def exportar_excel_con_grafico():
            cod = cb_ind.get()
            fecha_desde = entry_desde.get().strip() or None
            fecha_hasta = entry_hasta.get().strip() or None

            try:
                if fecha_desde:
                    datetime.strptime(fecha_desde, "%Y-%m-%d")
                if fecha_hasta:
                    datetime.strptime(fecha_hasta, "%Y-%m-%d")
            except ValueError:
                messagebox.showwarning(
                    "Fechas",
                    "Use el formato YYYY-MM-DD en ambas fechas."
                )
                return

            solo_usuario_id = None
            if solo_mias_var.get():
                solo_usuario_id = self.usuario.id_UsuarioSistema

            try:
                ruta = self.economia_service.exportar_historial_excel(
                    fecha_desde=fecha_desde,
                    fecha_hasta=fecha_hasta,
                    codigo_indicador=cod,
                    incluir_grafico=True,
                    solo_usuario_id=solo_usuario_id
                )
                if ruta:
                    messagebox.showinfo(
                        "Exportación",
                        f"Archivo Excel generado con datos y gráfico en:\n{ruta}"
                    )
                else:
                    messagebox.showwarning(
                        "Exportación",
                        "No se encontraron registros para exportar."
                    )
            except Exception as e:
                messagebox.showerror(
                    "Exportación",
                    f"Error al exportar gráfico:\n{e}"
                )

        # Botones de acción
        frm_botones = tk.Frame(win)
        frm_botones.pack(fill="x", padx=8, pady=4)

        tk.Button(frm_botones, text="Mostrar gráfico", command=dibujar_grafico)\
            .pack(side="left", padx=4)

        tk.Button(frm_botones, text="Exportar a Excel", command=exportar_excel_con_grafico)\
            .pack(side="left", padx=4)


    # =========================================================
    # LAYOUT EMPLEADO
    # =========================================================

    def _build_empleado_layout(self, parent):
        # columnas: menú izquierda, contenido centro, economía derecha
        parent.columnconfigure(0, weight=0)  # menú
        parent.columnconfigure(1, weight=1)  # contenido
        parent.columnconfigure(2, weight=0)  # economía

        # ---- MENÚ IZQUIERDA ----
        menu = tk.Frame(parent, bd=1, relief="groove")
        menu.grid(row=0, column=0, sticky="nsw", padx=(0, 10))

        tk.Label(menu, text="Opciones Empleado", font=("Arial", 11, "bold")).pack(pady=5)

        opciones = [
            ("Mis datos personales",    self._emp_mis_datos),
            ("Detalle laboral",         self._emp_detalle_laboral),
            ("Proyectos activos",       self._emp_proyectos_activos),
            ("Historial de proyectos",  self._emp_historial_proyectos),
            ("Compañeros",              self._emp_companeros),
            ("Editar datos personales", self._emp_editar_datos),
            ("Cambiar usuario",         self._emp_cambiar_usuario),
            ("Cambiar contraseña",      self._emp_cambiar_contrasena),
        ]

        for texto, cmd in opciones:
            tk.Button(menu, text=texto, command=cmd).pack(fill="x", padx=5, pady=2)

        # ---- PANEL CENTRAL (contenido) ----
        self.panel_central = tk.Frame(parent, bd=1, relief="groove")
        self.panel_central.grid(row=0, column=1, sticky="nsew")
        parent.rowconfigure(0, weight=1)

        self.lbl_contenido = tk.Label(
            self.panel_central,
            text="Selecciona una opción del menú de la izquierda.",
            font=("Arial", 12),
        )
        self.lbl_contenido.pack(padx=10, pady=10, anchor="nw")

        # ---- PANEL DERECHA (Economía) ----
        panel_eco = tk.Frame(parent, bd=1, relief="groove", width=240)
        panel_eco.grid(row=0, column=2, sticky="ns", padx=(10, 0))

        tk.Label(panel_eco, text="Economía", font=("Arial", 11, "bold")).pack(pady=5)

        self.lbl_usd = tk.Label(panel_eco, text="USD/CLP (dólar): —", anchor="w")
        self.lbl_usd.pack(fill="x", padx=8, pady=2)

        self.lbl_usd_inter = tk.Label(panel_eco, text="USD Acuerdo: —", anchor="w")
        self.lbl_usd_inter.pack(fill="x", padx=8, pady=2)

        self.lbl_eur = tk.Label(panel_eco, text="EUR/CLP: —", anchor="w")
        self.lbl_eur.pack(fill="x", padx=8, pady=2)

        self.lbl_uf = tk.Label(panel_eco, text="UF (CLP): —", anchor="w")
        self.lbl_uf.pack(fill="x", padx=8, pady=2)

        self.lbl_utm = tk.Label(panel_eco, text="UTM (CLP): —", anchor="w")
        self.lbl_utm.pack(fill="x", padx=8, pady=2)

        self.lbl_ipc = tk.Label(panel_eco, text="IPC (%): —", anchor="w")
        self.lbl_ipc.pack(fill="x", padx=8, pady=2)

        self.lbl_ipsa = tk.Label(panel_eco, text="IPSA (pts): —", anchor="w")
        self.lbl_ipsa.pack(fill="x", padx=8, pady=2)

        self.lbl_fecha_eco = tk.Label(
            panel_eco,
            text="Última actualización: —",
            font=("Arial", 8),
            anchor="w"
        )
        self.lbl_fecha_eco.pack(fill="x", padx=8, pady=(8, 4))

        tk.Button(
            panel_eco,
            text="Actualizar ahora",
            command=self._actualizar_economia_desde_api
        ).pack(fill="x", padx=8, pady=(5, 2))

        tk.Button(
            panel_eco,
            text="Ver historial…",
            command=self._ver_historial_economia_gui
        ).pack(fill="x", padx=8, pady=(0, 2))

        tk.Button(
            panel_eco,
            text="Ver gráfico…",
            command=self._ver_grafico_economia_gui
        ).pack(fill="x", padx=8, pady=(0, 2))

        tk.Button(
            panel_eco,
            text="Exportar historial…",
            command=self._exportar_historial_economia_gui
        ).pack(fill="x", padx=8, pady=(0, 8))

        # Cargar valores iniciales
        self._refrescar_panel_economia()







    # =========================================================
    # FUNCIONES EMPLEADO
    # =========================================================

    def _clear_central(self):
        """Limpia el panel central (se usa tanto por admin como por empleado)."""
        if self.panel_central is None:
            return
        for w in self.panel_central.winfo_children():
            w.destroy()

    # ---------- 1) Mis datos personales ----------
    def _emp_mis_datos(self):
        """Muestra los datos personales del empleado actual en el panel central."""
        self._clear_central()

        cont = tk.Frame(self.panel_central)
        cont.pack(fill="both", expand=True, padx=10, pady=10, anchor="nw")

        tk.Label(
            cont,
            text="Mis datos personales",
            font=("Arial", 13, "bold")
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))

        try:
            self.cursor.execute("""
                SELECT 
                    U.Nombre, U.Apellido, U.RUT, U.Email, U.Telefono, U.NombreUsuario,
                    D.Calle, D.Numero, D.Ciudad, D.Region, D.Pais, D.CodigoPostal
                FROM UsuarioSistema U
                LEFT JOIN Direccion D ON U.Direccion_idDireccion = D.id_Direccion
                WHERE U.id_UsuarioSistema = %s
            """, (self.usuario.id_UsuarioSistema,))
            u = self.cursor.fetchone()

            if not u:
                tk.Label(cont, text="No se encontraron datos del usuario actual.").grid(
                    row=1, column=0, columnspan=2, sticky="w"
                )
                return

            (nombre, apellido, rut, email, telefono, nombre_usuario,
             calle, numero, ciudad, region, pais, cp) = u

            calle  = calle  if calle  is not None else "(sin calle)"
            numero = numero if numero is not None else "-"
            ciudad = ciudad if ciudad is not None else "(sin ciudad)"
            region = region if region is not None else "(sin región)"
            pais   = pais   if pais   is not None else "(sin país)"
            cp     = cp     if cp     is not None else "(sin CP)"
            tel_mostrar = telefono if telefono else "(No registrado)"

            datos = [
                ("Nombre completo:", f"{nombre} {apellido}"),
                ("RUT:",             rut),
                ("Email:",           email),
                ("Teléfono:",        tel_mostrar),
                ("Usuario:",         nombre_usuario),
                ("Calle:",           f"{calle} {numero}"),
                ("Ciudad / Región:", f"{ciudad}, {region}"),
                ("País:",            pais),
                ("Código Postal:",   cp),
            ]

            for i, (label, valor) in enumerate(datos, start=1):
                tk.Label(cont, text=label, font=("Arial", 10, "bold")).grid(
                    row=i, column=0, sticky="e", pady=2, padx=5
                )
                tk.Label(cont, text=valor, font=("Arial", 10)).grid(
                    row=i, column=1, sticky="w", pady=2, padx=5
                )

        except Exception as e:
            tk.Label(
                cont,
                text=f"Ocurrió un error al obtener los datos personales:\n{e}",
                fg="red"
            ).grid(row=1, column=0, columnspan=2, sticky="w")


    # ---------- 2) Detalle laboral ----------

    def _emp_detalle_laboral(self):
        """Muestra el estado laboral actual del empleado."""
        self._clear_central()
        cont = self.panel_central

        tk.Label(cont, text="Detalle laboral", font=("Arial", 12, "bold"))\
            .grid(row=0, column=0, sticky="w", padx=10, pady=(10, 5))

        try:
            self.cursor.execute("""
                SELECT 
                    D.FechaContrato,
                    D.Salario,
                    D.TipoEmpleado,
                    Dep.NombreDepartamento,
                    CASE WHEN ED.Activo = TRUE THEN 'En servicio' ELSE 'Desvinculado' END AS Estado
                FROM DetalleEmpleado D
                JOIN EmpleadoDepartamento ED ON D.UsuarioSistema_idUsuarioSistema = ED.UsuarioSistema_idUsuarioSistema
                JOIN Departamento Dep ON ED.Departamento_idDepartamento = Dep.id_Departamento
                WHERE D.UsuarioSistema_idUsuarioSistema = %s
                ORDER BY ED.FechaAsignacion DESC
                LIMIT 1
            """, (self.usuario.id_UsuarioSistema,))
            detalle = self.cursor.fetchone()
        except Exception as e:
            tk.Label(cont, text=f"Error al obtener detalle laboral: {e}", fg="red")\
                .grid(row=1, column=0, sticky="w", padx=10, pady=5)
            return

        if not detalle:
            tk.Label(cont, text="No tienes historial laboral registrado.")\
                .grid(row=1, column=0, sticky="w", padx=10, pady=5)
            return

        fecha_contrato = detalle[0] or "(sin fecha)"
        salario_val = detalle[1] if detalle[1] is not None else 0
        tipo_emp = detalle[2] or "No especificado"
        depto = detalle[3] or "(sin departamento)"
        estado = detalle[4] or "(desconocido)"

        salario_fmt = f"{salario_val:,}".replace(",", ".")

        info = f"""
    Estado:              {estado}
    Departamento actual: {depto}
    Cargo / Tipo:        {tipo_emp}
    Salario:             ${salario_fmt}
    Fecha de Contrato:   {fecha_contrato}
    """.strip()

        txt = tk.Text(cont, width=70, height=8)
        txt.grid(row=1, column=0, padx=10, pady=5, sticky="nw")
        txt.insert("1.0", info)
        txt.config(state="disabled")

    # ---------- 3) Proyectos activos ----------

    def _emp_proyectos_activos(self):
        """Lista los proyectos activos del empleado."""
        self._clear_central()
        cont = self.panel_central

        tk.Label(cont, text="Proyectos activos", font=("Arial", 12, "bold"))\
            .pack(anchor="w", padx=10, pady=(10, 5))

        cols = ("proyecto", "fecha", "horas", "tarea")
        tree = ttk.Treeview(cont, columns=cols, show="headings")
        tree.pack(fill="both", expand=True, padx=10, pady=5)

        headers = ["Proyecto", "Fecha inscripción", "Horas asignadas", "Descripción tarea"]
        widths = [200, 130, 120, 400]

        for col, h, w in zip(cols, headers, widths):
            tree.heading(col, text=h)
            tree.column(col, width=w, anchor="w")

        scrollbar = ttk.Scrollbar(cont, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)

        try:
            self.cursor.execute("""
                SELECT 
                    P.NombreProyecto,
                    EP.FechaProyectoInscrito,
                    EP.CantidadHorasEmpleadoProyecto,
                    EP.DescripcionTareaProyecto
                FROM EmpleadoProyecto EP
                JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                WHERE ED.UsuarioSistema_idUsuarioSistema = %s
                AND EP.Activo = TRUE
                ORDER BY EP.FechaProyectoInscrito DESC
            """, (self.usuario.id_UsuarioSistema,))
            proyectos = self.cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Proyectos", f"Ocurrió un error al consultar: {e}")
            return

        if not proyectos:
            messagebox.showinfo("Proyectos", "Actualmente no estás asignado a ningún proyecto activo.")
            return

        for p in proyectos:
            nombre = p[0] or "(Sin nombre)"
            fecha = p[1]
            fecha_str = fecha.strftime("%Y-%m-%d") if fecha else "(Sin fecha)"
            horas = p[2] if p[2] is not None else 0
            tarea = p[3] or "(Sin descripción)"
            tree.insert("", "end", values=(nombre, fecha_str, horas, tarea))

    # ---------- 4) Historial de proyectos ----------

    def _emp_historial_proyectos(self):
        """Lista todo el historial de proyectos del empleado."""
        self._clear_central()
        cont = self.panel_central

        tk.Label(cont, text="Historial de proyectos", font=("Arial", 12, "bold"))\
            .pack(anchor="w", padx=10, pady=(10, 5))

        cols = ("proyecto", "fecha", "horas", "tarea", "estado")
        tree = ttk.Treeview(cont, columns=cols, show="headings")
        tree.pack(fill="both", expand=True, padx=10, pady=5)

        headers = ["Proyecto", "Fecha inscripción", "Horas", "Descripción tarea", "Estado"]
        widths = [200, 130, 80, 400, 100]

        for col, h, w in zip(cols, headers, widths):
            tree.heading(col, text=h)
            tree.column(col, width=w, anchor="w")

        scrollbar = ttk.Scrollbar(cont, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)

        try:
            self.cursor.execute("""
                SELECT 
                    P.NombreProyecto,
                    EP.FechaProyectoInscrito,
                    EP.CantidadHorasEmpleadoProyecto,
                    EP.DescripcionTareaProyecto,
                    CASE WHEN EP.Activo = TRUE THEN 'Activo' ELSE 'Inactivo' END AS Estado
                FROM EmpleadoProyecto EP
                JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                WHERE ED.UsuarioSistema_idUsuarioSistema = %s
                ORDER BY EP.FechaProyectoInscrito DESC
            """, (self.usuario.id_UsuarioSistema,))
            proyectos = self.cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Proyectos", f"Ocurrió un error al consultar: {e}")
            return

        if not proyectos:
            messagebox.showinfo("Proyectos", "No existe ningún historial de proyectos para tu usuario.")
            return

        for p in proyectos:
            nombre = p[0] or "(Sin nombre)"
            fecha = p[1]
            fecha_str = fecha.strftime("%Y-%m-%d") if fecha else "(Sin fecha)"
            horas = p[2] if p[2] is not None else 0
            tarea = p[3] or "(Sin descripción)"
            estado = p[4] or "(Desconocido)"
            tree.insert("", "end", values=(nombre, fecha_str, horas, tarea, estado))

    # ---------- 5) Compañeros ----------

    def _emp_companeros(self):
        """Muestra compañeros del mismo departamento (si los hay)."""
        self._clear_central()
        cont = self.panel_central

        tk.Label(cont, text="Compañeros de departamento", font=("Arial", 12, "bold"))\
            .pack(anchor="w", padx=10, pady=(10, 5))

        cols = ("nombre", "email", "telefono", "cargo")
        tree = ttk.Treeview(cont, columns=cols, show="headings")
        tree.pack(fill="both", expand=True, padx=10, pady=5)

        headers = ["Nombre", "Email", "Teléfono", "Cargo"]
        widths = [200, 220, 120, 150]

        for col, h, w in zip(cols, headers, widths):
            tree.heading(col, text=h)
            tree.column(col, width=w, anchor="w")

        scrollbar = ttk.Scrollbar(cont, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)

        try:
            # Departamento activo del empleado
            self.cursor.execute("""
                SELECT Departamento_idDepartamento
                FROM EmpleadoDepartamento
                WHERE UsuarioSistema_idUsuarioSistema = %s
                AND Activo = TRUE
                LIMIT 1
            """, (self.usuario.id_UsuarioSistema,))
            dep = self.cursor.fetchone()
            if not dep:
                messagebox.showinfo("Compañeros", "No estás asignado actualmente a ningún departamento.")
                return

            id_dep = dep[0]

            self.cursor.execute("""
                SELECT 
                    U.Nombre, U.Apellido, U.Email, U.Telefono, DE.TipoEmpleado
                FROM EmpleadoDepartamento ED
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN DetalleEmpleado DE ON U.id_UsuarioSistema = DE.UsuarioSistema_idUsuarioSistema
                WHERE ED.Departamento_idDepartamento = %s
                AND ED.Activo = TRUE
                AND U.id_UsuarioSistema <> %s
                ORDER BY U.Apellido, U.Nombre
            """, (id_dep, self.usuario.id_UsuarioSistema))
            companeros = self.cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Compañeros", f"Ocurrió un error al listar tus compañeros: {e}")
            return

        if not companeros:
            messagebox.showinfo("Compañeros", "No tienes compañeros actualmente en tu departamento.")
            return

        for c in companeros:
            nombre = (c[0] or "") + " " + (c[1] or "")
            email = c[2] or "(sin email)"
            telefono = c[3] or "(sin teléfono)"
            cargo = c[4] or "(sin cargo)"
            tree.insert("", "end", values=(nombre.strip(), email, telefono, cargo))

    # ---------- 6) Editar datos personales ----------

    def _emp_editar_datos(self):
        """Formulario para que el empleado edite sus datos + dirección."""
        self._clear_central()
        cont = self.panel_central

        tk.Label(cont, text="Editar datos personales", font=("Arial", 12, "bold"))\
            .grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(10, 5))

        try:
            self.cursor.execute("""
                SELECT 
                    U.Nombre, U.Apellido, U.Email, U.Telefono,
                    D.id_Direccion, D.Calle, D.Numero, D.Ciudad, D.Region, D.Pais, D.CodigoPostal
                FROM UsuarioSistema U
                LEFT JOIN Direccion D ON U.Direccion_idDireccion = D.id_Direccion
                WHERE U.id_UsuarioSistema = %s
            """, (self.usuario.id_UsuarioSistema,))
            datos = self.cursor.fetchone()
        except Exception as e:
            tk.Label(cont, text=f"Error al cargar datos: {e}", fg="red")\
                .grid(row=1, column=0, sticky="w", padx=10, pady=5)
            return

        if not datos:
            tk.Label(cont, text="No se pudo cargar la información actual.", fg="red")\
                .grid(row=1, column=0, sticky="w", padx=10, pady=5)
            return

        (nombre, apellido, email, telefono,
        id_dir, calle, numero, ciudad, region, pais, cp) = datos

        # --- campos personales ---
        labels = ["Nombre", "Apellido", "Email", "Teléfono",
                "Calle", "Número", "Ciudad", "Región", "País", "Código Postal"]
        valores = [nombre, apellido, email, telefono,
                calle or "", str(numero) if numero is not None else "",
                ciudad or "", region or "", pais or "", str(cp) if cp is not None else ""]

        entries = []
        for i, (lab, val) in enumerate(zip(labels, valores), start=1):
            tk.Label(cont, text=lab + ":").grid(row=i, column=0, sticky="e", padx=10, pady=3)
            e = tk.Entry(cont, width=35)
            e.grid(row=i, column=1, sticky="w", padx=5, pady=3)
            e.insert(0, val)
            entries.append(e)

        def guardar():
            nuevo_nombre      = entries[0].get().strip()
            nuevo_apellido    = entries[1].get().strip()
            nuevo_email       = entries[2].get().strip()
            nuevo_telefono    = entries[3].get().strip()
            nueva_calle       = entries[4].get().strip()
            nuevo_numero      = entries[5].get().strip()
            nueva_ciudad      = entries[6].get().strip()
            nueva_region      = entries[7].get().strip()
            nuevo_pais        = entries[8].get().strip()
            nuevo_codigo_post = entries[9].get().strip()

            if not nuevo_nombre or not nuevo_apellido:
                messagebox.showwarning("Validación", "Nombre y apellido no pueden estar vacíos.")
                return

            if nuevo_email and not validar_email(nuevo_email):
                messagebox.showwarning("Validación", "El email ingresado no es válido.")
                return

            if nuevo_telefono and not validar_telefono(nuevo_telefono):
                messagebox.showwarning("Validación", "El teléfono ingresado no es válido.")
                return

            # número y código postal opcionales -> int o NULL
            try:
                numero_int = int(nuevo_numero) if nuevo_numero else None
            except ValueError:
                messagebox.showwarning("Validación", "El número de dirección debe ser numérico.")
                return

            try:
                cp_int = int(nuevo_codigo_post) if nuevo_codigo_post else None
            except ValueError:
                messagebox.showwarning("Validación", "El código postal debe ser numérico.")
                return

            if not messagebox.askyesno("Confirmar", "¿Desea guardar los cambios?"):
                return

            try:
                # Actualizar usuario
                self.cursor.execute("""
                    UPDATE UsuarioSistema
                    SET Nombre=%s, Apellido=%s, Email=%s, Telefono=%s
                    WHERE id_UsuarioSistema=%s
                """, (nuevo_nombre, nuevo_apellido, nuevo_email or None,
                    nuevo_telefono or None, self.usuario.id_UsuarioSistema))

                # Actualizar dirección
                self.cursor.execute("""
                    UPDATE Direccion
                    SET Calle=%s, Numero=%s, Ciudad=%s, Region=%s, Pais=%s, CodigoPostal=%s
                    WHERE id_Direccion=%s
                """, (nueva_calle or None, numero_int, nueva_ciudad or None,
                    nueva_region or None, nuevo_pais or None, cp_int, id_dir))

                self.cnx.commit()
                messagebox.showinfo("Datos personales", "Tus datos han sido actualizados correctamente.")
            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error al guardar cambios:\n{e}")

        tk.Button(cont, text="Guardar cambios", command=guardar)\
            .grid(row=len(labels)+1, column=1, sticky="e", padx=10, pady=10)

    # ---------- 7) Cambiar usuario ----------

    def _emp_cambiar_usuario(self):
        """Permite cambiar el nombre de usuario del empleado."""
        win = tk.Toplevel(self)
        win.title("Cambiar nombre de usuario")
        win.geometry("350x150")
        win.grab_set()

        tk.Label(win, text=f"Usuario actual: {self.usuario.NombreUsuario}")\
            .pack(padx=10, pady=(10, 5))

        tk.Label(win, text="Nuevo nombre de usuario:").pack(padx=10, pady=3)
        entry_nuevo = tk.Entry(win, width=30)
        entry_nuevo.pack(padx=10, pady=3)

        def aplicar():
            nuevo_usuario = entry_nuevo.get().strip()
            if not nuevo_usuario:
                messagebox.showwarning("Validación", "El nombre de usuario no puede estar vacío.")
                return
            if len(nuevo_usuario) > 20:
                messagebox.showwarning("Validación", "El nombre de usuario no puede superar los 20 caracteres.")
                return

            # Verificar que no exista
            self.cursor.execute(
                "SELECT COUNT(*) FROM UsuarioSistema WHERE NombreUsuario = %s",
                (nuevo_usuario,)
            )
            if self.cursor.fetchone()[0] > 0:
                messagebox.showwarning("Validación", "Ese nombre de usuario ya está en uso.")
                return

            if not messagebox.askyesno("Confirmar", f"¿Cambiar tu usuario a '{nuevo_usuario}'?"):
                return

            try:
                self.cursor.execute("""
                    UPDATE UsuarioSistema
                    SET NombreUsuario = %s
                    WHERE id_UsuarioSistema = %s
                """, (nuevo_usuario, self.usuario.id_UsuarioSistema))
                self.cnx.commit()
                self.usuario.NombreUsuario = nuevo_usuario
                messagebox.showinfo("Usuario", "Nombre de usuario actualizado correctamente.")
                win.destroy()
            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error al cambiar el usuario:\n{e}")

        tk.Button(win, text="Aceptar", command=aplicar).pack(pady=10)

    # ---------- 8) Cambiar contraseña ----------

    def _emp_cambiar_contrasena(self):
        """Cambio de contraseña del empleado (con validación mínima)."""
        win = tk.Toplevel(self)
        win.title("Cambiar contraseña")
        win.geometry("350x200")
        win.grab_set()

        tk.Label(win, text="Contraseña actual:").grid(row=0, column=0, sticky="e", padx=10, pady=5)
        entry_actual = tk.Entry(win, show="*", width=25)
        entry_actual.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(win, text="Nueva contraseña:").grid(row=1, column=0, sticky="e", padx=10, pady=5)
        entry_nueva = tk.Entry(win, show="*", width=25)
        entry_nueva.grid(row=1, column=1, padx=5, pady=5)

        tk.Label(win, text="Repetir nueva contraseña:").grid(row=2, column=0, sticky="e", padx=10, pady=5)
        entry_rep = tk.Entry(win, show="*", width=25)
        entry_rep.grid(row=2, column=1, padx=5, pady=5)

        def aplicar():
            actual = entry_actual.get()
            nueva = entry_nueva.get()
            repetir = entry_rep.get()

            if not actual or not nueva or not repetir:
                messagebox.showwarning("Validación", "Complete todos los campos.")
                return

            # Verificar contraseña actual
            self.cursor.execute("""
                SELECT Contraseña FROM UsuarioSistema WHERE id_UsuarioSistema = %s
            """, (self.usuario.id_UsuarioSistema,))
            registro = self.cursor.fetchone()
            if not registro:
                messagebox.showerror("Error", "No se pudo obtener la contraseña actual.")
                return

            hash_actual = hashlib.sha256(actual.encode()).hexdigest()
            if hash_actual != registro[0]:
                messagebox.showwarning("Validación", "La contraseña actual es incorrecta.")
                return

            # Validar nueva
            if not validar_contrasena(nueva):
                messagebox.showwarning(
                    "Validación",
                    "La contraseña debe tener mínimo 8 caracteres, una mayúscula, un número y un símbolo."
                )
                return

            if nueva != repetir:
                messagebox.showwarning("Validación", "Las contraseñas nuevas no coinciden.")
                return

            nuevo_hash = hashlib.sha256(nueva.encode()).hexdigest()

            try:
                self.cursor.execute("""
                    UPDATE UsuarioSistema
                    SET Contraseña = %s
                    WHERE id_UsuarioSistema = %s
                """, (nuevo_hash, self.usuario.id_UsuarioSistema))
                self.cnx.commit()
                messagebox.showinfo("Contraseña", "Contraseña actualizada correctamente.")
                win.destroy()
            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error al cambiar la contraseña:\n{e}")

        tk.Button(win, text="Aceptar", command=aplicar)\
            .grid(row=3, column=1, sticky="e", padx=10, pady=10)





    # =========================================================
    # PANEL USUARIOS (ADMIN)
    # =========================================================
    def _show_usuarios(self):
        """Construye el subpanel de gestión de usuarios."""
        self._clear_central()

        cont = tk.Frame(self.panel_central)
        cont.pack(fill="both", expand=True, padx=10, pady=10, anchor="nw")

        tk.Label(cont, text="Gestión de usuarios", font=("Arial", 13, "bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 5)
        )
        tk.Label(cont, text="Seleccione una acción:", font=("Arial", 10)).grid(
            row=1, column=0, sticky="w", pady=(0, 10)
        )

        tk.Button(
            cont, text="Crear usuario", width=25, command=self._abrir_dialogo_crear_usuario
        ).grid(row=2, column=0, sticky="w", pady=2)

        tk.Button(
            cont, text="Editar usuario", width=25, command=self._abrir_dialogo_editar_usuario
        ).grid(row=3, column=0, sticky="w", pady=2)

        tk.Button(
            cont, text="Eliminar usuario", width=25, command=self._abrir_dialogo_eliminar_usuario
        ).grid(row=4, column=0, sticky="w", pady=2)

        tk.Button(
            cont, text="Listar usuarios", width=25, command=self._listar_usuarios_gui
        ).grid(row=5, column=0, sticky="w", pady=2)

        tk.Button(
            cont, text="Buscar usuario específico", width=25, command=self._buscar_usuario_gui
        ).grid(row=6, column=0, sticky="w", pady=2)

    # ---------- Crear usuario ----------
    def _abrir_dialogo_crear_usuario(self):
        win = tk.Toplevel(self)
        win.title("Crear usuario")
        win.grab_set()

        # ====== SECCIONES ======
        frm_personal = tk.LabelFrame(win, text="Datos personales")
        frm_personal.grid(row=0, column=0, padx=10, pady=5, sticky="nsew", columnspan=2)

        frm_dir = tk.LabelFrame(win, text="Dirección")
        frm_dir.grid(row=1, column=0, padx=10, pady=5, sticky="nsew", columnspan=2)

        frm_emp = tk.LabelFrame(win, text="Detalle empleado")
        frm_emp.grid(row=2, column=0, padx=10, pady=5, sticky="nsew", columnspan=2)

        # ====== DATOS PERSONALES ======
        labels_p = ["Nombre", "Apellido", "RUT", "Email", "Teléfono", "Nombre de usuario", "Contraseña"]
        entries_p = {}

        for i, txt in enumerate(labels_p):
            tk.Label(frm_personal, text=txt).grid(row=i, column=0, sticky="e", padx=5, pady=3)
            ent = tk.Entry(frm_personal, width=30, show="*" if txt == "Contraseña" else None)
            ent.grid(row=i, column=1, padx=5, pady=3)
            entries_p[txt] = ent

        tk.Label(frm_personal, text="Rol (admin/empleado):").grid(row=len(labels_p), column=0, sticky="e", padx=5, pady=3)
        entry_rol = tk.Entry(frm_personal, width=15)
        entry_rol.grid(row=len(labels_p), column=1, sticky="w", padx=5, pady=3)

        # ====== DIRECCIÓN ======
        labels_d = ["Calle", "Número", "Ciudad", "Región", "País", "Código Postal"]
        entries_d = {}
        for i, txt in enumerate(labels_d):
            tk.Label(frm_dir, text=txt).grid(row=i, column=0, sticky="e", padx=5, pady=3)
            ent = tk.Entry(frm_dir, width=30)
            ent.grid(row=i, column=1, padx=5, pady=3)
            entries_d[txt] = ent

        # ====== DETALLE EMPLEADO ======
        tk.Label(frm_emp, text="Salario (CLP):").grid(row=0, column=0, sticky="e", padx=5, pady=3)
        entry_salario = tk.Entry(frm_emp, width=20)
        entry_salario.grid(row=0, column=1, padx=5, pady=3)

        tk.Label(frm_emp, text="Tipo de empleado:").grid(row=1, column=0, sticky="e", padx=5, pady=3)
        entry_tipo_emp = tk.Entry(frm_emp, width=20)
        entry_tipo_emp.grid(row=1, column=1, padx=5, pady=3)

        tk.Label(frm_emp, text="Fecha contrato:").grid(
            row=2, column=0, sticky="e", padx=5, pady=3
        )
        entry_fecha = DateEntry(
            frm_emp,
            width=18,
            date_pattern="yyyy-mm-dd"  # ⬅️ para que salga YYYY-MM-DD
        )
        entry_fecha.grid(row=2, column=1, padx=5, pady=3)


        def guardar():
            # ----- personales -----
            nombre = entries_p["Nombre"].get().strip()
            apellido = entries_p["Apellido"].get().strip()
            rut = entries_p["RUT"].get().strip()
            email = entries_p["Email"].get().strip()
            telefono = entries_p["Teléfono"].get().strip()
            nombre_usuario = entries_p["Nombre de usuario"].get().strip()
            contrasena = entries_p["Contraseña"].get().strip()
            tipo_rol = entry_rol.get().strip().lower()

            # obligatorios
            if not (nombre and apellido and rut and email and nombre_usuario and contrasena and tipo_rol):
                messagebox.showwarning("Campos incompletos", "Complete todos los campos obligatorios, incluyendo el rol.")
                return

            # validar rut
            if not validar_rut(rut):
                messagebox.showwarning("RUT inválido", "El RUT ingresado no es válido.")
                return

            # email
            if not validar_email(email):
                messagebox.showwarning("Email inválido", "El email ingresado no es válido.")
                return

            # teléfono opcional
            if telefono and not validar_telefono(telefono):
                messagebox.showwarning("Teléfono inválido", "El teléfono ingresado no es válido.")
                return

            # nombre de usuario
            if len(nombre_usuario) > 20:
                messagebox.showwarning("Usuario inválido", "El nombre de usuario no puede superar los 20 caracteres.")
                return

            try:
                self.cursor.execute(
                    "SELECT COUNT(*) FROM UsuarioSistema WHERE NombreUsuario = %s",
                    (nombre_usuario,),
                )
                if self.cursor.fetchone()[0] > 0:
                    messagebox.showwarning("Usuario existente", "El nombre de usuario ya existe. Elija otro.")
                    return
            except Exception as e:
                messagebox.showerror("Error", f"Error comprobando nombre de usuario:\n{e}")
                return

            # contraseña
            if not validar_contrasena(contrasena):
                messagebox.showwarning(
                    "Contraseña inválida",
                    "La contraseña debe tener mínimo 8 caracteres, una mayúscula, un número y un símbolo.",
                )
                return
            contrasena_hash = hashlib.sha256(contrasena.encode()).hexdigest()

            # rol solo admin/empleado
            if tipo_rol not in ("admin", "empleado"):
                messagebox.showwarning("Rol inválido", "El rol debe ser 'admin' o 'empleado'.")
                return

            # ----- dirección -----
            calle = entries_d["Calle"].get().strip()
            numero = entries_d["Número"].get().strip()
            ciudad = entries_d["Ciudad"].get().strip()
            region = entries_d["Región"].get().strip()
            pais = entries_d["País"].get().strip()
            codigo_postal = entries_d["Código Postal"].get().strip()

            if not (calle and numero and ciudad and region and pais and codigo_postal):
                messagebox.showwarning("Dirección incompleta", "Complete todos los campos de dirección.")
                return

            try:
                numero_int = int(numero)
            except ValueError:
                messagebox.showwarning("Número inválido", "El número de calle debe ser un entero.")
                return

            try:
                cp_int = int(codigo_postal)
                if not (100000 <= cp_int <= 9999999):
                    messagebox.showwarning("Código postal inválido", "Código postal fuera de rango válido.")
                    return
            except ValueError:
                messagebox.showwarning("Código postal inválido", "El código postal debe ser numérico.")
                return

            # ----- detalle empleado -----
            salario = entry_salario.get().strip()
            if not salario:
                messagebox.showwarning("Salario inválido", "El salario no puede estar vacío.")
                return
            try:
                salario_int = int(salario)
                if salario_int <= 0:
                    messagebox.showwarning("Salario inválido", "El salario debe ser mayor que 0.")
                    return
                if salario_int > 10_000_000:
                    messagebox.showwarning("Salario inválido", "El salario no puede superar los 10 millones de CLP.")
                    return
            except ValueError:
                messagebox.showwarning(
                    "Salario inválido",
                    "El salario debe ser un número entero (sin puntos ni símbolos).",
                )
                return

            tipo_empleado = entry_tipo_emp.get().strip() or "No especificado"

            try:
                # DateEntry devuelve datetime.date
                fecha_obj = entry_fecha.get_date()
                fecha_contrato = fecha_obj.strftime("%Y-%m-%d")
            except Exception as e:
                messagebox.showwarning(
                    "Fecha inválida",
                    f"No se pudo obtener la fecha de contrato:\n{e}",
                )
                return

            # ----- INSERTS EN CADENA COMO EN Admin.crear_usuario -----
            try:
                # Direccion
                self.cursor.execute(
                    """
                    INSERT INTO Direccion (Calle, Numero, Ciudad, Region, Pais, CodigoPostal)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (calle, numero_int, ciudad, region, pais, cp_int),
                )
                id_direccion = self.cursor.lastrowid

                # Rol (siempre nuevo aunque exista otro igual)
                self.cursor.execute(
                    "INSERT INTO Rol (tipo_Rol) VALUES (%s)",
                    (tipo_rol,),
                )
                id_rol = self.cursor.lastrowid

                # UsuarioSistema
                self.cursor.execute(
                    """
                    INSERT INTO UsuarioSistema
                        (Nombre, Apellido, RUT, Email, Telefono,
                         NombreUsuario, Contraseña, Direccion_idDireccion, Rol_idRol)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        nombre,
                        apellido,
                        rut,
                        email,
                        telefono or None,
                        nombre_usuario,
                        contrasena_hash,
                        id_direccion,
                        id_rol,
                    ),
                )
                id_usuario = self.cursor.lastrowid

                # DetalleEmpleado
                self.cursor.execute(
                    """
                    INSERT INTO DetalleEmpleado
                        (FechaContrato, Salario, TipoEmpleado, UsuarioSistema_idUsuarioSistema)
                    VALUES (%s, %s, %s, %s)
                    """,
                    (fecha_contrato, salario_int, tipo_empleado, id_usuario),
                )

                self.cnx.commit()
                messagebox.showinfo("Éxito", f"Usuario creado correctamente (ID {id_usuario}).")
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error al crear el usuario:\n{e}")

        tk.Button(win, text="Guardar", command=guardar).grid(
            row=3, column=1, pady=10, sticky="e"
        )

    # ---------- Editar usuario (COMPLETO, como Admin.editar_usuario) ----------
    def _abrir_dialogo_editar_usuario(self):
        win = tk.Toplevel(self)
        win.title("Editar usuario")
        win.grab_set()

        tk.Label(win, text="ID de usuario a editar:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_id = tk.Entry(win, width=10)
        entry_id.grid(row=0, column=1, padx=5, pady=3)

        # ====== SECCIONES ======
        frm_personal = tk.LabelFrame(win, text="Datos personales")
        frm_personal.grid(row=1, column=0, columnspan=3, padx=10, pady=5, sticky="nsew")

        frm_dir = tk.LabelFrame(win, text="Dirección")
        frm_dir.grid(row=2, column=0, columnspan=3, padx=10, pady=5, sticky="nsew")

        frm_emp = tk.LabelFrame(win, text="Detalle empleado")
        frm_emp.grid(row=3, column=0, columnspan=3, padx=10, pady=5, sticky="nsew")

        # ====== DATOS PERSONALES ======
        labels_p = ["Nombre", "Apellido", "RUT", "Email", "Teléfono", "Nombre de usuario"]
        entries_p = {}
        for i, txt in enumerate(labels_p):
            tk.Label(frm_personal, text=txt).grid(row=i, column=0, sticky="e", padx=5, pady=3)
            ent = tk.Entry(frm_personal, width=30)
            ent.grid(row=i, column=1, padx=5, pady=3)
            entries_p[txt] = ent

        # Contraseña (opcional, nueva)
        tk.Label(frm_personal, text="Nueva contraseña:").grid(
            row=len(labels_p), column=0, sticky="e", padx=5, pady=3
        )
        entry_pass = tk.Entry(frm_personal, width=30, show="*")
        entry_pass.grid(row=len(labels_p), column=1, padx=5, pady=3)

        tk.Label(frm_personal, text="Confirmar contraseña:").grid(
            row=len(labels_p) + 1, column=0, sticky="e", padx=5, pady=3
        )
        entry_pass_conf = tk.Entry(frm_personal, width=30, show="*")
        entry_pass_conf.grid(row=len(labels_p) + 1, column=1, padx=5, pady=3)

        # Rol
        tk.Label(frm_personal, text="Rol (admin/empleado):").grid(
            row=len(labels_p) + 2, column=0, sticky="e", padx=5, pady=3
        )
        entry_rol = tk.Entry(frm_personal, width=15)
        entry_rol.grid(row=len(labels_p) + 2, column=1, sticky="w", padx=5, pady=3)

        # ====== DIRECCIÓN ======
        labels_d = ["Calle", "Número", "Ciudad", "Región", "País", "Código Postal"]
        entries_d = {}
        for i, txt in enumerate(labels_d):
            tk.Label(frm_dir, text=txt).grid(row=i, column=0, sticky="e", padx=5, pady=3)
            ent = tk.Entry(frm_dir, width=30)
            ent.grid(row=i, column=1, padx=5, pady=3)
            entries_d[txt] = ent

        # ====== DETALLE EMPLEADO ======
        tk.Label(frm_emp, text="Salario (CLP):").grid(row=0, column=0, sticky="e", padx=5, pady=3)
        entry_salario = tk.Entry(frm_emp, width=20)
        entry_salario.grid(row=0, column=1, padx=5, pady=3)

        tk.Label(frm_emp, text="Tipo de empleado:").grid(row=1, column=0, sticky="e", padx=5, pady=3)
        entry_tipo_emp = tk.Entry(frm_emp, width=20)
        entry_tipo_emp.grid(row=1, column=1, padx=5, pady=3)

        tk.Label(frm_emp, text="Fecha contrato:").grid(
            row=2, column=0, sticky="e", padx=5, pady=3
        )
        entry_fecha = DateEntry(
            frm_emp,
            width=18,
            date_pattern="yyyy-mm-dd"
        )
        entry_fecha.grid(row=2, column=1, padx=5, pady=3)

        # para mantener IDs / contraseña antigua / id detalle
        estado = {
            "id_usuario": None,
            "id_direccion": None,
            "id_rol": None,
            "id_detalle": None,
            "old_pass_hash": None,
        }

        def cargar():
            val = entry_id.get().strip()
            if not val.isdigit():
                messagebox.showwarning("ID inválido", "Ingrese un ID numérico de usuario.")
                return
            uid = int(val)
            try:
                self.cursor.execute(
                    """
                    SELECT 
                        U.Nombre, U.Apellido, U.RUT, U.Email, U.Telefono,
                        U.NombreUsuario, U.Contraseña,
                        U.Direccion_idDireccion, U.Rol_idRol,
                        D.Calle, D.Numero, D.Ciudad, D.Region, D.Pais, D.CodigoPostal,
                        R.tipo_Rol,
                        E.id_DetalleEmpleado, E.FechaContrato, E.Salario, E.TipoEmpleado
                    FROM UsuarioSistema U
                    INNER JOIN Direccion D ON U.Direccion_idDireccion = D.id_Direccion
                    INNER JOIN Rol R ON U.Rol_idRol = R.id_Rol
                    LEFT JOIN DetalleEmpleado E ON U.id_UsuarioSistema = E.UsuarioSistema_idUsuarioSistema
                    WHERE U.id_UsuarioSistema = %s
                    """,
                    (uid,),
                )
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showinfo("Sin resultados", "No se encontró ese usuario.")
                    return

                estado["id_usuario"] = uid
                estado["id_direccion"] = row[7]
                estado["id_rol"] = row[8]
                estado["id_detalle"] = row[16]
                estado["old_pass_hash"] = row[6]

                # personales
                entries_p["Nombre"].delete(0, tk.END)
                entries_p["Nombre"].insert(0, row[0])

                entries_p["Apellido"].delete(0, tk.END)
                entries_p["Apellido"].insert(0, row[1])

                entries_p["RUT"].delete(0, tk.END)
                entries_p["RUT"].insert(0, row[2])

                entries_p["Email"].delete(0, tk.END)
                entries_p["Email"].insert(0, row[3])

                entries_p["Teléfono"].delete(0, tk.END)
                entries_p["Teléfono"].insert(0, row[4] or "")

                entries_p["Nombre de usuario"].delete(0, tk.END)
                entries_p["Nombre de usuario"].insert(0, row[5])

                entry_rol.delete(0, tk.END)
                entry_rol.insert(0, row[15])

                # dirección
                entries_d["Calle"].delete(0, tk.END)
                entries_d["Calle"].insert(0, row[9])

                entries_d["Número"].delete(0, tk.END)
                entries_d["Número"].insert(0, str(row[10]))

                entries_d["Ciudad"].delete(0, tk.END)
                entries_d["Ciudad"].insert(0, row[11])

                entries_d["Región"].delete(0, tk.END)
                entries_d["Región"].insert(0, row[12])

                entries_d["País"].delete(0, tk.END)
                entries_d["País"].insert(0, row[13])

                entries_d["Código Postal"].delete(0, tk.END)
                entries_d["Código Postal"].insert(0, str(row[14]))

                # detalle empleado (puede ser None)
                entry_salario.delete(0, tk.END)
                entry_tipo_emp.delete(0, tk.END)

                if row[18] is not None:
                    entry_salario.insert(0, str(row[18]))
                if row[19] is not None:
                    entry_tipo_emp.insert(0, row[19])

                # fecha contrato si existe
                if row[17]:
                    try:
                        fecha_val = row[17]
                        if isinstance(fecha_val, (datetime, date)):
                            entry_fecha.set_date(fecha_val)
                        else:
                            entry_fecha.set_date(datetime.strptime(str(fecha_val), "%Y-%m-%d"))
                    except Exception:
                        pass  # si falla, se deja la fecha por defecto

            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error al cargar:\n{e}")

        def guardar_cambios():
            if estado["id_usuario"] is None:
                messagebox.showwarning("Atención", "Primero cargue un usuario por ID.")
                return

            # ---- personales ----
            nombre = entries_p["Nombre"].get().strip()
            apellido = entries_p["Apellido"].get().strip()
            rut = entries_p["RUT"].get().strip()
            email = entries_p["Email"].get().strip()
            telefono = entries_p["Teléfono"].get().strip()
            nombre_usuario = entries_p["Nombre de usuario"].get().strip()
            tipo_rol = entry_rol.get().strip().lower()

            if not (nombre and apellido and rut and email and nombre_usuario and tipo_rol):
                messagebox.showwarning(
                    "Campos incompletos",
                    "Complete todos los campos obligatorios (personales y rol).",
                )
                return

            if not validar_rut(rut):
                messagebox.showwarning("RUT inválido", "El RUT ingresado no es válido.")
                return

            if not validar_email(email):
                messagebox.showwarning("Email inválido", "El email ingresado no es válido.")
                return

            if telefono and not validar_telefono(telefono):
                messagebox.showwarning("Teléfono inválido", "El teléfono ingresado no es válido.")
                return

            if len(nombre_usuario) > 20:
                messagebox.showwarning(
                    "Usuario inválido",
                    "El nombre de usuario no puede superar los 20 caracteres.",
                )
                return

            if tipo_rol not in ("admin", "empleado"):
                messagebox.showwarning("Rol inválido", "El rol debe ser 'admin' o 'empleado'.")
                return

            # nombre de usuario único (excluyendo este ID)
            try:
                self.cursor.execute(
                    """
                    SELECT COUNT(*) FROM UsuarioSistema
                    WHERE NombreUsuario = %s AND id_UsuarioSistema <> %s
                    """,
                    (nombre_usuario, estado["id_usuario"]),
                )
                if self.cursor.fetchone()[0] > 0:
                    messagebox.showwarning(
                        "Usuario existente",
                        "Ya existe otro usuario con ese nombre. Elija otro.",
                    )
                    return
            except Exception as e:
                messagebox.showerror("Error", f"Error validando nombre de usuario:\n{e}")
                return

            # ---- contraseña ----
            nueva_pass = entry_pass.get().strip()
            conf_pass = entry_pass_conf.get().strip()
            if nueva_pass or conf_pass:
                # quiere cambiar
                if nueva_pass != conf_pass:
                    messagebox.showwarning(
                        "Contraseña",
                        "La confirmación no coincide con la nueva contraseña.",
                    )
                    return
                if not validar_contrasena(nueva_pass):
                    messagebox.showwarning(
                        "Contraseña inválida",
                        "La contraseña debe tener mínimo 8 caracteres, una mayúscula, un número y un símbolo.",
                    )
                    return
                pass_hash = hashlib.sha256(nueva_pass.encode()).hexdigest()
            else:
                # mantiene la anterior
                pass_hash = estado["old_pass_hash"]

            # ---- dirección ----
            calle = entries_d["Calle"].get().strip()
            numero = entries_d["Número"].get().strip()
            ciudad = entries_d["Ciudad"].get().strip()
            region = entries_d["Región"].get().strip()
            pais = entries_d["País"].get().strip()
            codigo_postal = entries_d["Código Postal"].get().strip()

            if not (calle and numero and ciudad and region and pais and codigo_postal):
                messagebox.showwarning("Dirección incompleta", "Complete todos los campos de dirección.")
                return

            try:
                numero_int = int(numero)
            except ValueError:
                messagebox.showwarning("Número inválido", "El número de calle debe ser entero.")
                return

            try:
                cp_int = int(codigo_postal)
                if not (100000 <= cp_int <= 9999999):
                    messagebox.showwarning("Código postal inválido", "Fuera de rango válido.")
                    return
            except ValueError:
                messagebox.showwarning("Código postal inválido", "Debe ser numérico.")
                return

            # ---- detalle empleado (opcional, pero si pones salario se valida todo) ----
            salario_txt = entry_salario.get().strip()
            tipo_emp_txt = entry_tipo_emp.get().strip() or "No especificado"

            salario_int = None
            fecha_contrato = None

            if salario_txt:
                try:
                    salario_int = int(salario_txt)
                    if salario_int <= 0:
                        messagebox.showwarning(
                            "Salario inválido",
                            "El salario debe ser mayor que 0.",
                        )
                        return
                    if salario_int > 10_000_000:
                        messagebox.showwarning(
                            "Salario inválido",
                            "El salario no puede superar los 10 millones de CLP.",
                        )
                        return
                except ValueError:
                    messagebox.showwarning(
                        "Salario inválido",
                        "El salario debe ser un entero (sin puntos ni símbolos).",
                    )
                    return

                try:
                    fecha_obj = entry_fecha.get_date()
                    fecha_contrato = fecha_obj.strftime("%Y-%m-%d")
                except Exception as e:
                    messagebox.showwarning(
                        "Fecha inválida",
                        f"No se pudo obtener la fecha de contrato:\n{e}",
                    )
                    return

            try:
                # --- actualizar ROL (igual que Admin.editar_usuario) ---
                self.cursor.execute(
                    "UPDATE Rol SET tipo_Rol = %s WHERE id_Rol = %s",
                    (tipo_rol, estado["id_rol"]),
                )

                # --- actualizar DIRECCIÓN ---
                self.cursor.execute(
                    """
                    UPDATE Direccion
                    SET Calle=%s, Numero=%s, Ciudad=%s, Region=%s, Pais=%s, CodigoPostal=%s
                    WHERE id_Direccion=%s
                    """,
                    (calle, numero_int, ciudad, region, pais, cp_int, estado["id_direccion"]),
                )

                # --- actualizar USUARIO ---
                self.cursor.execute(
                    """
                    UPDATE UsuarioSistema
                    SET Nombre=%s, Apellido=%s, RUT=%s, Email=%s,
                        Telefono=%s, NombreUsuario=%s, Contraseña=%s
                    WHERE id_UsuarioSistema=%s
                    """,
                    (
                        nombre,
                        apellido,
                        rut,
                        email,
                        telefono or None,
                        nombre_usuario,
                        pass_hash,
                        estado["id_usuario"],
                    ),
                )

                # --- actualizar / insertar DETALLE EMPLEADO ---
                if salario_int is not None:
                    if estado["id_detalle"]:
                        # update
                        self.cursor.execute(
                            """
                            UPDATE DetalleEmpleado
                            SET FechaContrato=%s, Salario=%s, TipoEmpleado=%s
                            WHERE id_DetalleEmpleado=%s
                            """,
                            (
                                fecha_contrato,
                                salario_int,
                                tipo_emp_txt,
                                estado["id_detalle"],
                            ),
                        )
                    else:
                        # insert
                        self.cursor.execute(
                            """
                            INSERT INTO DetalleEmpleado
                                (FechaContrato, Salario, TipoEmpleado, UsuarioSistema_idUsuarioSistema)
                            VALUES (%s, %s, %s, %s)
                            """,
                            (
                                fecha_contrato,
                                salario_int,
                                tipo_emp_txt,
                                estado["id_usuario"],
                            ),
                        )

                self.cnx.commit()
                messagebox.showinfo("Éxito", "Usuario actualizado correctamente.")
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error al guardar:\n{e}")

        tk.Button(win, text="Cargar", command=cargar).grid(row=0, column=2, padx=5, pady=3)
        tk.Button(win, text="Guardar cambios", command=guardar_cambios).grid(
            row=4, column=2, pady=10, sticky="e"
        )

    # ---------- Eliminar usuario ----------
    def _abrir_dialogo_eliminar_usuario(self):
        win = tk.Toplevel(self)
        win.title("Eliminar usuario")
        win.grab_set()

        tk.Label(win, text="Nombre de usuario a eliminar:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_user = tk.Entry(win, width=25)
        entry_user.grid(row=0, column=1, padx=5, pady=3)

        def eliminar():
            nombre_usuario = entry_user.get().strip()
            if not nombre_usuario:
                messagebox.showwarning("Atención", "Ingrese un nombre de usuario.")
                return

            try:
                # traemos datos completos para poder aplicar reglas tipo Admin.eliminar_usuario
                self.cursor.execute(
                    """
                    SELECT U.id_UsuarioSistema, U.Nombre, U.Apellido,
                           U.Direccion_idDireccion, U.Rol_idRol, R.tipo_Rol, U.NombreUsuario
                    FROM UsuarioSistema U
                    JOIN Rol R ON U.Rol_idRol = R.id_Rol
                    WHERE U.NombreUsuario = %s
                    """,
                    (nombre_usuario,),
                )
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showinfo("Sin resultados", "No se encontró ese usuario.")
                    return

                uid, nombre, apellido, id_dir, id_rol, tipo_rol, nom_user = row

                # no permitir auto-eliminar al usuario autenticado
                if hasattr(self.usuario, "id_UsuarioSistema") and uid == self.usuario.id_UsuarioSistema:
                    messagebox.showerror(
                        "Acción no permitida",
                        "No puedes eliminar el usuario con el que estás autenticado.",
                    )
                    return

                # si es admin, no permitir borrar al último admin
                if tipo_rol.lower() == "admin":
                    self.cursor.execute(
                        """
                        SELECT COUNT(*) FROM UsuarioSistema U
                        JOIN Rol R ON U.Rol_idRol = R.id_Rol
                        WHERE R.tipo_Rol = 'admin'
                        """
                    )
                    cantidad_admins = self.cursor.fetchone()[0]
                    if cantidad_admins <= 1:
                        messagebox.showerror(
                            "Acción no permitida",
                            "No se puede eliminar al último usuario con rol 'admin'.",
                        )
                        return

                resp = messagebox.askyesno(
                    "Confirmar eliminación",
                    f"¿Eliminar al usuario {nombre} {apellido} "
                    f"('{nom_user}')?\nEsta acción no se puede deshacer.",
                )
                if not resp:
                    return

                # eliminar detalle empleado primero
                self.cursor.execute(
                    "DELETE FROM DetalleEmpleado WHERE UsuarioSistema_idUsuarioSistema = %s",
                    (uid,),
                )
                # eliminar usuario
                self.cursor.execute(
                    "DELETE FROM UsuarioSistema WHERE id_UsuarioSistema=%s", (uid,)
                )

                # limpiar direccion si queda huérfana
                if id_dir is not None:
                    self.cursor.execute(
                        "SELECT COUNT(*) FROM UsuarioSistema WHERE Direccion_idDireccion = %s",
                        (id_dir,),
                    )
                    if self.cursor.fetchone()[0] == 0:
                        self.cursor.execute(
                            "DELETE FROM Direccion WHERE id_Direccion = %s",
                            (id_dir,),
                        )

                # limpiar rol si queda huérfano
                if id_rol is not None:
                    self.cursor.execute(
                        "SELECT COUNT(*) FROM UsuarioSistema WHERE Rol_idRol = %s",
                        (id_rol,),
                    )
                    if self.cursor.fetchone()[0] == 0:
                        self.cursor.execute(
                            "DELETE FROM Rol WHERE id_Rol = %s",
                            (id_rol,),
                        )

                self.cnx.commit()
                messagebox.showinfo("Éxito", "Usuario eliminado correctamente.")
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error al eliminar:\n{e}")

        tk.Button(win, text="Eliminar", command=eliminar).grid(
            row=1, column=1, pady=10, sticky="e"
        )

    # ---------- Listar usuarios ----------
    def _listar_usuarios_gui(self):
        try:
            self.cursor.execute(
                """
                SELECT U.id_UsuarioSistema, U.Nombre, U.Apellido, U.RUT, U.Email, U.Telefono,
                       U.NombreUsuario, R.tipo_Rol,
                       D.Calle, D.Numero, D.Ciudad, D.Region, D.Pais, D.CodigoPostal,
                       E.FechaContrato, E.Salario, E.TipoEmpleado
                FROM UsuarioSistema U
                INNER JOIN Rol R ON U.Rol_idRol = R.id_Rol
                INNER JOIN Direccion D ON U.Direccion_idDireccion = D.id_Direccion
                LEFT JOIN DetalleEmpleado E ON U.id_UsuarioSistema = E.UsuarioSistema_idUsuarioSistema
                ORDER BY U.id_UsuarioSistema
                """
            )
            usuarios = self.cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron obtener los usuarios:\n{e}")
            return

        if not usuarios:
            messagebox.showinfo("Usuarios", "No hay usuarios registrados.")
            return

        win = tk.Toplevel(self)
        win.title("Lista de usuarios")

        txt = tk.Text(win, width=120, height=25)
        txt.pack(padx=5, pady=5)

        txt.insert(
            "end",
            "LISTADO COMPLETO DE USUARIOS\n"
            + "=" * 80
            + "\n\n"
        )

        for u in usuarios:
            bloque = (
                f"ID: {u[0]} - {u[1]} {u[2]} (Usuario: {u[6]}, Rol: {u[7]})\n"
                f"RUT: {u[3]}, Email: {u[4]}, Teléfono: {u[5] or '-'}\n"
                f"Dirección: {u[8]} {u[9]}, {u[10]}, {u[11]}, {u[12]}, CP: {u[13]}\n"
                f"Detalle Empleado → Fecha Contrato: {u[14]}, Salario: {u[15]}, Tipo: {u[16]}\n"
                + "-" * 80
                + "\n"
            )
            txt.insert("end", bloque)

        txt.config(state="disabled")

    # ---------- Buscar usuario específico (detalles completos) ----------
    def _buscar_usuario_gui(self):
        win = tk.Toplevel(self)
        win.title("Buscar usuario")
        win.grab_set()

        tk.Label(win, text="Buscar por nombre de usuario:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_user = tk.Entry(win, width=25)
        entry_user.grid(row=0, column=1, padx=5, pady=3)

        result = tk.Text(win, width=100, height=12, state="disabled")
        result.grid(row=1, column=0, columnspan=3, padx=5, pady=5)

        def buscar():
            nombre_usuario = entry_user.get().strip()
            if not nombre_usuario:
                messagebox.showwarning("Atención", "Ingrese un nombre de usuario.")
                return
            try:
                self.cursor.execute(
                    """
                    SELECT U.id_UsuarioSistema, U.Nombre, U.Apellido, U.RUT, U.Email, U.Telefono,
                           U.NombreUsuario, R.tipo_Rol,
                           D.Calle, D.Numero, D.Ciudad, D.Region, D.Pais, D.CodigoPostal,
                           E.FechaContrato, E.Salario, E.TipoEmpleado
                    FROM UsuarioSistema U
                    INNER JOIN Rol R ON U.Rol_idRol = R.id_Rol
                    INNER JOIN Direccion D ON U.Direccion_idDireccion = D.id_Direccion
                    LEFT JOIN DetalleEmpleado E ON U.id_UsuarioSistema = E.UsuarioSistema_idUsuarioSistema
                    WHERE U.NombreUsuario = %s
                    """,
                    (nombre_usuario,),
                )
                row = self.cursor.fetchone()

                result.config(state="normal")
                result.delete("1.0", "end")

                if not row:
                    result.insert("end", "No se encontró ese usuario.\n")
                else:
                    bloque = (
                        f"ID: {row[0]} - {row[1]} {row[2]} (Usuario: {row[6]}, Rol: {row[7]})\n"
                        f"RUT: {row[3]}, Email: {row[4]}, Teléfono: {row[5] or '-'}\n"
                        f"Dirección: {row[8]} {row[9]}, {row[10]}, {row[11]}, {row[12]}, CP: {row[13]}\n"
                        f"Detalle Empleado → Fecha Contrato: {row[14]}, Salario: {row[15]}, Tipo: {row[16]}\n"
                    )
                    result.insert("end", bloque)

                result.config(state="disabled")

            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error al buscar:\n{e}")

        tk.Button(win, text="Buscar", command=buscar).grid(
            row=0, column=2, padx=5, pady=3
        )

    # =========================================================
    # PANEL PROYECTOS (ADMIN)
    # =========================================================
    def _show_proyectos(self):
        """Construye el subpanel de gestión de proyectos."""
        self._clear_central()

        cont = tk.Frame(self.panel_central)
        cont.pack(fill="both", expand=True, padx=10, pady=10, anchor="nw")

        tk.Label(cont, text="Gestión de proyectos", font=("Arial", 13, "bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 5)
        )
        tk.Label(cont, text="Seleccione una acción:", font=("Arial", 10)).grid(
            row=1, column=0, sticky="w", pady=(0, 10)
        )

        # Crear
        tk.Button(
            cont,
            text="Crear proyecto",
            width=25,
            command=self._abrir_dialogo_crear_proyecto
        ).grid(row=2, column=0, sticky="w", pady=2)

        # Editar
        tk.Button(
            cont,
            text="Editar proyecto",
            width=25,
            command=self._abrir_dialogo_editar_proyecto
        ).grid(row=3, column=0, sticky="w", pady=2)

        # Eliminar
        tk.Button(
            cont,
            text="Eliminar proyecto",
            width=25,
            command=self._abrir_dialogo_eliminar_proyecto
        ).grid(row=4, column=0, sticky="w", pady=2)

        # Listar
        tk.Button(
            cont,
            text="Listar proyectos",
            width=25,
            command=self._listar_proyectos_gui
        ).grid(row=5, column=0, sticky="w", pady=2)

        # Buscar
        tk.Button(
            cont,
            text="Buscar proyecto específico",
            width=25,
            command=self._buscar_proyecto_gui
        ).grid(row=6, column=0, sticky="w", pady=2)

    # ---------- Crear PROYECTOS ----------

    def _abrir_dialogo_crear_proyecto(self):
        win = tk.Toplevel(self)
        win.title("Crear proyecto")
        win.grab_set()

        # ---- Nombre ----
        tk.Label(win, text="Nombre del proyecto:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_nombre = tk.Entry(win, width=25)
        entry_nombre.grid(row=0, column=1, padx=5, pady=3)

        # ---- Fecha inicio (opcional) ----
        tk.Label(win, text="Fecha inicio (YYYY-MM-DD, opcional):").grid(
            row=1, column=0, sticky="e", padx=5, pady=3
        )
        entry_fecha = DateEntry(
            win,
            width=18,
            date_pattern="yyyy-mm-dd"   # para que sea 2025-03-21
        )
        entry_fecha.grid(row=1, column=1, padx=5, pady=3, sticky="w")

        # ---- Descripción (opcional) ----
        tk.Label(win, text="Descripción (opcional):").grid(
            row=2, column=0, sticky="ne", padx=5, pady=3
        )
        txt_desc = tk.Text(win, width=40, height=5)
        txt_desc.grid(row=2, column=1, padx=5, pady=3)

        def guardar():
            nombre = entry_nombre.get().strip()
            fecha_str = entry_fecha.get().strip()  # texto del DateEntry
            descripcion = txt_desc.get("1.0", "end").strip()

            # Validar nombre
            if not nombre:
                messagebox.showwarning(
                    "Campo obligatorio",
                    "El nombre del proyecto no puede estar vacío."
                )
                return

            if len(nombre) > 20:
                messagebox.showwarning(
                    "Nombre demasiado largo",
                    "El nombre del proyecto no puede superar los 20 caracteres."
                )
                return

             # Fecha opcional
            fecha = None
            if fecha_str:
                try:
                    fecha_obj = entry_fecha.get_date()   # datetime.date
                    fecha = fecha_obj.strftime("%Y-%m-%d")
                except Exception:
                    messagebox.showwarning(
                        "Fecha inválida",
                        "Seleccione una fecha válida o deje el campo vacío."
                    )
                    return

            # Validar longitud descripción (opcional)
            if descripcion and len(descripcion) > 300:
                messagebox.showwarning(
                    "Descripción demasiado larga",
                    "La descripción no puede superar los 300 caracteres."
                )
                return

            # Comprobar nombre duplicado
            try:
                self.cursor.execute(
                    "SELECT COUNT(*) FROM Proyecto WHERE LOWER(NombreProyecto) = LOWER(%s)",
                    (nombre,)
                )
                if self.cursor.fetchone()[0] > 0:
                    messagebox.showwarning(
                        "Proyecto existente",
                        "Ya existe un proyecto con ese nombre. Elija otro."
                    )
                    return
            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Ocurrió un error comprobando el nombre del proyecto:\n{e}"
                )
                return

            # Insertar
            try:
                self.cursor.execute(
                    """
                    INSERT INTO Proyecto (NombreProyecto, DescripcionProyecto, FechaInicioProyecto)
                    VALUES (%s, %s, %s)
                    """,
                    (nombre, descripcion or None, fecha)
                )
                self.cnx.commit()
                messagebox.showinfo(
                    "Éxito",
                    "Proyecto creado correctamente."
                )
                win.destroy()
            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror(
                    "Error",
                    f"No se pudo crear el proyecto:\n{e}"
                )

        tk.Button(win, text="Crear proyecto", command=guardar).grid(
            row=3, column=1, pady=10, sticky="e"
        )

    def _abrir_dialogo_editar_proyecto(self):
        win = tk.Toplevel(self)
        win.title("Editar proyecto")
        win.grab_set()

        # ---- ID a cargar ----
        tk.Label(win, text="ID de proyecto a editar:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_id = tk.Entry(win, width=10)
        entry_id.grid(row=0, column=1, padx=5, pady=3)

        btn_cargar = tk.Button(win, text="Cargar")
        btn_cargar.grid(row=0, column=2, padx=5, pady=3)

        # ---- Formulario ----
        frm = tk.LabelFrame(win, text="Datos del proyecto")
        frm.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

        # Nombre
        tk.Label(frm, text="Nombre del proyecto:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_nombre = tk.Entry(frm, width=25)
        entry_nombre.grid(row=0, column=1, padx=5, pady=3, sticky="w")

        # Fecha inicio
        tk.Label(frm, text="Fecha inicio (YYYY-MM-DD, opcional):").grid(
            row=1, column=0, sticky="e", padx=5, pady=3
        )
        entry_fecha = DateEntry(
            frm,
            width=18,
            date_pattern="yyyy-mm-dd"
        )
        entry_fecha.grid(row=1, column=1, padx=5, pady=3, sticky="w")

        # Descripción
        tk.Label(frm, text="Descripción (opcional):").grid(
            row=2, column=0, sticky="ne", padx=5, pady=3
        )
        txt_desc = tk.Text(frm, width=40, height=5)
        txt_desc.grid(row=2, column=1, padx=5, pady=3)

        # Estado interno
        estado = {"id_proyecto": None}

        def cargar():
            val = entry_id.get().strip()
            if not val.isdigit():
                messagebox.showwarning(
                    "ID inválido",
                    "Ingrese un ID numérico de proyecto."
                )
                return

            pro_id = int(val)
            try:
                self.cursor.execute(
                    """
                    SELECT id_Proyecto, NombreProyecto, FechaInicioProyecto, DescripcionProyecto
                    FROM Proyecto
                    WHERE id_Proyecto = %s
                    """,
                    (pro_id,)
                )
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showinfo(
                        "Sin resultados",
                        "No se encontró ese proyecto."
                    )
                    return

                estado["id_proyecto"] = row[0]

                # Rellenar campos
                entry_nombre.delete(0, tk.END)
                entry_nombre.insert(0, row[1] or "")

                entry_fecha.delete(0, tk.END)
                # row[2] es DATETIME → lo convertimos a 'YYYY-MM-DD' si existe
                if row[2]:
                    try:
                        fecha_val = row[2]  # puede ser datetime/date o string
                        if isinstance(fecha_val, (datetime, date)):
                            entry_fecha.set_date(fecha_val)
                        else:
                            entry_fecha.set_date(datetime.strptime(str(fecha_val)[:10], "%Y-%m-%d"))
                    except Exception:
                        # si algo raro pasa, dejamos la fecha por defecto
                        pass
                else:
                    # si no tiene fecha en BD, dejamos el campo vacío
                    entry_fecha.delete(0, "end")

                txt_desc.delete("1.0", tk.END)
                if row[3]:
                    txt_desc.insert("1.0", row[3])

            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Ocurrió un error al cargar el proyecto:\n{e}"
                )

        def guardar_cambios():
            if estado["id_proyecto"] is None:
                messagebox.showwarning(
                    "Atención",
                    "Primero cargue un proyecto por ID."
                )
                return

            nombre = entry_nombre.get().strip()
            fecha_str = entry_fecha.get().strip()
            descripcion = txt_desc.get("1.0", "end").strip()

            # Validaciones
            if not nombre:
                messagebox.showwarning(
                    "Campo obligatorio",
                    "El nombre del proyecto no puede estar vacío."
                )
                return

            if len(nombre) > 20:
                messagebox.showwarning(
                    "Nombre demasiado largo",
                    "El nombre no puede superar los 20 caracteres."
                )
                return

            if descripcion and len(descripcion) > 300:
                messagebox.showwarning(
                    "Descripción demasiado larga",
                    "La descripción no puede superar los 300 caracteres."
                )
                return

            # Validar fecha
            fecha = None
            if fecha_str:
                try:
                    fecha_obj = entry_fecha.get_date()
                    fecha = fecha_obj.strftime("%Y-%m-%d")
                except Exception:
                    messagebox.showwarning(
                        "Fecha inválida",
                        "La fecha debe ser válida o el campo debe quedar vacío."
                    )
                    return

            # Verificar nombre duplicado (excluyendo este ID)
            try:
                self.cursor.execute(
                    """
                    SELECT COUNT(*) FROM Proyecto
                    WHERE LOWER(NombreProyecto) = LOWER(%s)
                      AND id_Proyecto <> %s
                    """,
                    (nombre, estado["id_proyecto"])
                )
                if self.cursor.fetchone()[0] > 0:
                    messagebox.showwarning(
                        "Proyecto existente",
                        "Ya existe otro proyecto con ese nombre."
                    )
                    return
            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Error comprobando nombre del proyecto:\n{e}"
                )
                return

            # UPDATE
            try:
                self.cursor.execute(
                    """
                    UPDATE Proyecto
                    SET NombreProyecto=%s,
                        DescripcionProyecto=%s,
                        FechaInicioProyecto=%s
                    WHERE id_Proyecto=%s
                    """,
                    (nombre, descripcion or None, fecha, estado["id_proyecto"])
                )
                self.cnx.commit()
                messagebox.showinfo(
                    "Éxito",
                    "Proyecto actualizado correctamente."
                )
                win.destroy()
            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror(
                    "Error",
                    f"Ocurrió un error al guardar los cambios:\n{e}"
                )

        btn_cargar.config(command=cargar)

        tk.Button(
            win,
            text="Guardar cambios",
            command=guardar_cambios
        ).grid(row=2, column=2, padx=5, pady=10, sticky="e")

    def _abrir_dialogo_eliminar_proyecto(self):
        win = tk.Toplevel(self)
        win.title("Eliminar proyecto")
        win.grab_set()

        tk.Label(win, text="ID de proyecto a eliminar:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_id = tk.Entry(win, width=10)
        entry_id.grid(row=0, column=1, padx=5, pady=3)

        def eliminar():
            val = entry_id.get().strip()
            if not val.isdigit():
                messagebox.showwarning(
                    "ID inválido",
                    "Ingrese un ID numérico de proyecto."
                )
                return

            pro_id = int(val)

            try:
                # Traer datos básicos del proyecto
                self.cursor.execute(
                    "SELECT id_Proyecto, NombreProyecto FROM Proyecto WHERE id_Proyecto = %s",
                    (pro_id,)
                )
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showinfo(
                        "Sin resultados",
                        "No se encontró ese proyecto."
                    )
                    return

                _, nombre = row

                # Verificar si tiene empleados asignados (activos)
                self.cursor.execute(
                    """
                    SELECT COUNT(*) 
                    FROM EmpleadoProyecto 
                    WHERE Proyecto_idProyecto = %s AND Activo = TRUE
                    """,
                    (pro_id,)
                )
                count = self.cursor.fetchone()[0]

                if count > 0:
                    messagebox.showwarning(
                        "No se puede eliminar",
                        f"No se puede eliminar el proyecto.\n"
                        f"Tiene {count} empleado(s) asignado(s) actualmente."
                    )
                    return

                # Confirmación
                resp = messagebox.askyesno(
                    "Confirmar eliminación",
                    f"¿Eliminar el proyecto '{nombre}'?\n"
                    f"Esta acción no se puede deshacer."
                )
                if not resp:
                    return

                # Eliminar
                self.cursor.execute(
                    "DELETE FROM Proyecto WHERE id_Proyecto = %s",
                    (pro_id,)
                )
                self.cnx.commit()
                messagebox.showinfo(
                    "Éxito",
                    "Proyecto eliminado correctamente."
                )
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror(
                    "Error",
                    f"No se pudo eliminar el proyecto:\n{e}"
                )

        tk.Button(win, text="Eliminar", command=eliminar).grid(
            row=1, column=1, pady=10, sticky="e"
        )

    def _listar_proyectos_gui(self):
        try:
            self.cursor.execute(
                """
                SELECT id_Proyecto, NombreProyecto, FechaInicioProyecto, DescripcionProyecto
                FROM Proyecto
                ORDER BY id_Proyecto
                """
            )
            proyectos = self.cursor.fetchall()
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudieron obtener los proyectos:\n{e}"
            )
            return

        if not proyectos:
            messagebox.showinfo(
                "Proyectos",
                "No hay proyectos registrados."
            )
            return

        win = tk.Toplevel(self)
        win.title("Lista de proyectos")

        txt = tk.Text(win, width=110, height=25)
        txt.pack(padx=5, pady=5)

        txt.insert(
            "end",
            "LISTADO DE PROYECTOS\n" + "=" * 80 + "\n\n"
        )

        for p in proyectos:
            pro_id, nombre, fecha, desc = p
            fecha_str = str(fecha) if fecha else "(sin definir)"
            bloque = (
                f"ID: {pro_id} - {nombre}\n"
                f"Fecha inicio: {fecha_str}\n"
                f"Descripción: {desc or '(sin descripción)'}\n"
                + "-" * 80 + "\n"
            )
            txt.insert("end", bloque)

        txt.config(state="disabled")

    # =========================================================
    # PANEL DEPARTAMENTOS (ADMIN)
    # =========================================================
    def _show_departamentos(self):
        """Subpanel simple de menú para departamentos."""
        self._clear_central()

        cont = tk.Frame(self.panel_central)
        cont.pack(fill="both", expand=True, padx=10, pady=10, anchor="nw")

        tk.Label(
            cont,
            text="Gestión de departamentos",
            font=("Arial", 13, "bold")
        ).grid(row=0, column=0, sticky="w", pady=(0, 5))

        tk.Label(
            cont,
            text="Seleccione una acción:",
            font=("Arial", 10)
        ).grid(row=1, column=0, sticky="w", pady=(0, 10))

        tk.Button(
            cont,
            text="Crear departamento",
            width=25,
            command=self._abrir_dialogo_crear_departamento
        ).grid(row=2, column=0, sticky="w", pady=2)

        tk.Button(
            cont,
            text="Editar departamento",
            width=25,
            command=self._abrir_dialogo_editar_departamento
        ).grid(row=3, column=0, sticky="w", pady=2)

        tk.Button(
            cont,
            text="Eliminar departamento",
            width=25,
            command=self._abrir_dialogo_eliminar_departamento
        ).grid(row=4, column=0, sticky="w", pady=2)

        tk.Button(
            cont,
            text="Listar departamentos",
            width=25,
            command=self._listar_departamentos_gui
        ).grid(row=5, column=0, sticky="w", pady=2)

        tk.Button(
            cont,
            text="Buscar departamento",
            width=25,
            command=self._buscar_departamento_gui
        ).grid(row=6, column=0, sticky="w", pady=2)

    # ---------- Crear DEPARTAMENTOS ----------

    def _abrir_dialogo_crear_departamento(self):
        win = tk.Toplevel(self)
        win.title("Crear departamento")
        win.grab_set()

        frm = tk.LabelFrame(win, text="Datos del departamento")
        frm.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # Nombre
        tk.Label(frm, text="Nombre del departamento:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_nombre = tk.Entry(frm, width=30)
        entry_nombre.grid(row=0, column=1, padx=5, pady=3)

        # Tipo
        tk.Label(frm, text="Tipo de departamento:").grid(
            row=1, column=0, sticky="e", padx=5, pady=3
        )
        entry_tipo = tk.Entry(frm, width=30)
        entry_tipo.grid(row=1, column=1, padx=5, pady=3)

        # Descripción
        tk.Label(frm, text="Descripción (opcional):").grid(
            row=2, column=0, sticky="ne", padx=5, pady=3
        )
        txt_desc = tk.Text(frm, width=30, height=4)
        txt_desc.grid(row=2, column=1, padx=5, pady=3)

        # ----------- DIRECCIÓN (OBLIGATORIA) -----------
        tk.Label(frm, text="Dirección asociada:").grid(
            row=3, column=0, sticky="e", padx=5, pady=(8, 3)
        )

        # estas variables guardan el ID real y el texto mostrado
        dir_id_var = tk.IntVar(value=0)
        dir_text_var = tk.StringVar(value="(ninguna seleccionada)")

        entry_dir = tk.Entry(
            frm,
            width=40,
            textvariable=dir_text_var,
            state="readonly"
        )
        entry_dir.grid(row=3, column=1, padx=(5, 0), pady=(8, 3), sticky="w")

        btn_buscar_dir = tk.Button(
            frm,
            text="Buscar...",
            command=lambda: self._abrir_selector_direccion(dir_id_var, dir_text_var)
        )
        btn_buscar_dir.grid(row=3, column=2, padx=5, pady=(8, 3))

        def guardar_departamento():
            nombre = entry_nombre.get().strip()
            tipo = entry_tipo.get().strip()
            descripcion = txt_desc.get("1.0", "end").strip()

            # ---- Validaciones básicas ----
            if not nombre:
                messagebox.showwarning(
                    "Campo obligatorio",
                    "El nombre del departamento no puede estar vacío."
                )
                return

            if len(nombre) > 25:  # respeta VARCHAR(25)
                messagebox.showwarning(
                    "Nombre demasiado largo",
                    "El nombre no puede superar los 25 caracteres."
                )
                return

            if not tipo:
                messagebox.showwarning(
                    "Campo obligatorio",
                    "El tipo de departamento no puede estar vacío."
                )
                return

            if len(tipo) > 30:  # respeta VARCHAR(30)
                messagebox.showwarning(
                    "Tipo demasiado largo",
                    "El tipo no puede superar los 30 caracteres."
                )
                return

            if descripcion and len(descripcion) > 200:
                messagebox.showwarning(
                    "Descripción demasiado larga",
                    "La descripción no puede superar los 200 caracteres."
                )
                return

            # Dirección seleccionada mediante el buscador
            id_direccion = dir_id_var.get()
            if id_direccion == 0:
                messagebox.showwarning(
                    "Dirección requerida",
                    "Debe seleccionar una dirección con el botón 'Buscar...'."
                )
                return

            # ---- Verificar nombre duplicado ----
            try:
                self.cursor.execute(
                    """
                    SELECT COUNT(*) 
                    FROM Departamento 
                    WHERE LOWER(NombreDepartamento) = LOWER(%s)
                    """,
                    (nombre,),
                )
                if self.cursor.fetchone()[0] > 0:
                    messagebox.showwarning(
                        "Departamento existente",
                        "Ya existe un departamento con ese nombre."
                    )
                    return
            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Error comprobando nombre del departamento:\n{e}"
                )
                return

            # ---- INSERT ----
            try:
                self.cursor.execute(
                    """
                    INSERT INTO Departamento 
                        (NombreDepartamento, TipoDepartamento,
                         DescripcionDepartamento, Direccion_idDireccion)
                    VALUES (%s, %s, %s, %s)
                    """,
                    (nombre, tipo, descripcion or None, id_direccion),
                )
                self.cnx.commit()
                messagebox.showinfo(
                    "Éxito",
                    "Departamento creado correctamente."
                )
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror(
                    "Error",
                    f"Ocurrió un error al crear el departamento:\n{e}"
                )

        tk.Button(win, text="Guardar", command=guardar_departamento).grid(
            row=1, column=0, padx=10, pady=(0, 10), sticky="e"
        )

    def _abrir_selector_direccion(self, var_id, var_text):
        """
        Ventana modal para buscar y seleccionar una dirección.
        - var_id: tk.IntVar donde se guardará id_Direccion.
        - var_text: tk.StringVar con el texto legible de la dirección.
        """
        win = tk.Toplevel(self)
        win.title("Seleccionar dirección")
        win.grab_set()
        win.geometry("800x400")

        # ----- BARRA DE BÚSQUEDA -----
        frm_top = tk.Frame(win)
        frm_top.pack(fill="x", padx=10, pady=5)

        tk.Label(frm_top, text="Buscar:").pack(side="left")
        search_var = tk.StringVar()
        entry_search = tk.Entry(frm_top, textvariable=search_var, width=40)
        entry_search.pack(side="left", padx=5)

        tk.Label(frm_top, text="(ID, calle, ciudad, región, país, CP)").pack(
            side="left", padx=5
        )

        # ----- TREEVIEW -----
        frm_tree = tk.Frame(win)
        frm_tree.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("id", "calle", "num", "ciudad", "region", "pais", "cp")
        tree = ttk.Treeview(frm_tree, columns=cols, show="headings")
        tree.pack(side="left", fill="both", expand=True)

        headers = ["ID", "Calle", "Número", "Ciudad", "Región", "País", "CP"]
        for col, h in zip(cols, headers):
            tree.heading(col, text=h)
            tree.column(col, width=100, anchor="w")

        scrollbar = ttk.Scrollbar(frm_tree, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)

        # ----- CARGAR TODAS LAS DIRECCIONES -----
        try:
            self.cursor.execute(
                """
                SELECT id_Direccion, Calle, Numero, Ciudad, Region, Pais, CodigoPostal
                FROM Direccion
                ORDER BY id_Direccion
                """
            )
            todas = self.cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar las direcciones:\n{e}")
            win.destroy()
            return

        def refrescar_lista(filtro=""):
            filtro = filtro.lower()
            tree.delete(*tree.get_children())
            for row in todas:
                # row: (id, calle, numero, ciudad, region, pais, cp)
                texto_busq = " ".join(str(x) for x in row).lower()
                if filtro in texto_busq:
                    tree.insert("", "end", values=row)

        refrescar_lista()

        def on_search(*_):
            q = search_var.get().strip()
            refrescar_lista(q)

        search_var.trace_add("write", on_search)

        # ----- BOTONES INFERIORES -----
        frm_bottom = tk.Frame(win)
        frm_bottom.pack(fill="x", padx=10, pady=5)

        def seleccionar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning(
                    "Seleccione una fila",
                    "Debe seleccionar una dirección de la lista."
                )
                return
            values = tree.item(sel[0], "values")
            dir_id = int(values[0])
            texto = (
                f"{values[0]} - {values[1]} {values[2]}, "
                f"{values[3]}, {values[4]}, {values[5]} (CP {values[6]})"
            )
            var_id.set(dir_id)
            var_text.set(texto)
            win.destroy()

        tk.Button(frm_bottom, text="Seleccionar", command=seleccionar).pack(side="right")
        tk.Button(frm_bottom, text="Cancelar", command=win.destroy).pack(
            side="right", padx=5
        )

        # doble clic en la fila → seleccionar
        def on_double_click(event):
            seleccionar()

        tree.bind("<Double-1>", on_double_click)
        
    def _abrir_dialogo_editar_departamento(self):
        win = tk.Toplevel(self)
        win.title("Editar departamento")
        win.grab_set()

        # ---------- ID A EDITAR ----------
        tk.Label(win, text="ID de departamento a editar:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_id = tk.Entry(win, width=10)
        entry_id.grid(row=0, column=1, padx=5, pady=3)

        btn_cargar = tk.Button(win, text="Cargar")
        btn_cargar.grid(row=0, column=2, padx=5, pady=3)

        # ---------- FORMULARIO ----------
        frm = tk.LabelFrame(win, text="Datos del departamento")
        frm.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

        # Nombre
        tk.Label(frm, text="Nombre del departamento:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_nombre = tk.Entry(frm, width=30)
        entry_nombre.grid(row=0, column=1, padx=5, pady=3)

        # Tipo
        tk.Label(frm, text="Tipo de departamento:").grid(
            row=1, column=0, sticky="e", padx=5, pady=3
        )
        entry_tipo = tk.Entry(frm, width=30)
        entry_tipo.grid(row=1, column=1, padx=5, pady=3)

        # Descripción
        tk.Label(frm, text="Descripción (opcional):").grid(
            row=2, column=0, sticky="ne", padx=5, pady=3
        )
        txt_desc = tk.Text(frm, width=30, height=4)
        txt_desc.grid(row=2, column=1, padx=5, pady=3)

        # Dirección asociada (con buscador)
        tk.Label(frm, text="Dirección asociada:").grid(
            row=3, column=0, sticky="e", padx=5, pady=(8, 3)
        )

        dir_id_var = tk.IntVar(value=0)
        dir_text_var = tk.StringVar(value="(ninguna seleccionada)")

        entry_dir = tk.Entry(
            frm,
            width=40,
            textvariable=dir_text_var,
            state="readonly"
        )
        entry_dir.grid(row=3, column=1, padx=(5, 0), pady=(8, 3), sticky="w")

        btn_buscar_dir = tk.Button(
            frm,
            text="Buscar...",
            command=lambda: self._abrir_selector_direccion(dir_id_var, dir_text_var)
        )
        btn_buscar_dir.grid(row=3, column=2, padx=5, pady=(8, 3))

        # guardamos el ID de departamento actual
        estado = {
            "id_dep": None
        }

        # ---------- FUNCIONES ----------
        def cargar():
            val = entry_id.get().strip()
            if not val.isdigit():
                messagebox.showwarning(
                    "ID inválido",
                    "Ingrese un ID numérico de departamento."
                )
                return

            dep_id = int(val)
            try:
                self.cursor.execute(
                    """
                    SELECT D.id_Departamento,
                           D.NombreDepartamento,
                           D.TipoDepartamento,
                           D.DescripcionDepartamento,
                           D.Direccion_idDireccion,
                           Dir.Calle, Dir.Numero, Dir.Ciudad, Dir.Region, Dir.Pais, Dir.CodigoPostal
                    FROM Departamento D
                    JOIN Direccion Dir
                        ON D.Direccion_idDireccion = Dir.id_Direccion
                    WHERE D.id_Departamento = %s
                    """,
                    (dep_id,)
                )
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showinfo(
                        "Sin resultados",
                        "No se encontró ese departamento."
                    )
                    return

                estado["id_dep"] = row[0]

                # Rellenar campos
                entry_nombre.delete(0, tk.END)
                entry_nombre.insert(0, row[1])

                entry_tipo.delete(0, tk.END)
                entry_tipo.insert(0, row[2])

                txt_desc.delete("1.0", tk.END)
                if row[3]:
                    txt_desc.insert("1.0", row[3])

                dir_id_var.set(row[4])
                dir_text_var.set(
                    f"{row[4]} - {row[5]} {row[6]}, {row[7]}, {row[8]}, {row[9]} (CP {row[10]})"
                )

            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Ocurrió un error al cargar el departamento:\n{e}"
                )

        def guardar_cambios():
            if estado["id_dep"] is None:
                messagebox.showwarning(
                    "Atención",
                    "Primero cargue un departamento por ID."
                )
                return

            nombre = entry_nombre.get().strip()
            tipo = entry_tipo.get().strip()
            descripcion = txt_desc.get("1.0", "end").strip()
            id_direccion = dir_id_var.get()

            # Validaciones
            if not nombre:
                messagebox.showwarning(
                    "Campo obligatorio",
                    "El nombre del departamento no puede estar vacío."
                )
                return

            if len(nombre) > 25:
                messagebox.showwarning(
                    "Nombre demasiado largo",
                    "El nombre no puede superar los 25 caracteres."
                )
                return

            if not tipo:
                messagebox.showwarning(
                    "Campo obligatorio",
                    "El tipo de departamento no puede estar vacío."
                )
                return

            if len(tipo) > 30:
                messagebox.showwarning(
                    "Tipo demasiado largo",
                    "El tipo no puede superar los 30 caracteres."
                )
                return

            if descripcion and len(descripcion) > 200:
                messagebox.showwarning(
                    "Descripción demasiado larga",
                    "La descripción no puede superar los 200 caracteres."
                )
                return

            if id_direccion == 0:
                messagebox.showwarning(
                    "Dirección requerida",
                    "Debe seleccionar una dirección con el botón 'Buscar...'."
                )
                return

            # verificar nombre duplicado (excluyendo este mismo ID)
            try:
                self.cursor.execute(
                    """
                    SELECT COUNT(*)
                    FROM Departamento
                    WHERE LOWER(NombreDepartamento) = LOWER(%s)
                      AND id_Departamento <> %s
                    """,
                    (nombre, estado["id_dep"])
                )
                if self.cursor.fetchone()[0] > 0:
                    messagebox.showwarning(
                        "Departamento existente",
                        "Ya existe otro departamento con ese nombre."
                    )
                    return
            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Error comprobando nombre del departamento:\n{e}"
                )
                return

            # UPDATE
            try:
                self.cursor.execute(
                    """
                    UPDATE Departamento
                    SET NombreDepartamento=%s,
                        TipoDepartamento=%s,
                        DescripcionDepartamento=%s,
                        Direccion_idDireccion=%s
                    WHERE id_Departamento=%s
                    """,
                    (nombre, tipo, descripcion or None, id_direccion, estado["id_dep"])
                )
                self.cnx.commit()
                messagebox.showinfo(
                    "Éxito",
                    "Departamento actualizado correctamente."
                )
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror(
                    "Error",
                    f"Ocurrió un error al guardar los cambios:\n{e}"
                )

        btn_cargar.config(command=cargar)

        tk.Button(
            win,
            text="Guardar cambios",
            command=guardar_cambios
        ).grid(row=2, column=2, padx=5, pady=10, sticky="e")

    def _abrir_dialogo_eliminar_departamento(self):
        win = tk.Toplevel(self)
        win.title("Eliminar departamento")
        win.grab_set()

        tk.Label(win, text="ID de departamento a eliminar:").grid(
            row=0, column=0, sticky="e", padx=5, pady=3
        )
        entry_id = tk.Entry(win, width=10)
        entry_id.grid(row=0, column=1, padx=5, pady=3)

        def eliminar():
            val = entry_id.get().strip()
            if not val.isdigit():
                messagebox.showwarning(
                    "ID inválido",
                    "Ingrese un ID numérico de departamento."
                )
                return

            dep_id = int(val)
            try:
                # Traemos datos básicos para confirmar
                self.cursor.execute(
                    """
                    SELECT D.id_Departamento,
                           D.NombreDepartamento,
                           D.TipoDepartamento
                    FROM Departamento D
                    WHERE D.id_Departamento = %s
                    """,
                    (dep_id,)
                )
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showinfo(
                        "Sin resultados",
                        "No se encontró ese departamento."
                    )
                    return

                _, nombre, tipo = row

                resp = messagebox.askyesno(
                    "Confirmar eliminación",
                    f"¿Eliminar el departamento '{nombre}' "
                    f"({tipo})?\nEsta acción no se puede deshacer."
                )
                if not resp:
                    return

                # Intentamos eliminar
                self.cursor.execute(
                    "DELETE FROM Departamento WHERE id_Departamento = %s",
                    (dep_id,)
                )
                self.cnx.commit()
                messagebox.showinfo(
                    "Éxito",
                    "Departamento eliminado correctamente."
                )
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror(
                    "Error",
                    "No se pudo eliminar el departamento.\n"
                    "Es posible que esté siendo usado por otros registros.\n\n"
                    f"Detalle técnico:\n{e}"
                )

        tk.Button(win, text="Eliminar", command=eliminar).grid(
            row=1, column=1, pady=10, sticky="e"
        )

    def _listar_departamentos_gui(self):
        try:
            self.cursor.execute(
                """
                SELECT D.id_Departamento,
                       D.NombreDepartamento,
                       D.TipoDepartamento,
                       D.DescripcionDepartamento,
                       Dir.Calle, Dir.Numero, Dir.Ciudad,
                       Dir.Region, Dir.Pais, Dir.CodigoPostal
                FROM Departamento D
                JOIN Direccion Dir
                    ON D.Direccion_idDireccion = Dir.id_Direccion
                ORDER BY D.id_Departamento
                """
            )
            departamentos = self.cursor.fetchall()
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudieron obtener los departamentos:\n{e}"
            )
            return

        if not departamentos:
            messagebox.showinfo(
                "Departamentos",
                "No hay departamentos registrados."
            )
            return

        win = tk.Toplevel(self)
        win.title("Lista de departamentos")

        txt = tk.Text(win, width=120, height=25)
        txt.pack(padx=5, pady=5)

        txt.insert(
            "end",
            "LISTADO DE DEPARTAMENTOS\n"
            + "=" * 80
            + "\n\n"
        )

        for d in departamentos:
            (id_dep, nombre, tipo, desc,
             calle, num, ciudad, region, pais, cp) = d

            bloque = (
                f"ID: {id_dep} - {nombre} (Tipo: {tipo})\n"
                f"Descripción: {desc or '-'}\n"
                f"Dirección: {calle} {num}, {ciudad}, "
                f"{region}, {pais}, CP: {cp}\n"
                + "-" * 80
                + "\n"
            )
            txt.insert("end", bloque)

        txt.config(state="disabled")

    def _buscar_proyecto_gui(self):
        win = tk.Toplevel(self)
        win.title("Buscar proyecto")
        win.grab_set()

        # --- Modo de búsqueda ---
        tk.Label(win, text="Buscar por:").grid(
            row=0, column=0, padx=5, pady=5, sticky="e"
        )

        modo_var = tk.StringVar(value="id")

        rb_id = tk.Radiobutton(win, text="ID", variable=modo_var, value="id")
        rb_id.grid(row=0, column=1, sticky="w", padx=5, pady=5)

        rb_nombre = tk.Radiobutton(win, text="Nombre contiene", variable=modo_var, value="nombre")
        rb_nombre.grid(row=0, column=2, sticky="w", padx=5, pady=5)

        # --- Campo de valor ---
        tk.Label(win, text="Valor:").grid(
            row=1, column=0, padx=5, pady=5, sticky="e"
        )
        entry_valor = tk.Entry(win, width=30)
        entry_valor.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="w")

        # --- Área de resultados ---
        txt_resultados = tk.Text(win, width=100, height=20)
        txt_resultados.grid(row=3, column=0, columnspan=3, padx=5, pady=10)
        txt_resultados.config(state="disabled")

        def mostrar_resultados(filas):
            txt_resultados.config(state="normal")
            txt_resultados.delete("1.0", "end")

            if not filas:
                txt_resultados.insert("end", "No se encontraron proyectos con ese criterio.\n")
                txt_resultados.config(state="disabled")
                return

            txt_resultados.insert("end", "RESULTADOS DE BÚSQUEDA\n" + "=" * 80 + "\n\n")

            for p in filas:
                pro_id, nombre, fecha, desc = p
                fecha_str = str(fecha) if fecha else "(sin definir)"
                bloque = (
                    f"ID: {pro_id}\n"
                    f"Nombre: {nombre}\n"
                    f"Fecha inicio: {fecha_str}\n"
                    f"Descripción: {desc or '(sin descripción)'}\n"
                    + "-" * 80 + "\n"
                )
                txt_resultados.insert("end", bloque)

            txt_resultados.config(state="disabled")

        def buscar():
            valor = entry_valor.get().strip()
            modo = modo_var.get()

            if not valor:
                messagebox.showwarning(
                    "Campo vacío",
                    "Ingrese un valor para buscar."
                )
                return

            try:
                if modo == "id":
                    if not valor.isdigit():
                        messagebox.showwarning(
                            "ID inválido",
                            "Para buscar por ID, ingrese un número entero."
                        )
                        return

                    pro_id = int(valor)
                    self.cursor.execute(
                        """
                        SELECT id_Proyecto, NombreProyecto, FechaInicioProyecto, DescripcionProyecto
                        FROM Proyecto
                        WHERE id_Proyecto = %s
                        """,
                        (pro_id,)
                    )
                    filas = self.cursor.fetchall()

                else:  # buscar por nombre
                    # Usamos LIKE con LOWER para búsqueda flexible y limitamos resultados
                    self.cursor.execute(
                        """
                        SELECT id_Proyecto, NombreProyecto, FechaInicioProyecto, DescripcionProyecto
                        FROM Proyecto
                        WHERE LOWER(NombreProyecto) LIKE LOWER(%s)
                        ORDER BY id_Proyecto
                        LIMIT 50
                        """,
                        (f"%{valor}%",)
                    )
                    filas = self.cursor.fetchall()

                mostrar_resultados(filas)

            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Ocurrió un error al buscar proyectos:\n{e}"
                )

        btn_buscar = tk.Button(win, text="Buscar", command=buscar)
        btn_buscar.grid(row=2, column=2, padx=5, pady=5, sticky="e")

    def _buscar_departamento_gui(self):
        win = tk.Toplevel(self)
        win.title("Buscar departamento")
        win.grab_set()

        # --- Modo de búsqueda ---
        tk.Label(win, text="Buscar por:").grid(
            row=0, column=0, padx=5, pady=5, sticky="e"
        )

        modo_var = tk.StringVar(value="id")

        rb_id = tk.Radiobutton(win, text="ID", variable=modo_var, value="id")
        rb_id.grid(row=0, column=1, sticky="w", padx=5, pady=5)

        rb_nombre = tk.Radiobutton(win, text="Nombre contiene",
                                variable=modo_var, value="nombre")
        rb_nombre.grid(row=0, column=2, sticky="w", padx=5, pady=5)

        # --- Campo valor ---
        tk.Label(win, text="Valor:").grid(
            row=1, column=0, padx=5, pady=5, sticky="e"
        )
        entry_valor = tk.Entry(win, width=30)
        entry_valor.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="w")

        # --- Área resultados ---
        txt_resultados = tk.Text(win, width=110, height=20, state="disabled")
        txt_resultados.grid(row=3, column=0, columnspan=3, padx=5, pady=10)

        def mostrar_resultados(filas):
            txt_resultados.config(state="normal")
            txt_resultados.delete("1.0", "end")

            if not filas:
                txt_resultados.insert("end", "No se encontraron departamentos.\n")
                txt_resultados.config(state="disabled")
                return

            txt_resultados.insert("end",
                "RESULTADOS DE BÚSQUEDA\n" + "=" * 80 + "\n\n"
            )

            for d in filas:
                (id_dep, nombre, tipo, desc,
                calle, num, ciudad, region, pais, cp) = d

                bloque = (
                    f"ID: {id_dep} - {nombre} (Tipo: {tipo})\n"
                    f"Descripción: {desc or '-'}\n"
                    f"Dirección: {calle} {num}, {ciudad}, "
                    f"{region}, {pais}, CP: {cp}\n"
                    + "-" * 80 + "\n"
                )
                txt_resultados.insert("end", bloque)

            txt_resultados.config(state="disabled")

        def buscar():
            valor = entry_valor.get().strip()
            modo = modo_var.get()

            if not valor:
                messagebox.showwarning(
                    "Campo vacío",
                    "Ingrese un valor para buscar."
                )
                return

            try:
                if modo == "id":
                    if not valor.isdigit():
                        messagebox.showwarning(
                            "ID inválido",
                            "Para buscar por ID, ingrese un número entero."
                        )
                        return

                    dep_id = int(valor)
                    self.cursor.execute(
                        """
                        SELECT D.id_Departamento,
                            D.NombreDepartamento,
                            D.TipoDepartamento,
                            D.DescripcionDepartamento,
                            Dir.Calle, Dir.Numero, Dir.Ciudad,
                            Dir.Region, Dir.Pais, Dir.CodigoPostal
                        FROM Departamento D
                        JOIN Direccion Dir
                        ON D.Direccion_idDireccion = Dir.id_Direccion
                        WHERE D.id_Departamento = %s
                        """,
                        (dep_id,)
                    )
                    filas = self.cursor.fetchall()

                else:  # búsqueda por nombre, optimizada con LIKE + LIMIT
                    self.cursor.execute(
                        """
                        SELECT D.id_Departamento,
                            D.NombreDepartamento,
                            D.TipoDepartamento,
                            D.DescripcionDepartamento,
                            Dir.Calle, Dir.Numero, Dir.Ciudad,
                            Dir.Region, Dir.Pais, Dir.CodigoPostal
                        FROM Departamento D
                        JOIN Direccion Dir
                        ON D.Direccion_idDireccion = Dir.id_Direccion
                        WHERE LOWER(D.NombreDepartamento) LIKE LOWER(%s)
                        ORDER BY D.id_Departamento
                        LIMIT 50
                        """,
                        (f"%{valor}%",)
                    )
                    filas = self.cursor.fetchall()

                mostrar_resultados(filas)

            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Ocurrió un error al buscar departamentos:\n{e}"
                )

        tk.Button(win, text="Buscar", command=buscar).grid(
            row=2, column=2, padx=5, pady=5, sticky="e"
        )


    # =========================================================
    # PANEL ASIGNACIONES (ADMIN)
    # =========================================================
    def _show_asignaciones(self):
        self._clear_central() # mismo helper que usas en el resto

        cont = tk.Frame(self.panel_central)
        cont.pack(fill="both", expand=True, padx=10, pady=10, anchor="nw")

        lf_dep = tk.LabelFrame(cont, text="Empleado ↔ Departamento")
        lf_dep.grid(row=0, column=0, padx=10, pady=10, sticky="nw")

        tk.Button(
            lf_dep,
            text="Asignar empleado a departamento",
            width=35,
            command=self._asignar_empleado_departamento_gui
        ).grid(row=0, column=0, sticky="w", pady=2)

        tk.Button(
            lf_dep,
            text="Listar empleados por departamento",
            width=35,
            command=self._listar_empleados_departamento_gui
        ).grid(row=1, column=0, sticky="w", pady=2)

        tk.Button(
            lf_dep,
            text="Editar asignación empleado-departamento",
            width=35,
            command=self._editar_asignacion_empleado_departamento_gui
        ).grid(row=2, column=0, sticky="w", pady=2)

        tk.Button(
            lf_dep,
            text="Eliminar asignación empleado-departamento",
            width=35,
            command=self._eliminar_asignacion_empleado_departamento_gui
        ).grid(row=3, column=0, sticky="w", pady=2)

        # ------------------ Empleado ↔ Proyecto ------------------ #
        lf_pro = tk.LabelFrame(cont, text="Empleado ↔ Proyecto")
        lf_pro.grid(row=0, column=1, padx=10, pady=10, sticky="nw")

        tk.Button(
            lf_pro,
            text="Asignar empleado a proyecto",
            width=35,
            command=self._asignar_empleado_proyecto_gui
        ).grid(row=0, column=0, sticky="w", pady=2)

        tk.Button(
            lf_pro,
            text="Editar asignación (horas / tarea)",
            width=35,
            command=self._editar_empleado_proyecto_gui
        ).grid(row=1, column=0, sticky="w", pady=2)

        tk.Button(
            lf_pro,
            text="Desasignar empleado de proyecto",
            width=35,
            command=self._desasignar_empleado_proyecto_gui
        ).grid(row=2, column=0, sticky="w", pady=2)

        tk.Button(
            lf_pro,
            text="Ver detalle empleado-proyecto",
            width=35,
            command=self._ver_detalle_empleado_proyecto_gui
        ).grid(row=3, column=0, sticky="w", pady=2)

    # ---------- ASIGNAR usuarios ----------

    def _asignar_empleado_departamento_gui(self):
        win = tk.Toplevel(self)
        win.title("Asignar empleado a departamento")
        win.grab_set()

        # --- Cargar empleados ---
        self.cursor.execute("""
            SELECT id_UsuarioSistema, Nombre, Apellido, NombreUsuario
            FROM UsuarioSistema
            ORDER BY id_UsuarioSistema
        """)
        empleados = self.cursor.fetchall()
        if not empleados:
            messagebox.showinfo("Sin datos", "No hay usuarios registrados.")
            win.destroy()
            return

        emp_map = {
            f"{e[0]} - {e[1]} {e[2]} ({e[3]})": e[0]
            for e in empleados
        }

        # --- Cargar departamentos ---
        self.cursor.execute("""
            SELECT id_Departamento, NombreDepartamento
            FROM Departamento
            ORDER BY id_Departamento
        """)
        departamentos = self.cursor.fetchall()
        if not departamentos:
            messagebox.showinfo("Sin datos", "No hay departamentos registrados.")
            win.destroy()
            return

        dep_map = {
            f"{d[0]} - {d[1]}": d[0]
            for d in departamentos
        }

        tk.Label(win, text="Empleado:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        cb_emp = ttk.Combobox(win, width=45, values=list(emp_map.keys()), state="readonly")
        cb_emp.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(win, text="Departamento:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        cb_dep = ttk.Combobox(win, width=45, values=list(dep_map.keys()), state="readonly")
        cb_dep.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        def asignar():
            sel_emp = cb_emp.get()
            sel_dep = cb_dep.get()
            if not sel_emp or not sel_dep:
                messagebox.showwarning("Faltan datos", "Seleccione empleado y departamento.")
                return

            id_emp = emp_map[sel_emp]
            id_dep = dep_map[sel_dep]

            try:
                # ¿Ya existe asignación?
                self.cursor.execute("""
                    SELECT id_EmpleadoDepartamento, Activo
                    FROM EmpleadoDepartamento
                    WHERE UsuarioSistema_idUsuarioSistema = %s
                    AND Departamento_idDepartamento = %s
                """, (id_emp, id_dep))
                row = self.cursor.fetchone()

                if row:
                    if row[1]:
                        messagebox.showinfo("Asignación existente",
                                            "El empleado ya está asignado a este departamento.")
                        return
                    else:
                        self.cursor.execute("""
                            UPDATE EmpleadoDepartamento
                            SET Activo = TRUE, FechaAsignacion = CURRENT_TIMESTAMP
                            WHERE id_EmpleadoDepartamento = %s
                        """, (row[0],))
                        self.cnx.commit()
                        messagebox.showinfo("OK", "Asignación reactivada correctamente.")
                        win.destroy()
                        return

                # Nueva asignación
                self.cursor.execute("""
                    INSERT INTO EmpleadoDepartamento
                        (Departamento_idDepartamento, UsuarioSistema_idUsuarioSistema)
                    VALUES (%s, %s)
                """, (id_dep, id_emp))
                self.cnx.commit()
                messagebox.showinfo("OK", "Empleado asignado correctamente.")
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error al asignar:\n{e}")

        tk.Button(win, text="Asignar", command=asignar)\
            .grid(row=2, column=1, padx=5, pady=10, sticky="e")

    def _listar_empleados_departamento_gui(self):
        win = tk.Toplevel(self)
        win.title("Empleados en departamentos")
        win.grab_set()

        txt = tk.Text(win, width=110, height=25, state="disabled")
        txt.pack(padx=10, pady=10)

        try:
            self.cursor.execute("""
                SELECT 
                    ED.id_EmpleadoDepartamento,
                    U.id_UsuarioSistema, U.Nombre, U.Apellido, U.NombreUsuario,
                    D.id_Departamento, D.NombreDepartamento,
                    R.tipo_Rol,
                    ED.Activo, ED.FechaAsignacion
                FROM EmpleadoDepartamento ED
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
                JOIN Rol R ON U.Rol_idRol = R.id_Rol
                ORDER BY D.NombreDepartamento, U.Nombre
            """)
            filas = self.cursor.fetchall()

            txt.config(state="normal")
            txt.delete("1.0", "end")

            if not filas:
                txt.insert("end", "No hay empleados asignados a departamentos.\n")
            else:
                for a in filas:
                    estado = "Activo" if a[8] else "Inactivo"
                    bloque = (
                        f"ID Asignación: {a[0]} | Fecha: {a[9]}\n"
                        f"Empleado: {a[2]} {a[3]} ({a[4]}) [ID {a[1]}]\n"
                        f"Departamento: {a[6]} (ID {a[5]})\n"
                        f"Rol: {a[7]} | Estado: {estado}\n"
                        + "-" * 80 + "\n"
                    )
                    txt.insert("end", bloque)

            txt.config(state="disabled")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al listar:\n{e}")

    def _editar_asignacion_empleado_departamento_gui(self):
        win = tk.Toplevel(self)
        win.title("Editar asignación empleado-departamento")
        win.grab_set()

        tk.Label(win, text="ID asignación:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        entry_id = tk.Entry(win, width=10)
        entry_id.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        modo = tk.StringVar(value="depto")
        tk.Radiobutton(win, text="Cambiar departamento", variable=modo, value="depto")\
            .grid(row=1, column=0, columnspan=2, sticky="w", padx=5)
        tk.Radiobutton(win, text="Activar / desactivar", variable=modo, value="estado")\
            .grid(row=2, column=0, columnspan=2, sticky="w", padx=5)

        # para cambio de depto
        tk.Label(win, text="Nuevo departamento:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
        cb_dep = ttk.Combobox(win, width=35, state="readonly")
        cb_dep.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        # precargar departamentos
        self.cursor.execute("SELECT id_Departamento, NombreDepartamento FROM Departamento ORDER BY id_Departamento")
        deps = self.cursor.fetchall()
        dep_map = {f"{d[0]} - {d[1]}": d[0] for d in deps}
        cb_dep["values"] = list(dep_map.keys())

        def aplicar():
            texto_id = entry_id.get().strip()
            if not texto_id.isdigit():
                messagebox.showwarning("ID inválido", "Ingrese un ID numérico de asignación.")
                return
            id_asig = int(texto_id)

            try:
                # datos actuales
                self.cursor.execute("""
                    SELECT id_EmpleadoDepartamento, Departamento_idDepartamento, UsuarioSistema_idUsuarioSistema, Activo
                    FROM EmpleadoDepartamento
                    WHERE id_EmpleadoDepartamento = %s
                """, (id_asig,))
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showerror("No encontrado", "No existe una asignación con ese ID.")
                    return

                id_dep_actual = row[1]
                id_usuario = row[2]
                estado_actual = row[3]

                if modo.get() == "estado":
                    nuevo_estado = not estado_actual
                    self.cursor.execute("""
                        UPDATE EmpleadoDepartamento
                        SET Activo = %s, FechaAsignacion = CURRENT_TIMESTAMP
                        WHERE id_EmpleadoDepartamento = %s
                    """, (nuevo_estado, id_asig))
                    self.cnx.commit()
                    messagebox.showinfo(
                        "OK",
                        f"Asignación {'activada' if nuevo_estado else 'desactivada'} correctamente."
                    )
                    win.destroy()
                    return

                # cambiar departamento
                sel = cb_dep.get()
                if not sel:
                    messagebox.showwarning("Falta dato", "Seleccione un nuevo departamento.")
                    return
                nuevo_dep = dep_map[sel]
                if nuevo_dep == id_dep_actual:
                    messagebox.showinfo("Sin cambios", "El empleado ya está en ese departamento.")
                    return

                # ¿ya existe esa combinación?
                self.cursor.execute("""
                    SELECT id_EmpleadoDepartamento, Activo
                    FROM EmpleadoDepartamento
                    WHERE UsuarioSistema_idUsuarioSistema = %s
                    AND Departamento_idDepartamento = %s
                """, (id_usuario, nuevo_dep))
                existe = self.cursor.fetchone()

                if existe:
                    if existe[1]:
                        messagebox.showerror(
                            "Asignación existente",
                            "El empleado ya está activo en ese departamento."
                        )
                        return
                    else:
                        self.cursor.execute("""
                            UPDATE EmpleadoDepartamento
                            SET Activo = TRUE, FechaAsignacion = CURRENT_TIMESTAMP
                            WHERE id_EmpleadoDepartamento = %s
                        """, (existe[0],))
                        self.cnx.commit()
                        messagebox.showinfo("OK", "Asignación reactivada en el nuevo departamento.")
                        win.destroy()
                        return

                # actualizar depto
                self.cursor.execute("""
                    UPDATE EmpleadoDepartamento
                    SET Departamento_idDepartamento = %s, FechaAsignacion = CURRENT_TIMESTAMP
                    WHERE id_EmpleadoDepartamento = %s
                """, (nuevo_dep, id_asig))
                self.cnx.commit()
                messagebox.showinfo("OK", "Departamento actualizado correctamente.")
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error:\n{e}")

        tk.Button(win, text="Aplicar cambios", command=aplicar)\
            .grid(row=4, column=1, padx=5, pady=10, sticky="e")

    def _eliminar_asignacion_empleado_departamento_gui(self):
        win = tk.Toplevel(self)
        win.title("Eliminar asignación empleado-departamento")
        win.grab_set()

        tk.Label(win, text="ID de asignación a eliminar:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        entry_id = tk.Entry(win, width=10)
        entry_id.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        def ejecutar():
            texto_id = entry_id.get().strip()
            if not texto_id.isdigit():
                messagebox.showwarning("ID inválido", "Ingrese un ID numérico.")
                return
            id_asig = int(texto_id)

            try:
                # datos de la asignación
                self.cursor.execute("""
                    SELECT 
                        ED.id_EmpleadoDepartamento,
                        U.Nombre, U.Apellido,
                        D.NombreDepartamento
                    FROM EmpleadoDepartamento ED
                    JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                    JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
                    WHERE ED.id_EmpleadoDepartamento = %s
                """, (id_asig,))
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showerror("No encontrado", "No existe esa asignación.")
                    return

                # ¿tiene proyectos?
                self.cursor.execute("""
                    SELECT COUNT(*)
                    FROM EmpleadoProyecto
                    WHERE EmpleadoDepartamento_idEmpleadoDepartamento = %s
                """, (id_asig,))
                n_proj = self.cursor.fetchone()[0]
                if n_proj > 0:
                    messagebox.showerror(
                        "Dependencias",
                        f"No se puede eliminar: el empleado participa en {n_proj} proyecto(s)."
                    )
                    return

                if not messagebox.askyesno(
                    "Confirmar",
                    f"¿Eliminar la asignación de {row[1]} {row[2]} "
                    f"del departamento '{row[3]}'?"
                ):
                    return

                self.cursor.execute(
                    "DELETE FROM EmpleadoDepartamento WHERE id_EmpleadoDepartamento = %s",
                    (id_asig,)
                )
                self.cnx.commit()
                messagebox.showinfo("OK", "Asignación eliminada correctamente.")
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error:\n{e}")

        tk.Button(win, text="Eliminar", command=ejecutar)\
            .grid(row=1, column=1, padx=5, pady=10, sticky="e")

    # =========================================================
    #  EMPLEADO ↔ PROYECTO (GUI)
    # =========================================================

    def _asignar_empleado_proyecto_gui(self):
        win = tk.Toplevel(self)
        win.title("Asignar empleado a proyecto")
        win.grab_set()

        # Empleados activos en algún departamento
        self.cursor.execute("""
            SELECT 
                ED.id_EmpleadoDepartamento,
                U.Nombre, U.Apellido, U.NombreUsuario,
                D.NombreDepartamento
            FROM EmpleadoDepartamento ED
            JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
            JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
            WHERE ED.Activo = TRUE
            ORDER BY U.Nombre
        """)
        empleados = self.cursor.fetchall()
        if not empleados:
            messagebox.showinfo("Sin datos", "No hay empleados activos en departamentos.")
            win.destroy()
            return

        emp_map = {
            f"{e[0]} - {e[1]} {e[2]} ({e[3]}) / {e[4]}": e[0]
            for e in empleados
        }

        # Proyectos
        self.cursor.execute("SELECT id_Proyecto, NombreProyecto FROM Proyecto ORDER BY id_Proyecto")
        proyectos = self.cursor.fetchall()
        if not proyectos:
            messagebox.showinfo("Sin datos", "No hay proyectos registrados.")
            win.destroy()
            return

        proy_map = {
            f"{p[0]} - {p[1]}": p[0]
            for p in proyectos
        }

        tk.Label(win, text="Empleado (EmpleadoDepartamento):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        cb_emp = ttk.Combobox(win, width=55, values=list(emp_map.keys()), state="readonly")
        cb_emp.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(win, text="Proyecto:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        cb_proy = ttk.Combobox(win, width=40, values=list(proy_map.keys()), state="readonly")
        cb_proy.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        tk.Label(win, text="Horas asignadas:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        entry_horas = tk.Entry(win, width=10)
        entry_horas.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        tk.Label(win, text="Descripción tarea (opcional):").grid(row=3, column=0, padx=5, pady=5, sticky="ne")
        txt_desc = tk.Text(win, width=50, height=4)
        txt_desc.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        def asignar():
            sel_emp = cb_emp.get().strip()
            sel_proy = cb_proy.get().strip()
            horas_txt = entry_horas.get().strip()

            if not sel_emp or not sel_proy:
                messagebox.showwarning("Faltan datos", "Seleccione empleado y proyecto.")
                return

            if not horas_txt.isdigit() or int(horas_txt) <= 0:
                messagebox.showwarning("Horas inválidas", "Ingrese un número entero mayor a 0.")
                return

            horas = int(horas_txt)
            desc = txt_desc.get("1.0", "end").strip() or None
            id_emp_dep = emp_map[sel_emp]
            id_proy = proy_map[sel_proy]

            try:
                # ¿Existe ya la asignación?
                self.cursor.execute("""
                    SELECT id_DetalleProyecto, Activo
                    FROM EmpleadoProyecto
                    WHERE EmpleadoDepartamento_idEmpleadoDepartamento = %s
                    AND Proyecto_idProyecto = %s
                """, (id_emp_dep, id_proy))
                row = self.cursor.fetchone()

                if row:
                    if row[1]:
                        messagebox.showinfo("Asignación existente",
                                            "El empleado ya está asignado activamente a este proyecto.")
                        return
                    else:
                        self.cursor.execute("""
                            UPDATE EmpleadoProyecto
                            SET Activo = TRUE,
                                FechaProyectoInscrito = CURRENT_TIMESTAMP,
                                CantidadHorasEmpleadoProyecto = %s,
                                DescripcionTareaProyecto = %s
                            WHERE id_DetalleProyecto = %s
                        """, (horas, desc, row[0]))
                        self.cnx.commit()
                        messagebox.showinfo("OK", "Asignación reactivada correctamente.")
                        win.destroy()
                        return

                # Nueva asignación
                self.cursor.execute("""
                    INSERT INTO EmpleadoProyecto
                    (EmpleadoDepartamento_idEmpleadoDepartamento, Proyecto_idProyecto,
                     CantidadHorasEmpleadoProyecto, DescripcionTareaProyecto)
                    VALUES (%s, %s, %s, %s)
                """, (id_emp_dep, id_proy, horas, desc))
                self.cnx.commit()
                messagebox.showinfo("OK", "Empleado asignado al proyecto correctamente.")
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error al asignar:\n{e}")

        tk.Button(win, text="Asignar", command=asignar)\
            .grid(row=4, column=1, padx=5, pady=10, sticky="e")

    def _editar_empleado_proyecto_gui(self):
        win = tk.Toplevel(self)
        win.title("Editar asignación empleado-proyecto")
        win.grab_set()

        tk.Label(win, text="ID DetalleProyecto:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        entry_id = tk.Entry(win, width=10)
        entry_id.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Label(win, text="Nuevas horas:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        entry_horas = tk.Entry(win, width=10)
        entry_horas.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        tk.Label(win, text="Nueva descripción:").grid(row=2, column=0, padx=5, pady=5, sticky="ne")
        txt_desc = tk.Text(win, width=50, height=4)
        txt_desc.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        def cargar_y_editar():
            txt_id = entry_id.get().strip()
            if not txt_id.isdigit():
                messagebox.showwarning("ID inválido", "Ingrese un ID numérico de detalle.")
                return
            id_det = int(txt_id)

            try:
                self.cursor.execute("""
                    SELECT 
                        EP.id_DetalleProyecto,
                        U.Nombre, U.Apellido,
                        P.NombreProyecto,
                        EP.CantidadHorasEmpleadoProyecto,
                        EP.DescripcionTareaProyecto
                    FROM EmpleadoProyecto EP
                    JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                    JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                    JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                    WHERE EP.Activo = TRUE AND EP.id_DetalleProyecto = %s
                """, (id_det,))
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showerror("No encontrado", "No existe una asignación activa con ese ID.")
                    return

                # Datos actuales
                horas_actual = row[4]
                desc_actual = row[5] or ""

                # Leer entradas
                horas_txt = entry_horas.get().strip()
                nueva_desc = txt_desc.get("1.0", "end").strip()

                if not horas_txt:
                    horas = horas_actual
                else:
                    if not horas_txt.isdigit() or int(horas_txt) <= 0:
                        messagebox.showwarning("Horas inválidas", "Ingrese un entero mayor a 0.")
                        return
                    horas = int(horas_txt)

                if not nueva_desc:
                    nueva_desc = desc_actual

                if not messagebox.askyesno(
                    "Confirmar",
                    f"Actualizar asignación de {row[1]} {row[2]} en proyecto '{row[3]}'?"
                ):
                    return

                self.cursor.execute("""
                    UPDATE EmpleadoProyecto
                    SET CantidadHorasEmpleadoProyecto = %s,
                        DescripcionTareaProyecto = %s
                    WHERE id_DetalleProyecto = %s
                """, (horas, nueva_desc, id_det))
                self.cnx.commit()
                messagebox.showinfo("OK", "Asignación actualizada correctamente.")
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error:\n{e}")

        tk.Button(win, text="Actualizar", command=cargar_y_editar)\
            .grid(row=3, column=1, padx=5, pady=10, sticky="e")

    def _desasignar_empleado_proyecto_gui(self):
        win = tk.Toplevel(self)
        win.title("Desasignar empleado de proyecto")
        win.grab_set()

        tk.Label(win, text="ID DetalleProyecto a desasignar:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        entry_id = tk.Entry(win, width=10)
        entry_id.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        def ejecutar():
            txt_id = entry_id.get().strip()
            if not txt_id.isdigit():
                messagebox.showwarning("ID inválido", "Ingrese un ID numérico.")
                return
            id_det = int(txt_id)

            try:
                self.cursor.execute("""
                    SELECT 
                        EP.id_DetalleProyecto,
                        U.Nombre, U.Apellido,
                        P.NombreProyecto
                    FROM EmpleadoProyecto EP
                    JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                    JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                    JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                    WHERE EP.id_DetalleProyecto = %s AND EP.Activo = TRUE
                """, (id_det,))
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showerror("No encontrado", "No existe una asignación activa con ese ID.")
                    return

                if not messagebox.askyesno(
                    "Confirmar",
                    f"¿Desasignar a {row[1]} {row[2]} del proyecto '{row[3]}'?"
                ):
                    return

                self.cursor.execute("""
                    UPDATE EmpleadoProyecto
                    SET Activo = FALSE
                    WHERE id_DetalleProyecto = %s
                """, (id_det,))
                self.cnx.commit()
                messagebox.showinfo("OK", "Empleado desasignado del proyecto.")
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error:\n{e}")

        tk.Button(win, text="Desasignar", command=ejecutar)\
            .grid(row=1, column=1, padx=5, pady=10, sticky="e")

    def _ver_detalle_empleado_proyecto_gui(self):
        win = tk.Toplevel(self)
        win.title("Detalle empleado-proyecto")
        win.grab_set()

        txt = tk.Text(win, width=120, height=30, state="disabled")
        txt.pack(padx=10, pady=10)

        try:
            self.cursor.execute("""
                SELECT 
                    EP.id_DetalleProyecto,
                    U.Nombre, U.Apellido, U.NombreUsuario,
                    P.NombreProyecto,
                    EP.FechaProyectoInscrito,
                    EP.CantidadHorasEmpleadoProyecto,
                    EP.DescripcionTareaProyecto,
                    CASE WHEN EP.Activo = TRUE THEN 'Activo' ELSE 'Inactivo' END AS Estado,
                    Dpto.NombreDepartamento
                FROM EmpleadoProyecto EP
                JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                JOIN Departamento Dpto ON ED.Departamento_idDepartamento = Dpto.id_Departamento
                ORDER BY EP.id_DetalleProyecto
            """)
            filas = self.cursor.fetchall()

            txt.config(state="normal")
            txt.delete("1.0", "end")

            if not filas:
                txt.insert("end", "No hay asignaciones registradas.\n")
            else:
                for a in filas:
                    bloque = (
                        f"ID Detalle:      {a[0]}\n"
                        f"Empleado:        {a[1]} {a[2]} ({a[3]})\n"
                        f"Departamento:    {a[9]}\n"
                        f"Proyecto:        {a[4]}\n"
                        f"Fecha inscripción: {a[5]}\n"
                        f"Horas:           {a[6]}\n"
                        f"Descripción:     {a[7] if a[7] else '(sin descripción)'}\n"
                        f"Estado:          {a[8]}\n"
                        + "-" * 90 + "\n"
                    )
                    txt.insert("end", bloque)

            txt.config(state="disabled")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al cargar el detalle:\n{e}")









    # =========================================================
    # PANEL INFORMES (ADMIN) - VERSION GUI
    # =========================================================
    def _show_informes(self):
        # Limpiar panel central y usarlo como contenedor
        self._clear_central()
        cont = self.panel_central

        tk.Label(
            cont,
            text="Gestión de Informes",
            font=("Arial", 13, "bold")
        ).grid(row=0, column=0, sticky="w", pady=(0, 5))

        tk.Label(
            cont,
            text="Seleccione una acción:",
            font=("Arial", 10)
        ).grid(row=1, column=0, sticky="w", pady=(0, 10))

        # 1) Informe Empleado-Proyecto (general / por proyecto / por departamento)
        tk.Button(
            cont,
            text="Generar informe Empleado-Proyecto",
            width=40,
            command=self._generar_informe_ep
        ).grid(row=2, column=0, sticky="w", pady=3)

        # 2) Informe por empleado
        tk.Button(
            cont,
            text="Generar informe por Empleado",
            width=40,
            command=self._generar_informe_por_empleado_gui
        ).grid(row=3, column=0, sticky="w", pady=3)

        # 3) Listar informes
        tk.Button(
            cont,
            text="Listar informes generados",
            width=40,
            command=self._listar_informes_gui
        ).grid(row=4, column=0, sticky="w", pady=3)

        # 4) Ver detalle informe
        tk.Button(
            cont,
            text="Ver detalle de informe",
            width=40,
            command=self._detalle_informe_gui
        ).grid(row=5, column=0, sticky="w", pady=3)

        # 5) Eliminar informe
        tk.Button(
            cont,
            text="Eliminar informe",
            width=40,
            command=self._eliminar_informe_gui
        ).grid(row=6, column=0, sticky="w", pady=3)

        # 6) Exportar Excel rápido (sin registrar)
        tk.Button(
            cont,
            text="Exportar Excel rápido (sin registrar)",
            width=40,
            command=self._exportar_excel_empleado_proyecto_gui
        ).grid(row=7, column=0, sticky="w", pady=10)

    # -------------------- 1) GENERAR INFORME EP --------------------

    def _generar_informe_ep(self):
        """
        Ventana para elegir tipo de informe Empleado-Proyecto:
        - General
        - Por proyecto
        - Por departamento
        """
        win = tk.Toplevel(self)
        win.title("Generar informe Empleado-Proyecto")
        win.grab_set()

        tk.Label(win, text="Seleccione tipo de informe:").grid(row=0, column=0, columnspan=2, pady=5, padx=5)

        modo = tk.StringVar(value="general")

        rb_general = tk.Radiobutton(win, text="General", variable=modo, value="general")
        rb_general.grid(row=1, column=0, sticky="w", padx=10)

        rb_proyecto = tk.Radiobutton(win, text="Por proyecto", variable=modo, value="proyecto")
        rb_proyecto.grid(row=2, column=0, sticky="w", padx=10)

        rb_dep = tk.Radiobutton(win, text="Por departamento", variable=modo, value="departamento")
        rb_dep.grid(row=3, column=0, sticky="w", padx=10)

        # Combobox para proyectos
        tk.Label(win, text="Proyecto:").grid(row=2, column=1, padx=5, pady=2, sticky="e")
        cb_pro = ttk.Combobox(win, state="disabled", width=35)
        cb_pro.grid(row=2, column=2, padx=5, pady=2, sticky="w")

        # Combobox para departamentos
        tk.Label(win, text="Departamento:").grid(row=3, column=1, padx=5, pady=2, sticky="e")
        cb_dep = ttk.Combobox(win, state="disabled", width=35)
        cb_dep.grid(row=3, column=2, padx=5, pady=2, sticky="w")

        # Cargar proyectos y departamentos para comboboxes
        try:
            self.cursor.execute("SELECT id_Proyecto, NombreProyecto FROM Proyecto ORDER BY id_Proyecto")
            proyectos = self.cursor.fetchall()
            proj_map = {f"{p[0]} - {p[1]}": p[0] for p in proyectos}
            cb_pro["values"] = list(proj_map.keys())

            self.cursor.execute("SELECT id_Departamento, NombreDepartamento FROM Departamento ORDER BY id_Departamento")
            departamentos = self.cursor.fetchall()
            dep_map = {f"{d[0]} - {d[1]}": d[0] for d in departamentos}
            cb_dep["values"] = list(dep_map.keys())
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar proyectos/departamentos:\n{e}")
            win.destroy()
            return

        def actualizar_estado():
            m = modo.get()
            if m == "proyecto":
                cb_pro.config(state="readonly")
                cb_dep.config(state="disabled")
                cb_dep.set("")
            elif m == "departamento":
                cb_dep.config(state="readonly")
                cb_pro.config(state="disabled")
                cb_pro.set("")
            else:
                cb_pro.config(state="disabled")
                cb_dep.config(state="disabled")
                cb_pro.set("")
                cb_dep.set("")

        rb_general.config(command=actualizar_estado)
        rb_proyecto.config(command=actualizar_estado)
        rb_dep.config(command=actualizar_estado)

        def generar():
            m = modo.get()
            id_obj = None

            if m == "proyecto":
                sel = cb_pro.get()
                if not sel:
                    messagebox.showwarning("Faltan datos", "Seleccione un proyecto.")
                    return
                id_obj = proj_map[sel]

            elif m == "departamento":
                sel = cb_dep.get()
                if not sel:
                    messagebox.showwarning("Faltan datos", "Seleccione un departamento.")
                    return
                id_obj = dep_map[sel]

            try:
                self._generar_informe_empleado_proyecto_gui(m, id_obj)
                win.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error al generar el informe:\n{e}")

        tk.Button(win, text="Generar", command=generar)\
            .grid(row=4, column=2, pady=10, sticky="e")

    def _generar_informe_empleado_proyecto_gui(self, modo, id_obj):
        """
        Versión GUI de generar_informe_empleado_proyecto.
        NO usa pedir_input ni print. Hace todo vía SQL y messagebox.
        """
        # Determinar filtro SQL
        if modo == "proyecto":
            filtro_sql = "WHERE P.id_Proyecto = %s"
            params = (id_obj,)
            tipo_informe = "Por Proyecto"
        elif modo == "departamento":
            filtro_sql = "WHERE D.id_Departamento = %s"
            params = (id_obj,)
            tipo_informe = "Por Departamento"
        else:
            filtro_sql = ""
            params = ()
            tipo_informe = "General"

        # Consulta principal (misma que en Admin.generar_informe_empleado_proyecto)
        self.cursor.execute(f"""
            SELECT 
                EP.id_DetalleProyecto,
                CONCAT(U.Nombre, ' ', U.Apellido) AS Empleado,
                D.NombreDepartamento,
                P.NombreProyecto,
                EP.FechaProyectoInscrito,
                EP.CantidadHorasEmpleadoProyecto,
                EP.DescripcionTareaProyecto,
                CASE WHEN EP.Activo THEN 'Activo' ELSE 'Inactivo' END AS Estado
            FROM EmpleadoProyecto EP
            JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
            JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
            JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
            JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
            {filtro_sql}
            ORDER BY D.NombreDepartamento, P.NombreProyecto, U.Nombre
        """, params)
        datos = self.cursor.fetchall()

        if not datos:
            messagebox.showinfo("Sin datos", "No se encontraron registros para generar el informe.")
            return

        # Crear carpeta ./informes si no existe
        os.makedirs("./informes", exist_ok=True)

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Empleados-Proyectos"

        encabezados = [
            "ID Detalle", "Empleado", "Departamento", "Proyecto",
            "Fecha Inscripción", "Horas Asignadas", "Descripción", "Estado"
        ]
        ws.append(encabezados)

        for fila in datos:
            ws.append(list(fila))

        nombre_archivo = f"informe_{tipo_informe.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta = os.path.join("./informes", nombre_archivo)
        wb.save(ruta)

        # Registrar en tabla Informe (usamos primer id_DetalleProyecto como referencia)
        id_detalle = datos[0][0]
        descripcion = f"Informe {tipo_informe.lower()} generado con {len(datos)} registros."

        self.cursor.execute("""
            INSERT INTO Informe (NombreInforme, FechaConsulta, DescripcionInforme, TipoInforme, EmpleadoProyecto_idDetalleProyecto)
            VALUES (%s, NOW(), %s, %s, %s)
        """, (nombre_archivo, descripcion, tipo_informe, id_detalle))
        self.cnx.commit()

        messagebox.showinfo(
            "Informe generado",
            f"Informe '{tipo_informe}' generado correctamente.\n"
            f"Archivo: {ruta}\nRegistros: {len(datos)}"
        )

    # -------------------- 2) INFORME POR EMPLEADO --------------------

    def _generar_informe_por_empleado_gui(self):
        """
        Versión GUI de generar_informe_por_empleado.
        Pide empleado con Combobox y genera el Excel + registro.
        """
        win = tk.Toplevel(self)
        win.title("Informe por empleado")
        win.grab_set()

        tk.Label(win, text="Seleccione empleado:").grid(row=0, column=0, padx=5, pady=5, sticky="e")

        try:
            self.cursor.execute("""
                SELECT U.id_UsuarioSistema, CONCAT(U.Nombre, ' ', U.Apellido) AS Empleado
                FROM UsuarioSistema U
                JOIN Rol R ON U.Rol_idRol = R.id_Rol
                WHERE R.tipo_Rol = 'empleado'
                ORDER BY U.Apellido
            """)
            empleados = self.cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar empleados:\n{e}")
            win.destroy()
            return

        if not empleados:
            messagebox.showinfo("Sin datos", "No hay empleados registrados.")
            win.destroy()
            return

        emp_map = {f"{e[0]} - {e[1]}": e[0] for e in empleados}
        cb_emp = ttk.Combobox(win, state="readonly", width=40, values=list(emp_map.keys()))
        cb_emp.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        def generar():
            sel = cb_emp.get()
            if not sel:
                messagebox.showwarning("Faltan datos", "Seleccione un empleado.")
                return

            id_emp = emp_map[sel]

            try:
                # Datos personales
                self.cursor.execute("""
                    SELECT 
                        U.Nombre, U.Apellido, U.RUT, U.Email, U.Telefono,
                        D.Calle, D.Numero, D.Ciudad, D.Region, D.Pais, D.CodigoPostal
                    FROM UsuarioSistema U
                    LEFT JOIN Direccion D ON U.Direccion_idDireccion = D.id_Direccion
                    WHERE U.id_UsuarioSistema = %s
                """, (id_emp,))
                datos_personales = self.cursor.fetchone()

                if not datos_personales:
                    messagebox.showerror("Error", "No se encontró información personal del empleado.")
                    return

                # Proyectos
                self.cursor.execute("""
                    SELECT 
                        EP.id_DetalleProyecto,
                        P.NombreProyecto,
                        Dep.NombreDepartamento,
                        EP.FechaProyectoInscrito,
                        EP.CantidadHorasEmpleadoProyecto,
                        EP.DescripcionTareaProyecto,
                        CASE WHEN EP.Activo THEN 'Activo' ELSE 'Inactivo' END AS Estado
                    FROM EmpleadoProyecto EP
                    JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                    JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                    JOIN Departamento Dep ON ED.Departamento_idDepartamento = Dep.id_Departamento
                    WHERE ED.UsuarioSistema_idUsuarioSistema = %s
                    ORDER BY EP.FechaProyectoInscrito DESC
                """, (id_emp,))
                proyectos = self.cursor.fetchall()

                if not proyectos:
                    messagebox.showinfo(
                        "Sin proyectos",
                        "El empleado no tiene proyectos asignados. No se generará informe."
                    )
                    return

                id_detalle = proyectos[0][0]

                os.makedirs("./informes", exist_ok=True)

                wb = Workbook()
                ws = wb.active
                ws.title = "Informe Empleado"

                ws.append(["INFORME DE EMPLEADO"])
                ws.append([])

                encabezados_personales = [
                    "Nombre", "Apellido", "RUT", "Email", "Teléfono",
                    "Calle", "Número", "Ciudad", "Región", "País", "Código Postal"
                ]
                ws.append(encabezados_personales)
                ws.append(list(datos_personales))
                ws.append([])

                ws.append(["PROYECTOS ASOCIADOS"])
                ws.append([])
                encabezados_proyectos = [
                    "ID Detalle", "Proyecto", "Departamento", "Fecha Inscripción",
                    "Horas Asignadas", "Descripción Tarea", "Estado"
                ]
                ws.append(encabezados_proyectos)

                for fila in proyectos:
                    ws.append(list(fila))

                nombre_archivo = f"informe_empleado_{id_emp}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                ruta = os.path.join("./informes", nombre_archivo)
                wb.save(ruta)

                descripcion = f"Informe individual del empleado ID {id_emp}, con {len(proyectos)} proyectos listados."

                self.cursor.execute("""
                    INSERT INTO Informe (NombreInforme, FechaConsulta, DescripcionInforme, TipoInforme, EmpleadoProyecto_idDetalleProyecto)
                    VALUES (%s, NOW(), %s, %s, %s)
                """, (nombre_archivo, descripcion, "Por Empleado", id_detalle))
                self.cnx.commit()

                messagebox.showinfo(
                    "Informe generado",
                    f"Informe generado correctamente.\nArchivo: {ruta}\nProyectos incluidos: {len(proyectos)}"
                )
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error al generar informe:\n{e}")

        tk.Button(win, text="Generar informe", command=generar)\
            .grid(row=1, column=1, pady=10, sticky="e")

    # -------------------- 3) LISTAR INFORMES --------------------

    def _listar_informes_gui(self):
        win = tk.Toplevel(self)
        win.title("Listado de informes")
        win.geometry("900x500")
        win.grab_set()

        txt = tk.Text(win, width=120, height=30)
        txt.pack(fill="both", expand=True)

        try:
            self.cursor.execute("""
                SELECT id_Informe, NombreInforme, FechaConsulta, TipoInforme, DescripcionInforme
                FROM Informe
                ORDER BY FechaConsulta DESC
            """)
            informes = self.cursor.fetchall()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar informes:\n{e}")
            win.destroy()
            return

        if not informes:
            txt.insert("end", "No hay informes registrados.")
            txt.config(state="disabled")
            return

        for inf in informes:
            bloque = (
                f"ID: {inf[0]}\n"
                f"Nombre: {inf[1]}\n"
                f"Fecha: {inf[2]}\n"
                f"Tipo: {inf[3]}\n"
                f"Descripción: {inf[4]}\n"
                + "-"*80 + "\n"
            )
            txt.insert("end", bloque)

        txt.config(state="disabled")

    # -------------------- 4) DETALLE INFORME --------------------

    def _detalle_informe_gui(self):
        win = tk.Toplevel(self)
        win.title("Ver detalle informe")
        win.grab_set()

        tk.Label(win, text="ID de informe:").grid(row=0, column=0, pady=5, padx=5)
        entry = tk.Entry(win, width=10)
        entry.grid(row=0, column=1, pady=5, padx=5)

        result = tk.Text(win, width=80, height=15)
        result.grid(row=1, column=0, columnspan=3, padx=5, pady=5)

        def buscar():
            id_inf = entry.get().strip()
            if not id_inf.isdigit():
                messagebox.showwarning("Error", "Ingrese un ID válido")
                return

            self.cursor.execute("""
                SELECT id_Informe, NombreInforme, FechaConsulta, DescripcionInforme, TipoInforme
                FROM Informe
                WHERE id_Informe = %s
            """, (int(id_inf),))
            inf = self.cursor.fetchone()

            result.config(state="normal")
            result.delete("1.0", "end")

            if not inf:
                result.insert("end", "Informe no encontrado.")
            else:
                result.insert(
                    "end",
                    f"ID: {inf[0]}\n"
                    f"Nombre: {inf[1]}\n"
                    f"Fecha: {inf[2]}\n"
                    f"Tipo: {inf[4]}\n"
                    f"Descripción: {inf[3]}\n"
                )

            result.config(state="disabled")

        tk.Button(win, text="Buscar", command=buscar)\
            .grid(row=0, column=2, padx=5)

    # -------------------- 5) ELIMINAR INFORME --------------------

    def _eliminar_informe_gui(self):
        win = tk.Toplevel(self)
        win.title("Eliminar informe")
        win.grab_set()

        tk.Label(win, text="ID de informe a eliminar:").grid(row=0, column=0, pady=5, padx=5, sticky="e")
        entry = tk.Entry(win, width=10)
        entry.grid(row=0, column=1, pady=5, padx=5, sticky="w")

        def eliminar():
            id_inf = entry.get().strip()
            if not id_inf.isdigit():
                messagebox.showwarning("Error", "Ingrese un ID numérico válido")
                return
            id_inf_int = int(id_inf)

            try:
                # Obtener nombre archivo
                self.cursor.execute("SELECT NombreInforme FROM Informe WHERE id_Informe = %s", (id_inf_int,))
                row = self.cursor.fetchone()
                if not row:
                    messagebox.showerror("No encontrado", "No existe un informe con ese ID.")
                    return

                nombre = row[0]
                ruta = os.path.join("./informes", nombre)

                if not messagebox.askyesno(
                    "Confirmar",
                    f"Se eliminará el informe '{nombre}' y su registro.\n¿Desea continuar?"
                ):
                    return

                # Eliminar archivo físico si existe
                if os.path.exists(ruta):
                    os.remove(ruta)

                # Eliminar registro BD
                self.cursor.execute("DELETE FROM Informe WHERE id_Informe = %s", (id_inf_int,))
                self.cnx.commit()

                messagebox.showinfo("OK", f"Informe '{nombre}' eliminado correctamente.")
                win.destroy()

            except Exception as e:
                self.cnx.rollback()
                messagebox.showerror("Error", f"Ocurrió un error al eliminar informe:\n{e}")

        tk.Button(win, text="Eliminar", command=eliminar)\
            .grid(row=1, column=1, pady=10, sticky="e")

    # -------------------- 6) EXPORTAR EXCEL RÁPIDO --------------------

    def _exportar_excel_empleado_proyecto_gui(self):
        """
        Versión GUI de exportar_excel_empleado_proyecto.
        Igual que generar_informe_ep, pero NO registra en tabla Informe.
        """
        win = tk.Toplevel(self)
        win.title("Exportar Excel Empleado-Proyecto")
        win.grab_set()

        tk.Label(win, text="Seleccione tipo de exportación:").grid(row=0, column=0, columnspan=2, pady=5, padx=5)

        modo = tk.StringVar(value="general")

        rb_general = tk.Radiobutton(win, text="General", variable=modo, value="general")
        rb_general.grid(row=1, column=0, sticky="w", padx=10)

        rb_proyecto = tk.Radiobutton(win, text="Por proyecto", variable=modo, value="proyecto")
        rb_proyecto.grid(row=2, column=0, sticky="w", padx=10)

        rb_dep = tk.Radiobutton(win, text="Por departamento", variable=modo, value="departamento")
        rb_dep.grid(row=3, column=0, sticky="w", padx=10)

        tk.Label(win, text="Proyecto:").grid(row=2, column=1, padx=5, pady=2, sticky="e")
        cb_pro = ttk.Combobox(win, state="disabled", width=35)
        cb_pro.grid(row=2, column=2, padx=5, pady=2, sticky="w")

        tk.Label(win, text="Departamento:").grid(row=3, column=1, padx=5, pady=2, sticky="e")
        cb_dep = ttk.Combobox(win, state="disabled", width=35)
        cb_dep.grid(row=3, column=2, padx=5, pady=2, sticky="w")

        try:
            self.cursor.execute("SELECT id_Proyecto, NombreProyecto FROM Proyecto ORDER BY id_Proyecto")
            proyectos = self.cursor.fetchall()
            proj_map = {f"{p[0]} - {p[1]}": p[0] for p in proyectos}
            cb_pro["values"] = list(proj_map.keys())

            self.cursor.execute("SELECT id_Departamento, NombreDepartamento FROM Departamento ORDER BY id_Departamento")
            departamentos = self.cursor.fetchall()
            dep_map = {f"{d[0]} - {d[1]}": d[0] for d in departamentos}
            cb_dep["values"] = list(dep_map.keys())
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar proyectos/departamentos:\n{e}")
            win.destroy()
            return

        def actualizar_estado():
            m = modo.get()
            if m == "proyecto":
                cb_pro.config(state="readonly")
                cb_dep.config(state="disabled")
                cb_dep.set("")
            elif m == "departamento":
                cb_dep.config(state="readonly")
                cb_pro.config(state="disabled")
                cb_pro.set("")
            else:
                cb_pro.config(state="disabled")
                cb_dep.config(state="disabled")
                cb_pro.set("")
                cb_dep.set("")

        rb_general.config(command=actualizar_estado)
        rb_proyecto.config(command=actualizar_estado)
        rb_dep.config(command=actualizar_estado)

        def exportar():
            m = modo.get()
            id_obj = None

            if m == "proyecto":
                sel = cb_pro.get()
                if not sel:
                    messagebox.showwarning("Faltan datos", "Seleccione un proyecto.")
                    return
                id_obj = proj_map[sel]
            elif m == "departamento":
                sel = cb_dep.get()
                if not sel:
                    messagebox.showwarning("Faltan datos", "Seleccione un departamento.")
                    return
                id_obj = dep_map[sel]

            # Determinar filtro
            if m == "proyecto":
                filtro_sql = "WHERE P.id_Proyecto = %s"
                params = (id_obj,)
                tipo = "Por Proyecto"
            elif m == "departamento":
                filtro_sql = "WHERE D.id_Departamento = %s"
                params = (id_obj,)
                tipo = "Por Departamento"
            else:
                filtro_sql = ""
                params = ()
                tipo = "General"

            try:
                self.cursor.execute(f"""
                    SELECT 
                        EP.id_DetalleProyecto,
                        CONCAT(U.Nombre, ' ', U.Apellido) AS Empleado,
                        D.NombreDepartamento,
                        P.NombreProyecto,
                        EP.FechaProyectoInscrito,
                        EP.CantidadHorasEmpleadoProyecto,
                        EP.DescripcionTareaProyecto,
                        CASE WHEN EP.Activo THEN 'Activo' ELSE 'Inactivo' END AS Estado
                    FROM EmpleadoProyecto EP
                    JOIN EmpleadoDepartamento ED ON EP.EmpleadoDepartamento_idEmpleadoDepartamento = ED.id_EmpleadoDepartamento
                    JOIN UsuarioSistema U ON ED.UsuarioSistema_idUsuarioSistema = U.id_UsuarioSistema
                    JOIN Proyecto P ON EP.Proyecto_idProyecto = P.id_Proyecto
                    JOIN Departamento D ON ED.Departamento_idDepartamento = D.id_Departamento
                    {filtro_sql}
                    ORDER BY D.NombreDepartamento, P.NombreProyecto, U.Nombre
                """, params)
                resultados = self.cursor.fetchall()

                if not resultados:
                    messagebox.showinfo("Sin datos", "No se encontraron registros para exportar.")
                    return

                os.makedirs("./informes", exist_ok=True)

                wb = Workbook()
                ws = wb.active
                ws.title = "Empleados-Proyectos"

                encabezados = [
                    "ID Detalle", "Empleado", "Departamento",
                    "Proyecto", "Fecha Inscripción",
                    "Horas Asignadas", "Descripción", "Estado"
                ]
                ws.append(encabezados)

                for fila in resultados:
                    ws.append(list(fila))

                nombre_archivo = f"export_{tipo.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                ruta = os.path.join("./informes", nombre_archivo)
                wb.save(ruta)

                messagebox.showinfo(
                    "Exportación completada",
                    f"Archivo guardado en: {ruta}\nRegistros exportados: {len(resultados)}"
                )
                win.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Error al exportar Excel:\n{e}")

        tk.Button(win, text="Exportar", command=exportar)\
            .grid(row=4, column=2, pady=10, sticky="e")






    # =========================================================
    # HERRAMIENTAS BD de Administrador
    # =========================================================
    def _show_bd_tools(self):
        from database.database import crear_admin, crear_datos_ejemplo, limpiarBaseDeDatos

        self._clear_central()

        titulo = tk.Label(
            self.panel_central,
            text="Herramientas de Base de Datos",
            font=("Arial", 13, "bold"),
        )
        titulo.pack(padx=10, pady=(10, 5), anchor="nw")

        desc = tk.Label(
            self.panel_central,
            text="Usa estas opciones con cuidado. Afectan directamente los datos del sistema.",
            font=("Arial", 10),
            fg="#555",
        )
        desc.pack(padx=10, pady=(0, 15), anchor="nw")

        def accion_crear_admin():
            try:
                crear_admin(self.cnx, self.cursor)
                messagebox.showinfo("OK", "Administrador por defecto verificado/creado.")
            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error: {e}")

        def accion_datos_ejemplo():
            try:
                crear_datos_ejemplo(self.cnx, self.cursor)
                messagebox.showinfo("OK", "Datos de ejemplo preparados correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error: {e}")

        def accion_limpiar():
            resp = messagebox.askyesno(
                "Confirmar",
                "Esto vaciará varias tablas del sistema.\n¿Seguro que quieres continuar?",
            )
            if not resp:
                return
            try:
                limpiarBaseDeDatos(self.cursor, self.cnx)
                messagebox.showinfo("OK", "Base de datos limpiada correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error: {e}")

        tk.Button(
            self.panel_central,
            text="Crear/verificar admin por defecto",
            command=accion_crear_admin,
        ).pack(padx=10, pady=5, anchor="nw")

        tk.Button(
            self.panel_central,
            text="Cargar datos de ejemplo",
            command=accion_datos_ejemplo,
        ).pack(padx=10, pady=5, anchor="nw")

        tk.Button(
            self.panel_central,
            text="Limpiar base de datos",
            command=accion_limpiar,
        ).pack(padx=10, pady=(15, 5), anchor="nw")
    






    # =========================================================
    # UTILIDADES GENERALES - Navegación
    # =========================================================
    def _clear_central(self):
        if self.panel_central is None:
            return
        for w in self.panel_central.winfo_children():
            w.destroy()

    def _show_text(self, texto):
        self._clear_central()
        txt = tk.Text(self.panel_central, wrap="word")
        txt.pack(fill="both", expand=True, padx=10, pady=10)
        txt.insert("end", texto)
        txt.config(state="disabled")


    def _set_contenido(self, texto):
        self._clear_central()
        self.lbl_contenido = tk.Label(self.panel_central, text=texto, font=("Arial", 12))
        self.lbl_contenido.pack(padx=10, pady=10, anchor="nw")

    # Navegación
    def _go_to_dashboard(self, usuario, rol):
        self.destroy()
        new_dash = DashboardView(self.master, self.cnx, self.cursor, usuario, rol)
        new_dash.pack(fill="both", expand=True)

    def _logout(self):
        if not messagebox.askyesno("Cerrar sesión", "¿Seguro que desea cerrar sesión?"):
            return

        # 1. destruir el dashboard actual
        self.destroy()

        # 2. volver al login, igual que en main.py
        from interfaces.login_view import LoginView

        def on_login_success(usuario_obj, rol):
            # destruir el login y crear un nuevo dashboard limpio
            login.destroy()
            nuevo_dashboard = DashboardView(self.master, self.cnx, self.cursor, usuario_obj, rol)
            nuevo_dashboard.pack(fill="both", expand=True)

        login = LoginView(self.master, self.cnx, self.cursor, on_login_success)
        login.pack(fill="both", expand=True)




